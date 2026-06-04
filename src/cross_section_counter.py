"""
Cross-Section Traffic Counter Module
截面流量统计模块

Features:
- Manage cross-section (count-line) configuration per entry/direction
- Count vehicle crossings using ByteTrack trajectories and segment intersect
- Export Excel reports
- Support YAML configuration persistence

Author: UAV Detection System
"""

from __future__ import annotations

import math
import uuid
from dataclasses import dataclass, field, asdict
from pathlib import Path
from typing import Dict, List, Set, Tuple, Any, Optional

import cv2
import numpy as np
import yaml


# ═══════════════════════════════════════════════════════════════
# Data Structures
# ═══════════════════════════════════════════════════════════════

@dataclass
class CrossSectionLine:
    """A single cross-section count line for traffic counting."""
    line_id: str                      # unique identifier
    entry_name: str                   # entry name (进口道)
    direction: str                    # travel direction (转向方向)
    x1: int; y1: int                  # start point (original video coordinates)
    x2: int; y2: int                  # end point (original video coordinates)
    color: Tuple[int, int, int] = (0, 255, 255)  # BGR display color
    is_drawn: bool = True

    def to_dict(self) -> dict:
        return {
            "line_id": self.line_id,
            "entry_name": self.entry_name,
            "direction": self.direction,
            "x1": self.x1, "y1": self.y1,
            "x2": self.x2, "y2": self.y2,
            "color": list(self.color),
            "is_drawn": self.is_drawn,
        }

    @classmethod
    def from_dict(cls, d: dict) -> "CrossSectionLine":
        return cls(
            line_id=d["line_id"],
            entry_name=d["entry_name"],
            direction=d["direction"],
            x1=d["x1"], y1=d["y1"],
            x2=d["x2"], y2=d["y2"],
            color=tuple(d.get("color", [0, 255, 255])),
            is_drawn=d.get("is_drawn", True),
        )


# ═══════════════════════════════════════════════════════════════
# Geometry Helpers
# ═══════════════════════════════════════════════════════════════

def _orientation(ax: float, ay: float, bx: float, by: float, cx: float, cy: float) -> int:
    """Compute orientation of three points: 0=collinear, 1=clockwise, 2=counter-clockwise."""
    val = (by - ay) * (cx - bx) - (bx - ax) * (cy - by)
    if abs(val) < 1e-9:
        return 0
    return 1 if val > 0 else 2


def _on_segment(ax: float, ay: float, bx: float, by: float, cx: float, cy: float) -> bool:
    """Check if point b lies on segment ac (assuming collinear)."""
    return min(ax, cx) <= bx <= max(ax, cx) and min(ay, cy) <= by <= max(ay, cy)


def segments_intersect(
    p1x: float, p1y: float, p2x: float, p2y: float,
    q1x: float, q1y: float, q2x: float, q2y: float,
) -> bool:
    """
    Determine whether two line segments intersect (including endpoints).
    Uses standard CCW (Counter-Clockwise) orientation test.
    """
    o1 = _orientation(p1x, p1y, p2x, p2y, q1x, q1y)
    o2 = _orientation(p1x, p1y, p2x, p2y, q2x, q2y)
    o3 = _orientation(q1x, q1y, q2x, q2y, p1x, p1y)
    o4 = _orientation(q1x, q1y, q2x, q2y, p2x, p2y)

    if o1 != o2 and o3 != o4:
        return True

    # Collinear cases
    if o1 == 0 and _on_segment(p1x, p1y, q1x, q1y, p2x, p2y):
        return True
    if o2 == 0 and _on_segment(p1x, p1y, q2x, q2y, p2x, p2y):
        return True
    if o3 == 0 and _on_segment(q1x, q1y, p1x, p1y, q2x, q2y):
        return True
    if o4 == 0 and _on_segment(q1x, q1y, p2x, p2y, q2x, q2y):
        return True

    return False


# ═══════════════════════════════════════════════════════════════
# Color Allocator
# ═══════════════════════════════════════════════════════════════

_PALETTE_BGR = [
    (231, 76, 60),    # red
    (52, 152, 219),   # blue
    (46, 204, 113),   # green
    (155, 89, 182),   # purple
    (241, 196, 15),   # yellow
    (230, 126, 34),   # orange
    (26, 188, 156),   # cyan
    (236, 112, 99),   # pink
]

_PALETTE_SHADE = [
    (192, 57, 43),    # dark red
    (41, 128, 185),   # dark blue
    (39, 174, 96),    # dark green
    (142, 68, 173),   # dark purple
    (243, 156, 18),   # dark yellow
    (211, 84, 0),     # dark orange
    (22, 160, 133),   # dark cyan
    (231, 76, 60),    # red
]


def get_entry_color(entry_index: int, line_index: int = 0) -> Tuple[int, int, int]:
    """Assign a color based on entry index and line index."""
    base = _PALETTE_BGR[entry_index % len(_PALETTE_BGR)]
    if line_index > 0:
        return _PALETTE_SHADE[(entry_index + line_index) % len(_PALETTE_SHADE)]
    return base


# ═══════════════════════════════════════════════════════════════
# Core Counter
# ═══════════════════════════════════════════════════════════════

class CrossSectionCounter:
    """
    Cross-section traffic counter.

    Usage:
        counter = CrossSectionCounter(cs_lines)
        for each frame:
            for each tracked object:
                counter.update(tracker_id, class_id, bbox_xyxy)
        stats = counter.get_stats()
    """

    def __init__(self, cs_lines: List[CrossSectionLine]):
        self.cs_lines: List[CrossSectionLine] = [
            cl for cl in cs_lines if cl.is_drawn
        ]
        # tracker_id → previous position (x, y)
        self._prev_positions: Dict[int, Tuple[float, float]] = {}
        # counted set: (tracker_id, line_id)
        self._counted: Set[Tuple[int, str]] = set()
        # stats: line_id → {class_id: count}
        self._stats: Dict[str, Dict[int, int]] = {
            cl.line_id: {} for cl in self.cs_lines
        }
        # entry-aggregated stats: entry_name → {direction: {class_id: count}}
        self._entry_stats: Dict[str, Dict[str, Dict[int, int]]] = {}

    def update(self, tracker_id: int, class_id: int, bbox_xyxy: np.ndarray):
        """
        Update traffic count for a single tracked object.

        Args:
            tracker_id: ByteTrack-assigned tracking ID
            class_id: object class ID
            bbox_xyxy: [x1, y1, x2, y2] detection bounding box
        """
        x1, y1, x2, y2 = bbox_xyxy
        # Use bottom-center of bbox as vehicle reference point (aerial top-down view)
        cx = (x1 + x2) / 2.0
        cy = y2  # bottom

        prev = self._prev_positions.get(tracker_id)
        if prev is not None:
            px, py = prev
            for cl in self.cs_lines:
                key = (tracker_id, cl.line_id)
                if key in self._counted:
                    continue
                # Check if movement segment (px,py)→(cx,cy) intersects the cross-section line
                if segments_intersect(px, py, cx, cy, cl.x1, cl.y1, cl.x2, cl.y2):
                    self._counted.add(key)
                    # Accumulate stats
                    self._stats[cl.line_id][class_id] = self._stats[cl.line_id].get(class_id, 0) + 1
                    # Sync entry_stats
                    entry = self._entry_stats.setdefault(cl.entry_name, {})
                    direction_map = entry.setdefault(cl.direction, {})
                    direction_map[class_id] = direction_map.get(class_id, 0) + 1

        self._prev_positions[tracker_id] = (cx, cy)

    def remove_tracker(self, tracker_id: int):
        """Clean up memory when a target disappears (optional)."""
        self._prev_positions.pop(tracker_id, None)

    def get_stats(self) -> Dict[str, Any]:
        """
        Get current counting statistics.

        Returns:
            {
                "by_line": {line_id: {class_id: count, ...}, ...},
                "by_entry": {
                    entry_name: {
                        direction: {class_id: count, ...},
                        ...
                    },
                    ...
                },
                "lines": [CrossSectionLine, ...]
            }
        """
        return {
            "by_line": {lid: dict(cid_counts) for lid, cid_counts in self._stats.items()},
            "by_entry": {
                entry: {dir_: dict(c) for dir_, c in dirs.items()}
                for entry, dirs in self._entry_stats.items()
            },
            "lines": self.cs_lines,
        }

    def draw_on_frame(self, frame: np.ndarray, display_scale: float = 1.0) -> np.ndarray:
        """
        Draw cross-section lines and live counts on a frame.

        Args:
            frame: BGR image
            display_scale: scaling factor for line thickness / font size
        """
        from src.zh_draw import put_chinese_text, get_chinese_text_size

        annotated = frame.copy()
        h, w = annotated.shape[:2]
        thickness = max(1, int(3 * display_scale))
        font_size = max(10, int(14 * display_scale))
        radius = max(3, int(5 * display_scale))

        for cl in self.cs_lines:
            # Draw line
            cv2.line(annotated, (cl.x1, cl.y1), (cl.x2, cl.y2), cl.color, thickness)
            # Endpoints
            cv2.circle(annotated, (cl.x1, cl.y1), radius, (0, 255, 0), -1)
            cv2.circle(annotated, (cl.x2, cl.y2), radius, (0, 0, 255), -1)

            # Label at midpoint
            mx = (cl.x1 + cl.x2) // 2
            my = (cl.y1 + cl.y2) // 2
            total = sum(self._stats.get(cl.line_id, {}).values())
            label = f"{cl.entry_name}→{cl.direction}: {total}"
            tw, th = get_chinese_text_size(label, font_size=font_size)
            tx = max(0, min(mx - tw // 2, w - tw - 2))
            ty = max(th + 4, my - 8)
            # Semi-transparent background
            overlay = annotated.copy()
            cv2.rectangle(overlay, (tx - 4, ty - th - 4), (tx + tw + 4, ty + 4), (0, 0, 0), -1)
            cv2.addWeighted(overlay, 0.6, annotated, 0.4, 0, annotated)
            put_chinese_text(annotated, label, (tx, ty), font_size=font_size, color=(255, 255, 255))

        return annotated


# ═══════════════════════════════════════════════════════════════
# Excel Export
# ═══════════════════════════════════════════════════════════════

def export_cross_section_excel(
    stats: Dict[str, Any],
    class_labels_zh: Dict[int, str],
    video_name: str,
    duration_str: str,
    detect_time_str: str = "",
    output_path: Path = None,
) -> Path:
    """
    Export cross-section traffic statistics to an Excel file.

    Args:
        stats: result from CrossSectionCounter.get_stats()
        class_labels_zh: {class_id: Chinese label}
        video_name: video filename
        duration_str: video duration string, e.g. "00:05:00"
        detect_time_str: actual detection time, e.g. "00:02:15"
        output_path: output .xlsx path

    Returns:
        actual written file path
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError as _e:
        raise ImportError("Excel export requires openpyxl, install: pip install openpyxl") from _e

    wb = Workbook()
    ws = wb.active
    ws.title = "Cross-Section Traffic"

    # Styles
    header_font = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Title rows
    ws.merge_cells("A1:D1")
    ws["A1"] = f"Video: {video_name}"
    ws["A1"].font = Font(name="Microsoft YaHei", bold=True, size=12)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("A2:D2")
    ws["A2"] = f"Duration: {duration_str}"
    ws["A2"].font = Font(name="Microsoft YaHei", size=11)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    if detect_time_str:
        ws.merge_cells("A3:D3")
        ws["A3"] = f"Detection Time: {detect_time_str}"
        ws["A3"].font = Font(name="Microsoft YaHei", size=11, color="4A90D9")
        ws["A3"].alignment = Alignment(horizontal="left", vertical="center")

    # Collect all class_ids
    all_class_ids: Set[int] = set()
    for entry_dirs in stats.get("by_entry", {}).values():
        for dir_counts in entry_dirs.values():
            all_class_ids.update(dir_counts.keys())
    sorted_class_ids = sorted(all_class_ids)

    # Header row
    header_row = 5 if detect_time_str else 4
    headers = ["Entry", "Direction"] + [class_labels_zh.get(cid, f"Class {cid}") for cid in sorted_class_ids] + ["Total"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data rows
    data_row = header_row + 1
    by_entry = stats.get("by_entry", {})
    for entry_name in sorted(by_entry.keys()):
        directions = by_entry[entry_name]
        for direction in sorted(directions.keys()):
            counts = directions[direction]
            row_values = [entry_name, direction]
            row_total = 0
            for cid in sorted_class_ids:
                c = counts.get(cid, 0)
                row_values.append(c)
                row_total += c
            row_values.append(row_total)

            for col_idx, val in enumerate(row_values, 1):
                cell = ws.cell(row=data_row, column=col_idx, value=val)
                cell.alignment = Alignment(horizontal="center" if col_idx > 2 else "left", vertical="center")
                cell.border = thin_border
                if col_idx <= 2:
                    cell.font = Font(name="Microsoft YaHei", size=10)
                else:
                    cell.font = Font(name="Consolas", size=10)
            data_row += 1

    # Column widths
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 14
    for i, cid in enumerate(sorted_class_ids, 3):
        name = class_labels_zh.get(cid, f"Class {cid}")
        ws.column_dimensions[get_column_letter(i)].width = max(10, len(name) * 2 + 2)
    ws.column_dimensions[get_column_letter(len(headers))].width = 10

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    return output_path


# ═══════════════════════════════════════════════════════════════
# Configuration Persistence
# ═══════════════════════════════════════════════════════════════

_CS_CONFIG_DIR = "cross_section_lines"


def _config_dir(project_root: Path) -> Path:
    d = project_root / "configs" / _CS_CONFIG_DIR
    d.mkdir(parents=True, exist_ok=True)
    return d


def save_cs_lines_config(
    cs_lines: List[CrossSectionLine],
    video_path: Path,
    project_root: Path,
) -> Path:
    """
    Save cross-section line config to YAML, keyed by video filename.

    Returns:
        saved file path
    """
    config_dir = _config_dir(project_root)
    safe_name = video_path.stem.replace(" ", "_").replace(".", "_")
    config_path = config_dir / f"{safe_name}.yaml"

    data = {
        "video_name": video_path.name,
        "video_path": str(video_path),
        "entries": {},
    }
    for cl in cs_lines:
        entry = data["entries"].setdefault(cl.entry_name, {"lines": []})
        entry["lines"].append(cl.to_dict())

    with open(config_path, "w", encoding="utf-8") as f:
        yaml.dump(data, f, allow_unicode=True, default_flow_style=False, sort_keys=False)

    return config_path


def load_cs_lines_config(
    video_path: Path,
    project_root: Path,
) -> List[CrossSectionLine]:
    """
    Load cross-section line config from YAML.

    Returns:
        list of CrossSectionLine; empty list if not found
    """
    config_dir = _config_dir(project_root)
    safe_name = video_path.stem.replace(" ", "_").replace(".", "_")
    config_path = config_dir / f"{safe_name}.yaml"

    if not config_path.exists():
        return []

    try:
        with open(config_path, "r", encoding="utf-8") as f:
            data = yaml.safe_load(f) or {}
    except Exception:
        return []

    lines: List[CrossSectionLine] = []
    for entry_name, entry_data in data.get("entries", {}).items():
        for ld in entry_data.get("lines", []):
            ld["entry_name"] = entry_name
            lines.append(CrossSectionLine.from_dict(ld))
    return lines
