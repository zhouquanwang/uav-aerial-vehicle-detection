"""
无人机航拍视频检测系统 - 专业UI界面 v4.0 (优化版)
UAV Aerial Vehicle Detection System - Optimized UI

优化内容：
- 线程安全：分离视频读取线程与UI线程，使用命令队列通信
- 性能优化：双缓冲帧缓存、预渲染、流式检测处理
- 交互优化：时间轴拖动防抖、精确帧率控制
- 内存优化：逐帧流式处理，避免一次性加载所有帧

Author: UAV Detection System
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog
from pathlib import Path
from typing import Optional, Callable, Tuple, List, Dict
import cv2
from PIL import Image, ImageTk
import threading
import time
import numpy as np
import yaml
import sys
import math
import uuid
from queue import Queue, Empty, Full
from dataclasses import dataclass
from enum import Enum, auto

# 项目路径（兼容 PyInstaller 打包）
if getattr(sys, "frozen", False):
    PROJECT_ROOT = Path(sys.executable).parent
else:
    PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))

from src.class_labels import MODEL_CLASS_CONFIGS, label_for_class, parse_class_labels_zh
from src.device_util import device_status_line, resolve_device
from src.zh_draw import (create_chinese_label_annotator, resolve_chinese_font_path,
                          put_chinese_text, get_chinese_text_size)
from src.cross_section_counter import (
    CrossSectionLine, CrossSectionCounter,
    export_cross_section_excel,
    save_cs_lines_config, load_cs_lines_config,
    get_entry_color,
)
from src.auto_traffic_counter import AutoTrafficCounter


# ═══════════════════════════════════════════════════════════════
# 深色科技主题设计系统 (Dark Tech Design System)
# ═══════════════════════════════════════════════════════════════
class DARK:
    """深色科技主题配色"""
    BG          = "#0a1628"   # 最深背景
    PANEL       = "#0f1f3a"   # 面板背景
    CARD        = "#142848"   # 卡片/输入区背景
    SURFACE     = "#192d4d"   # 悬浮层背景
    HEADER      = "#0d1a30"   # 标题栏背景
    BORDER      = "#1a3050"   # 分隔线/边框
    BORDER_LIGHT= "#234060"   # 浅色边框

    # 文字
    TEXT        = "#e0e6f0"   # 主文字
    TEXT_SUB    = "#8a95a5"   # 次要文字
    TEXT_MUTED  = "#556070"   # 辅助文字
    TEXT_HINT   = "#3a4a5a"   # 提示文字

    # 强调色
    ACCENT      = "#00d4aa"   # 主强调-青色
    ACCENT_HOVER= "#00f0c0"   # 悬停
    BLUE        = "#3b82f6"   # 蓝色
    PURPLE      = "#8b5cf6"   # 紫色
    GREEN       = "#10b981"   # 绿色
    ORANGE      = "#f59e0b"   # 橙色
    RED         = "#ef4444"   # 红色
    YELLOW      = "#facc15"   # 黄色

    # 功能色
    BTN_PRIMARY  = "#00d4aa"  # 主按钮
    BTN_SUCCESS  = "#10b981"  # 成功按钮
    BTN_DANGER   = "#ef4444"  # 危险按钮
    BTN_DISABLED = "#2a3a50"  # 禁用按钮
    BTN_DISABLED_TEXT = "#4a5a6a"  # 禁用按钮文字

# 别名简写
C = DARK

# 字体系统
FONT_TITLE    = ("Microsoft YaHei", 16, "bold")
FONT_HEADER   = ("Microsoft YaHei", 12, "bold")
FONT_SECTION  = ("Microsoft YaHei", 10, "bold")
FONT_BODY     = ("Microsoft YaHei", 9)
FONT_SMALL    = ("Microsoft YaHei", 8)
FONT_TINY     = ("Microsoft YaHei", 7)
FONT_MONO     = ("Consolas", 9, "bold")
FONT_MONO_SM  = ("Consolas", 9)
FONT_TIME     = ("Consolas", 14, "bold")


# ═══════════════════════════════════════════════════════════════
# 线程安全数据结构
# ═══════════════════════════════════════════════════════════════

class PlayerCommand(Enum):
    PLAY = auto()
    PAUSE = auto()
    STOP = auto()
    SEEK = auto()
    SPEED = auto()
    EXIT = auto()


@dataclass
class Command:
    type: PlayerCommand
    value: Optional[int] = None


class FrameBuffer:
    """
    线程安全帧缓冲区 - 存储原始帧(ndarray)，由UI线程负责渲染与叠加
    """
    def __init__(self, maxsize: int = 3, max_cache: int = 10):
        self._queue: Queue = Queue(maxsize=maxsize)
        self._cache: Dict[int, np.ndarray] = {}
        self._max_cache = max_cache
        self._lock = threading.Lock()

    def set_display_size(self, width: int, height: int):
        with self._lock:
            self._cache.clear()

    def put(self, frame_num: int, frame: np.ndarray) -> bool:
        """放入原始帧到缓冲区，返回是否成功"""
        # 缓存原始帧副本
        with self._lock:
            if len(self._cache) >= self._max_cache:
                if self._cache:
                    oldest = min(self._cache.keys())
                    del self._cache[oldest]
            self._cache[frame_num] = frame.copy()

        # 放入队列（满时丢弃旧帧）
        try:
            self._queue.put_nowait((frame_num, frame.copy()))
            return True
        except Full:
            try:
                self._queue.get_nowait()
                self._queue.put_nowait((frame_num, frame.copy()))
                return True
            except (Empty, Full):
                return False

    def get(self) -> Optional[Tuple[int, np.ndarray]]:
        """从队列获取最新帧（ndarray）"""
        try:
            return self._queue.get_nowait()
        except Empty:
            return None

    def get_cached(self, frame_num: int) -> Optional[np.ndarray]:
        """从缓存获取指定帧"""
        with self._lock:
            frame = self._cache.get(frame_num)
            return frame.copy() if frame is not None else None

    def clear(self):
        """清空缓冲区和缓存"""
        with self._lock:
            self._cache.clear()
        while not self._queue.empty():
            try:
                self._queue.get_nowait()
            except Empty:
                break


# ═══════════════════════════════════════════════════════════════
# 视频播放线程
# ═══════════════════════════════════════════════════════════════

class VideoPlayerThread(threading.Thread):
    """
    独立视频读取线程 - 只负责解码，不操作任何UI
    通过 command_queue 接收命令，通过 frame_buffer 输出帧
    """
    
    def __init__(self, video_path: Path, frame_buffer: FrameBuffer, 
                 command_queue: Queue, fps: float = 30.0,
                 start_frame: int = 0, end_frame: int | None = None):
        super().__init__(daemon=True, name="VideoPlayer")
        self.video_path = video_path
        self.frame_buffer = frame_buffer
        self.command_queue = command_queue
        self.target_fps = fps
        self._start_frame = start_frame
        self._end_frame = end_frame  # None 表示无限制
        self._running = threading.Event()
        self._paused = threading.Event()
        self._paused.set()  # 初始暂停
        self._speed = 1.0
        self._current_frame = 0
        self._total_frames = 0
        self._cap: Optional[cv2.VideoCapture] = None
        self._seek_requested = threading.Event()
        self._seek_target = 0
        
    def run(self):
        """主循环"""
        self._cap = cv2.VideoCapture(str(self.video_path))
        if not self._cap.isOpened():
            print(f"[Player] 无法打开视频: {self.video_path}")
            return
            
        self._total_frames = int(self._cap.get(cv2.CAP_PROP_FRAME_COUNT))
        if self._end_frame is None:
            self._end_frame = self._total_frames - 1
        # 确保起始位置在有效范围内
        self._cap.set(cv2.CAP_PROP_POS_FRAMES, self._start_frame)
        self._current_frame = self._start_frame
        actual_fps = self._cap.get(cv2.CAP_PROP_FPS) or self.target_fps
        self._running.set()
        
        frame_duration = 1.0 / actual_fps
        next_frame_time = time.time()
        
        while self._running.is_set():
            # 处理命令队列（非阻塞）
            self._process_commands()
            
            # 处理seek请求（优先于暂停，暂停状态下也能跳转并显示帧）
            if self._seek_requested.is_set():
                self._cap.set(cv2.CAP_PROP_POS_FRAMES, self._seek_target)
                self._current_frame = self._seek_target
                self._seek_requested.clear()
                next_frame_time = time.time()
                # 读取一帧放入缓冲区，让UI能显示跳转后的画面
                ret, frame = self._cap.read()
                if ret:
                    self.frame_buffer.put(self._current_frame, frame)
                    self._current_frame = int(self._cap.get(cv2.CAP_PROP_POS_FRAMES)) - 1
                continue
                
            # 如果暂停，等待
            if self._paused.is_set():
                time.sleep(0.05)
                next_frame_time = time.time()
                continue
                
            # 帧率控制
            current_time = time.time()
            if current_time < next_frame_time:
                sleep_time = next_frame_time - current_time
                if sleep_time > 0.005:
                    time.sleep(sleep_time)
                continue
                
            # 读取帧
            ret, frame = self._cap.read()
            if not ret:
                # 到达末尾，回到起帧位置
                self._cap.set(cv2.CAP_PROP_POS_FRAMES, self._start_frame)
                self._current_frame = self._start_frame
                next_frame_time = time.time()
                continue

            # 检查是否超出终点范围，超出则回跳
            if self._current_frame >= self._end_frame:
                self._cap.set(cv2.CAP_PROP_POS_FRAMES, self._start_frame)
                self._current_frame = self._start_frame
                next_frame_time = time.time()
                continue
                
            self._current_frame = int(self._cap.get(cv2.CAP_PROP_POS_FRAMES)) - 1
            
            # 放入缓冲区
            self.frame_buffer.put(self._current_frame, frame)
            
            # 计算下一帧时间
            adjusted_duration = frame_duration / self._speed
            next_frame_time += adjusted_duration
            
            # 如果落后太多，重置时间
            if time.time() > next_frame_time + adjusted_duration * 2:
                next_frame_time = time.time()
                
        # 清理
        if self._cap:
            self._cap.release()
        print("[Player] 线程已退出")
        
    def _process_commands(self):
        """处理命令队列"""
        while True:
            try:
                cmd = self.command_queue.get_nowait()
            except Empty:
                break
                
            if cmd.type == PlayerCommand.PLAY:
                self._paused.clear()
                print(f"[Player] 播放，速度: {self._speed}x")
            elif cmd.type == PlayerCommand.PAUSE:
                self._paused.set()
                print("[Player] 暂停")
            elif cmd.type == PlayerCommand.STOP:
                self._paused.set()
                self._seek_target = self._start_frame
                self._seek_requested.set()
            elif cmd.type == PlayerCommand.SEEK and cmd.value is not None:
                # 限制在起终点范围内
                _t = max(0, min(cmd.value, self._total_frames - 1))
                _t = max(self._start_frame, min(_t, self._end_frame))
                self._seek_target = _t
                self._seek_requested.set()
            elif cmd.type == PlayerCommand.SPEED and cmd.value is not None:
                self._speed = max(0.25, min(cmd.value, 4.0))
                print(f"[Player] 速度设置为: {self._speed}x")
            elif cmd.type == PlayerCommand.EXIT:
                self._running.clear()
                
    def get_current_frame(self) -> int:
        """获取当前帧号（线程安全）"""
        return self._current_frame
        
    def get_total_frames(self) -> int:
        return self._total_frames
        
    def is_playing(self) -> bool:
        return not self._paused.is_set()
        
    def stop(self):
        """请求停止线程"""
        self._running.clear()


# ═══════════════════════════════════════════════════════════════
# 时间轴滑块组件
# ═══════════════════════════════════════════════════════════════

class TimelineSlider(tk.Canvas):
    """
    专业时间轴滑块 - 防抖优化版
    """
    
    def __init__(
        self,
        parent,
        width: int = 800,
        height: int = 90,
        min_val: int = 0,
        max_val: int = 100,
        start_value: int = 0,
        end_value: int = 100,
        command: Optional[Callable] = None,
        **kwargs
    ):
        super().__init__(
            parent,
            width=width,
            height=height,
            bg=C.PANEL,
            highlightthickness=0,
            **kwargs
        )

        self.min_val = min_val
        self.max_val = max_val
        self.start_value = start_value
        self.end_value = end_value
        self.command = command

        # 尺寸参数
        self.padx = 30
        self.track_top = 18
        self.track_bottom = 30
        self.playhead_x = 0

        self.dragging: Optional[str] = None
        self._drag_debounce_id: Optional[str] = None
        self._pending_seek_frame: Optional[int] = None

        # 绑定事件
        self.bind("<Button-1>", self._on_click)
        self.bind("<B1-Motion>", self._on_drag)
        self.bind("<ButtonRelease-1>", self._on_release)
        self.bind("<Configure>", lambda e: self._schedule_draw())

        self._draw_id = None
        self._draw()

    def _schedule_draw(self):
        if self._draw_id:
            self.after_cancel(self._draw_id)
        self._draw_id = self.after(10, self._redraw)

    def _redraw(self):
        self.delete("all")
        self._draw()

    def _value_to_x(self, value: int) -> float:
        if self.max_val == self.min_val:
            return self.padx
        ratio = (value - self.min_val) / (self.max_val - self.min_val)
        usable_width = self.winfo_width() - 2 * self.padx
        return self.padx + ratio * usable_width

    def _x_to_value(self, x: float) -> int:
        usable_width = self.winfo_width() - 2 * self.padx
        ratio = (x - self.padx) / usable_width
        ratio = max(0, min(1, ratio))
        return int(self.min_val + ratio * (self.max_val - self.min_val))

    def _draw(self):
        """绘制专业时间轴"""
        width = self.winfo_width()
        height = self.winfo_height()
        if width <= 1:
            width = 800

        # 背景
        self.create_rectangle(0, 0, width, height, fill=C.PANEL, outline="")

        # 计算位置
        sx = self._value_to_x(self.start_value)
        ex = self._value_to_x(self.end_value)

        # 时间刻度
        self._draw_time_ticks(width)

        # 轨道背景
        self.create_rectangle(
            self.padx, self.track_top,
            width - self.padx, self.track_bottom,
            fill=C.BORDER, outline=""
        )

        # 选中范围
        self.create_rectangle(
            sx, self.track_top,
            ex, self.track_bottom,
            fill=C.ACCENT, outline="", stipple="gray25"
        )

        # 播放头
        if self.playhead_x >= self.padx and self.playhead_x <= width - self.padx:
            self.create_polygon(
                self.playhead_x - 5, self.track_top - 4,
                self.playhead_x + 5, self.track_top - 4,
                self.playhead_x, self.track_top + 1,
                fill=C.TEXT, outline=C.TEXT
            )
            self.create_line(
                self.playhead_x, self.track_top,
                self.playhead_x, self.track_bottom,
                fill=C.ACCENT, width=2
            )

        # 滑块手柄
        self._draw_range_handle(sx, self.track_top, self.track_bottom, C.GREEN, "start")
        self._draw_range_handle(ex, self.track_top, self.track_bottom, C.ORANGE, "end")

        # 范围信息
        info_y = self.track_bottom + 12
        self.create_text(
            self.padx, info_y,
            text=f"起始: {self.start_value}帧 ({self._format_frame_time(self.start_value)})",
            fill=C.GREEN, font=("Consolas", 7), anchor="w"
        )
        self.create_text(
            width - self.padx, info_y,
            text=f"结束: {self.end_value}帧 ({self._format_frame_time(self.end_value)})",
            fill=C.ORANGE, font=("Consolas", 7), anchor="e"
        )

    def _draw_time_ticks(self, width: int):
        """绘制时间刻度"""
        total_frames = self.max_val - self.min_val
        if total_frames <= 0:
            return

        fps = 30
        total_seconds = total_frames / fps

        if total_seconds <= 60:
            tick_interval_sec = 5
        elif total_seconds <= 300:
            tick_interval_sec = 15
        elif total_seconds <= 900:
            tick_interval_sec = 30
        else:
            tick_interval_sec = 60

        tick_interval_frames = int(tick_interval_sec * fps)
        num_ticks = max(1, total_frames // tick_interval_frames)
        usable_width = width - 2 * self.padx

        for i in range(num_ticks + 1):
            x = self.padx + usable_width * i / num_ticks
            if x > width - self.padx:
                break

            self.create_line(x, self.track_top - 4, x, self.track_top - 1, fill=C.TEXT_MUTED, width=1)
            frame_num = int(i * tick_interval_frames)
            if frame_num <= self.max_val:
                time_str = self._format_frame_time(frame_num)
                self.create_text(x, self.track_top - 7, text=time_str, fill=C.TEXT_SUB, font=("Consolas", 7))

    def _draw_range_handle(self, x: float, top: int, bottom: int, color: str, handle_type: str):
        """绘制范围手柄"""
        h = bottom - top
        y = top + h // 2
        r = 5
        self.create_oval(x - r, y - r * 2, x + r, y + r * 2,
                        fill=color, outline=C.BG, width=1.5)
        self.create_line(x - 2, y, x + 2, y, fill=C.BG, width=1.5)

    def _format_frame_time(self, frame: int) -> str:
        total_sec = frame / 30
        m, s = divmod(int(total_sec), 60)
        h, m = divmod(m, 60)
        if h > 0:
            return f"{h}:{m:02d}:{s:02d}"
        return f"{m:02d}:{s:02d}"

    def _on_click(self, event):
        """处理点击"""
        sx = self._value_to_x(self.start_value)
        ex = self._value_to_x(self.end_value)

        # 检查滑块（单击立即跳转到对应位置预览，起终点值仅通过拖拽修改）
        for name, pos in [("start", sx), ("end", ex)]:
            if abs(event.x - pos) < 12 and abs(event.y - (self.track_top + self.track_bottom) / 2) < 15:
                self.dragging = name
                self._seek_to_position(event.x)  # 单击立即跳转预览
                return

        # 点击轨道跳转（限制在起终点范围内）
        if self.padx <= event.x <= self.winfo_width() - self.padx:
            if self.track_top - 3 <= event.y <= self.track_bottom + 3:
                self._seek_to_position(event.x)
                return

    def _on_drag(self, event):
        """处理拖拽 - 带防抖"""
        if not self.dragging:
            return

        if self.dragging == "playhead":
            self._seek_to_position(event.x)
            return

        if self.dragging == "start":
            new_val = self._x_to_value(event.x)
            new_val = max(self.min_val, min(new_val, self.end_value - 1))
            if new_val != self.start_value:
                self.start_value = new_val
                self._redraw()
                # 防抖：延迟通知
                self._pending_seek_frame = new_val
                self._debounce_notify("range", self.start_value, self.end_value)
                # 记录最后拖动的手柄
                if self.command and hasattr(self.command, "__self__"):
                    try:
                        self.command.__self__._last_dragged_handle = "start"
                    except:
                        pass

        elif self.dragging == "end":
            new_val = self._x_to_value(event.x)
            new_val = max(self.start_value + 1, min(new_val, self.max_val))
            if new_val != self.end_value:
                self.end_value = new_val
                self._redraw()
                self._pending_seek_frame = new_val
                self._debounce_notify("range", self.start_value, self.end_value)
                # 记录最后拖动的手柄
                if self.command and hasattr(self.command, "__self__"):
                    try:
                        self.command.__self__._last_dragged_handle = "end"
                    except:
                        pass

    def _debounce_notify(self, action: str, *args):
        """防抖通知 - 拖动停止后才实际跳转"""
        if self._drag_debounce_id:
            self.after_cancel(self._drag_debounce_id)
        # 拖动时只更新UI，300ms后才通知外部跳转
        self._drag_debounce_id = self.after(300, self._commit_notify, action, args)

    def _commit_notify(self, action: str, args: tuple):
        """提交最终通知"""
        if self.command:
            # 拖动结束时才通知seek
            self.command(action, *args, seek_frame=True)

    def _seek_to_position(self, x: float):
        """跳转到指定位置（限制在起终点范围内）"""
        frame = self._x_to_value(x)
        # 限制在起终点范围内
        frame = max(self.start_value, min(frame, self.end_value))
        self.playhead_x = self._value_to_x(frame)
        self._redraw()
        if self.command:
            self.command("seek", frame)

    def _on_release(self, event):
        """释放拖拽"""
        if self.dragging and self.dragging != "playhead":
            # 记录最后拖动的手柄
            if self.command and hasattr(self.command, "__self__"):
                try:
                    self.command.__self__._last_dragged_handle = self.dragging
                except:
                    pass
            # 确保最终通知发出
            if self._drag_debounce_id:
                self.after_cancel(self._drag_debounce_id)
                self._drag_debounce_id = None
            if self.command:
                self.command("range", self.start_value, self.end_value, seek_frame=True)
        self.dragging = None

    def set_playhead(self, frame: int):
        """设置播放头位置"""
        self.playhead_x = self._value_to_x(frame)
        self._redraw()

    def set_range(self, start: int, end: int):
        self.start_value = start
        self.end_value = end
        self._redraw()

    def configure(self, min_val: int = None, max_val: int = None,
                 start_value: int = None, end_value: int = None):
        if min_val is not None:
            self.min_val = min_val
        if max_val is not None:
            self.max_val = max_val
        if start_value is not None:
            self.start_value = start_value
        if end_value is not None:
            self.end_value = end_value
        self._redraw()


# ═══════════════════════════════════════════════════════════════
# 参数滑块组件
# ═══════════════════════════════════════════════════════════════

class IntScale(tk.Frame):
    """整数型滑块（支持拖动和直接输入）- 深色科技主题"""
    def __init__(self, parent, label, min_val, max_val, default_val, command=None, **kwargs):
        super().__init__(parent, **kwargs)
        self.configure(bg=C.CARD)
        self.min_val = min_val
        self.max_val = max_val
        self.command = command

        self.label_text = tk.Label(self, text=label, bg=C.CARD, fg=C.TEXT_SUB,
                                   font=FONT_SMALL, width=6, anchor=tk.W)
        self.label_text.pack(side=tk.LEFT, padx=(0, 2))

        self.entry_var = tk.StringVar(value=str(default_val))
        self.value_entry = tk.Entry(self, textvariable=self.entry_var,
                                    bg=C.SURFACE, fg=C.ACCENT,
                                    font=FONT_MONO, width=3, justify=tk.CENTER,
                                    relief=tk.FLAT, bd=0, insertbackground=C.ACCENT)
        self.value_entry.pack(side=tk.LEFT, padx=(0, 2))
        self.value_entry.bind("<Return>", self._on_entry_changed)
        self.value_entry.bind("<FocusOut>", self._on_entry_changed)

        self.scale = ttk.Scale(self, from_=min_val, to=max_val, orient=tk.HORIZONTAL,
                              variable=tk.IntVar(value=default_val), command=self._on_scale_changed)
        self._scale_var = self.scale.cget("variable")
        self.scale.configure(variable=self._scale_var)
        self.scale.pack(side=tk.LEFT, fill=tk.X, expand=True)

    def _on_scale_changed(self, val):
        int_val = int(float(val))
        self.entry_var.set(str(int_val))
        if self.command:
            self.command(int_val)

    def _on_entry_changed(self, event=None):
        try:
            raw = self.entry_var.get().strip()
            int_val = int(raw)
            int_val = max(self.min_val, min(self.max_val, int_val))
            self.entry_var.set(str(int_val))
            self.scale.set(int_val)
            if self.command:
                self.command(int_val)
        except ValueError:
            self.entry_var.set(str(self.get()))

    def get(self):
        try:
            return int(self.entry_var.get().strip())
        except ValueError:
            return int(float(self.scale.get()))

    def set(self, val):
        self.entry_var.set(str(val))
        self.scale.set(val)


class FloatScale(tk.Frame):
    """浮点型滑块（支持拖动和直接输入）- 深色科技主题"""
    def __init__(self, parent, label, min_val, max_val, default_val, step=0.05, decimals=2, command=None, **kwargs):
        super().__init__(parent, **kwargs)
        self.configure(bg=C.CARD)
        self.decimals = decimals
        self.min_val = min_val
        self.max_val = max_val
        self.command = command

        self.label_text = tk.Label(self, text=label, bg=C.CARD, fg=C.TEXT_SUB,
                                   font=FONT_SMALL, width=6, anchor=tk.W)
        self.label_text.pack(side=tk.LEFT, padx=(0, 2))

        self.entry_var = tk.StringVar(value=f"{default_val:.{decimals}f}")
        self.value_entry = tk.Entry(self, textvariable=self.entry_var,
                                    bg=C.SURFACE, fg=C.ACCENT,
                                    font=FONT_MONO, width=4, justify=tk.CENTER,
                                    relief=tk.FLAT, bd=0, insertbackground=C.ACCENT)
        self.value_entry.pack(side=tk.LEFT, padx=(0, 2))
        self.value_entry.bind("<Return>", self._on_entry_changed)
        self.value_entry.bind("<FocusOut>", self._on_entry_changed)

        self.scale = ttk.Scale(self, from_=min_val, to=max_val, orient=tk.HORIZONTAL,
                              variable=tk.DoubleVar(value=default_val), command=self._on_scale_changed)
        self._scale_var = self.scale.cget("variable")
        self.scale.configure(variable=self._scale_var)
        self.scale.pack(side=tk.LEFT, fill=tk.X, expand=True)

    def _on_scale_changed(self, val):
        float_val = float(val)
        self.entry_var.set(f"{float_val:.{self.decimals}f}")
        if self.command:
            self.command(float_val)

    def _on_entry_changed(self, event=None):
        try:
            raw = self.entry_var.get().strip()
            float_val = float(raw)
            float_val = max(self.min_val, min(self.max_val, float_val))
            self.entry_var.set(f"{float_val:.{self.decimals}f}")
            self.scale.set(float_val)
            if self.command:
                self.command(float_val)
        except ValueError:
            self.entry_var.set(f"{self.get():.{self.decimals}f}")

    def get(self):
        try:
            return float(self.entry_var.get().strip())
        except ValueError:
            return float(self.scale.get())

    def set(self, val):
        self.entry_var.set(f"{val:.{self.decimals}f}")
        self.scale.set(val)


# ═══════════════════════════════════════════════════════════════
# 主界面类
# ═══════════════════════════════════════════════════════════════

class UAVDetectionUI:
    """无人机航拍视频检测系统主界面 - 优化版"""

    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("无人机航拍视频检测系统")
        self.root.geometry("1600x980")
        self.root.minsize(1400, 900)
        self.root.configure(bg=C.BG)

        # ── ttk 深色主题 ──
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("TScrollbar",
            background=C.CARD, troughcolor=C.BG, arrowsize=12,
            arrowcolor=C.TEXT_SUB, bordercolor=C.BG, lightcolor=C.CARD, darkcolor=C.CARD)
        style.configure("Treeview",
            background=C.BG, foreground=C.TEXT, fieldbackground=C.BG,
            borderwidth=0, font=FONT_BODY)
        style.configure("Treeview.Heading",
            background=C.CARD, foreground=C.ACCENT, font=("Microsoft YaHei",9,"bold"),
            borderwidth=0)
        style.map("Treeview", background=[("selected", C.ACCENT)], foreground=[("selected", C.BG)])
        style.configure("TCombobox",
            fieldbackground=C.SURFACE, background=C.CARD, foreground=C.TEXT,
            arrowcolor=C.ACCENT, selectbackground=C.CARD, selectforeground=C.TEXT)
        style.map("TCombobox", fieldbackground=[("readonly", C.SURFACE)])
        style.configure("TScale", background=C.CARD, troughcolor=C.BORDER, sliderlength=14)

        # ═══════ 核心变量 ═══════
        self.video_path: Optional[Path] = None
        
        # 线程安全组件
        self.frame_buffer = FrameBuffer(maxsize=3, max_cache=10)
        self.command_queue: Queue = Queue(maxsize=10)
        self.player_thread: Optional[VideoPlayerThread] = None
        
        # UI状态（仅主线程访问）
        self.is_playing = False
        self.is_detecting = False
        self._detect_stats = {}
        self._detect_class_labels: Dict[int, str] = {}
        self._detect_model_name = ""
        self._traffic_stats = None
        self.current_frame = 0
        self.total_frames = 0
        self.fps = 30.0
        self.play_timer = None

        # 时间范围
        self.start_frame = 0
        self.end_frame = 0

        # 当前帧图像（Tkinter PhotoImage，必须在主线程创建）
        self.current_photo: Optional[ImageTk.PhotoImage] = None

        # 视频显示尺寸
        self.display_width = 920
        self.display_height = 540
        self.frame_buffer.set_display_size(self.display_width, self.display_height)

        # 检测参数
        self.model_select_var = tk.IntVar(value=3)
        self.road_model_var = tk.StringVar(value="hsv")  # 道路边界检测模型
        self.device_var = tk.StringVar(value="auto")
        self.confidence_threshold = 0.2
        self.slice_height = 640
        self.slice_width = 640
        self.overlap_h = 0.1
        self.overlap_w = 0.1
        self.skip_interval = 2
        self.max_frames = None

        # 检测功能
        self._vehicle_detect = True
        self._slice_detect = True
        self._enable_tracking = True
        self.var_use_sahi = tk.BooleanVar(value=True)           # 启用SAHI切片检测
        self.var_write_output = tk.BooleanVar(value=True)       # 输出检测视频
        self.var_export_traffic = tk.BooleanVar(value=True)     # 输出流量统计表
        self.var_enable_detection = tk.BooleanVar(value=True)    # 目标检测
        self.var_enable_tracking = tk.BooleanVar(value=True)     # 轨迹跟踪
        self.var_enable_road_detect = tk.BooleanVar(value=False) # 启用车道边界检测
        self.var_auto_traffic = tk.BooleanVar(value=False)       # 自动流量统计
        self.var_show_trajectory = tk.BooleanVar(value=True)   # 实时绘制轨迹
        self.var_export_trajectory_img = tk.BooleanVar(value=True)  # 输出分类轨迹图
        self.var_export_trajectory_data = tk.BooleanVar(value=True)  # 输出轨迹坐标数据

        # 流量统计设置
        self.var_traffic_count = tk.BooleanVar(value=False)
        self.var_show_count_lines = tk.BooleanVar(value=False)  # 检测时显示车型标注
        self.var_show_road_mask = tk.BooleanVar(value=False)    # 检测时标注道路边界
        self._road_segmenter = None  # 道路分割器（延迟加载）
        # 视频分辨率（用于判断是否启用切片）
        self._video_width = 0
        self._video_height = 0
        self.count_lines: List[CrossSectionLine] = []
        self._entry_counter = 0          # 进口道自动命名计数器
        self._line_draw_mode: Optional[str] = None  # None / "start" / "end"
        self._line_start_point: Optional[Tuple[int, int]] = None
        self._line_temp_entry = ""       # 当前画线所属进口道
        self._line_temp_direction = ""   # 当前画线所属方向
        self._line_preview_id: Optional[int] = None  # Canvas 橡皮筋线ID（若用Canvas）

        # 流量统计UI组件引用
        self.tree_traffic = None
        self.lbl_draw_hint = None

        # 自动流量统计UI组件
        self.traj_canvas = None
        self._traj_placeholder_id = None

        # ═══════ UI组件引用 ═══════
        self.status_indicator = None
        self.video_label = None
        self.timeline = None
        self.frame_info = None
        self.time_info = None
        self.time_remaining = None
        self.speed_var = None
        self.text_result = None
        self.stage_label = None
        self.progress_bar = None
        self.progress_rect = None
        self.progress_percent = None
        self.play_btn = None
        self.btn_start = None
        self.btn_stop = None
        self.btn_preview = None
        self.btn_select_video = None

        # 加载配置
        self.config = self._load_config()

        # 构建界面
        self._create_layout()
        self._create_menu()

        # 窗口关闭处理
        self.root.protocol("WM_DELETE_WINDOW", self._on_close)

    def _load_config(self) -> dict:
        """加载配置文件"""
        config_path = PROJECT_ROOT / "configs" / "inference.yaml"
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                return yaml.safe_load(f) or {}
        except Exception as e:
            print(f"加载配置失败: {e}")
            return {}

    def _save_config(self):
        """保存配置到YAML"""
        config_path = PROJECT_ROOT / "configs" / "inference.yaml"
        try:
            self.config['model_select'] = self.model_select_var.get()
            self.config['confidence_threshold'] = self.confidence_threshold
            self.config['slice_height'] = self.slice_h_var.get()
            self.config['slice_width'] = self.slice_w_var.get()
            self.config['overlap_height_ratio'] = self.overlap_h_var.get()
            self.config['overlap_width_ratio'] = self.overlap_w_var.get()
            self.config['detect_skip_interval'] = self.skip_interval
            self.config['max_frames'] = self.max_frames

            with open(config_path, 'w', encoding='utf-8') as f:
                yaml.dump(self.config, f, allow_unicode=True, default_flow_style=False)
            return True
        except Exception as e:
            self._log(f"保存配置失败: {e}")
            return False

    # ═════════════════════════════════════════
    # 菜单栏
    # ═════════════════════════════════════════
    def _create_menu(self):
        menubar = tk.Menu(self.root, bg=C.BG, fg=C.TEXT, font=FONT_BODY,
                         activebackground=C.ACCENT, activeforeground=C.BG)
        self.root.config(menu=menubar, bg=C.BG)

        file_menu = tk.Menu(menubar, tearoff=0, bg=C.CARD, fg=C.TEXT, font=FONT_BODY,
                          activebackground=C.ACCENT, activeforeground=C.BG)
        menubar.add_cascade(label="  文件  ", menu=file_menu)
        file_menu.add_command(label="  打开视频...", command=self.select_video, accelerator="Ctrl+O")
        file_menu.add_command(label="  加载配置...", command=self.load_config_file)
        file_menu.add_command(label="  保存配置", command=self._save_config)
        file_menu.add_separator()
        file_menu.add_command(label="  导出结果...", command=self.export_results)
        file_menu.add_separator()
        file_menu.add_command(label="  退出", command=self._on_close, accelerator="Alt+F4")

        detect_menu = tk.Menu(menubar, tearoff=0, bg=C.CARD, fg=C.TEXT, font=FONT_BODY,
                            activebackground=C.ACCENT, activeforeground=C.BG)
        menubar.add_cascade(label="  检测  ", menu=detect_menu)
        detect_menu.add_command(label="  开始检测", command=self.start_detection, accelerator="F5")
        detect_menu.add_command(label="  停止检测", command=self.stop_detection, accelerator="Esc")
        detect_menu.add_separator()
        detect_menu.add_command(label="  视频预览", command=self.toggle_play, accelerator="Space")

        mark_menu = tk.Menu(menubar, tearoff=0, bg=C.CARD, fg=C.TEXT, font=FONT_BODY,
                          activebackground=C.ACCENT, activeforeground=C.BG)
        menubar.add_cascade(label="  标注  ", menu=mark_menu)
        mark_menu.add_command(label="  开始流量统计线标注", command=self.start_traffic_line_mark)
        mark_menu.add_command(label="  清除所有流量统计线", command=self.clear_all_traffic_lines)
        mark_menu.add_separator()
        mark_menu.add_command(label="  保存截面线配置", command=self.save_traffic_config)
        mark_menu.add_command(label="  加载截面线配置", command=self.load_traffic_config)

        help_menu = tk.Menu(menubar, tearoff=0, bg=C.CARD, fg=C.TEXT, font=FONT_BODY,
                          activebackground=C.ACCENT, activeforeground=C.BG)
        menubar.add_cascade(label="  帮助  ", menu=help_menu)
        help_menu.add_command(label="  使用说明", command=self.show_help)
        help_menu.add_command(label="  参数说明", command=self.show_param_help)
        help_menu.add_command(label="  关于", command=self.show_about)

    # ═════════════════════════════════════════
    # 主布局
    # ═════════════════════════════════════════
    def _create_layout(self):
        # ── 全局根窗口背景 ──
        self.root.configure(bg=C.BG)

        # ── 标题栏 - 深色科技风格 ──
        header = tk.Frame(self.root, bg=C.HEADER, height=56)
        header.pack(fill=tk.X)
        header.pack_propagate(False)

        # 左侧标题区域
        title_frame = tk.Frame(header, bg=C.HEADER)
        title_frame.pack(side=tk.LEFT, padx=20, pady=8)

        tk.Label(
            title_frame, text="无人机航拍视频检测系统",
            bg=C.HEADER, fg=C.TEXT,
            font=FONT_TITLE
        ).pack(anchor=tk.W)

        tk.Label(
            title_frame, text="UAV Aerial Vehicle Detection System  v4.0",
            bg=C.HEADER, fg=C.ACCENT,
            font=("Arial", 8)
        ).pack(anchor=tk.W)

        # 右侧设备指示灯
        status_frame = tk.Frame(header, bg=C.HEADER)
        status_frame.pack(side=tk.RIGHT, padx=20, pady=8)


        # 分隔线
        tk.Frame(self.root, bg=C.BORDER, height=1).pack(fill=tk.X)

        # 主容器
        main_container = tk.Frame(self.root, bg=C.BG)
        main_container.pack(fill=tk.BOTH, expand=True)

        # 左侧文件浏览器
        browser_frame = tk.Frame(main_container, bg=C.PANEL, width=250)
        browser_frame.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 2), pady=5)
        browser_frame.pack_propagate(False)
        self._create_file_browser(browser_frame)

        # 左侧视频区域
        left_panel = tk.Frame(main_container, bg=C.BG)
        left_panel.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(2, 2), pady=5)
        self._create_video_panel(left_panel)

        # 右侧参数面板
        right_panel = tk.Frame(main_container, bg=C.PANEL, width=380)
        right_panel.pack(side=tk.RIGHT, fill=tk.Y, padx=(2, 0), pady=0)
        right_panel.pack_propagate(False)
        self._create_param_panel(right_panel)

        # 扫描工作区文件
        self.root.after(200, self._scan_file_browser)

    # ═════════════════════════════════════════
    # 视频面板
    # ═════════════════════════════════════════
    def _create_video_panel(self, parent: tk.Frame):
        """创建视频播放器界面"""

        # 视频显示框
        video_frame = tk.Frame(parent, bg=C.BG, highlightthickness=1,
                              highlightbackground=C.BORDER)
        video_frame.pack(fill=tk.BOTH, expand=True)

        self.video_label = tk.Label(
            video_frame,
            text="请选择视频文件开始\n支持 MP4、AVI、MOV 等格式",
            bg=C.PANEL, fg=C.TEXT_SUB,
            font=("Microsoft YaHei", 14)
        )
        self.video_label.pack(fill=tk.BOTH, expand=True)

        # 显示尺寸
        self.display_width = 920
        self.display_height = 540
        self.video_label.bind("<Button-1>", self._on_video_click)

        # ─── 时间轴区域 ───
        timeline_frame = tk.Frame(parent, bg=C.PANEL, height=65)
        timeline_frame.pack(fill=tk.X, pady=(3, 0))
        timeline_frame.pack_propagate(False)

        tk.Label(timeline_frame, text="⏱ 时间轴 - 拖动滑块设置检测起止范围",
                bg=C.PANEL, fg=C.TEXT_SUB,
                font=FONT_SMALL).pack(anchor=tk.W, padx=5, pady=(2, 1))

        self.timeline = TimelineSlider(
            timeline_frame,
            height=52,
            command=self._on_timeline_changed
        )
        self.timeline.pack(fill=tk.X, padx=5, pady=1)

        # ─── 播放控制条 ───
        control_frame = tk.Frame(parent, bg=C.BG)
        control_frame.pack(fill=tk.X, pady=(5, 0))

        left_controls = tk.Frame(control_frame, bg=C.BG)
        left_controls.pack(side=tk.LEFT, padx=10)

        def _ctrl_btn(parent, text, cmd, accent=False, font_size=16, w=3):
            return tk.Button(parent, text=text, command=cmd,
                bg=C.ACCENT if accent else C.CARD, fg=C.BG if accent else C.TEXT_SUB,
                relief=tk.FLAT, width=w, height=1, font=("Arial", font_size), cursor="hand2",
                activebackground=C.ACCENT, activeforeground=C.BG)

        self.play_btn = _ctrl_btn(left_controls, "▶", self.toggle_play, accent=True)
        self.play_btn.pack(side=tk.LEFT, padx=2)

        _ctrl_btn(left_controls, "■", self.stop_playback, font_size=12, w=2).pack(side=tk.LEFT, padx=2)
        _ctrl_btn(left_controls, "◀◀", lambda: self._seek_relative(-1), font_size=9, w=3).pack(side=tk.LEFT, padx=(5,2))
        _ctrl_btn(left_controls, "▶▶", lambda: self._seek_relative(1), font_size=9, w=3).pack(side=tk.LEFT, padx=2)
        _ctrl_btn(left_controls, "|◀", lambda: self._seek_absolute(self.start_frame), font_size=9, w=2).pack(side=tk.LEFT, padx=(5,2))
        _ctrl_btn(left_controls, "▶|", lambda: self._seek_absolute(self.end_frame), font_size=9, w=2).pack(side=tk.LEFT, padx=2)

        # 中间时间信息
        center_frame = tk.Frame(control_frame, bg=C.BG)
        center_frame.pack(side=tk.LEFT, fill=tk.X, expand=True)

        self.time_info = tk.Label(center_frame, text="00:00:00",
            bg=C.BG, fg=C.ACCENT, font=FONT_TIME)
        self.time_info.pack(pady=(8, 0))

        self.frame_info = tk.Label(center_frame, text="帧: 0 / 0",
            bg=C.BG, fg=C.TEXT_MUTED, font=FONT_MONO_SM)
        self.frame_info.pack()

        # 右侧控制
        right_controls = tk.Frame(control_frame, bg=C.BG)
        right_controls.pack(side=tk.RIGHT, padx=10)

        self.time_remaining = tk.Label(right_controls, text="-00:00:00",
            bg=C.BG, fg=C.TEXT_MUTED, font=("Consolas", 11))
        self.time_remaining.pack(pady=(5, 0))

        speed_frame = tk.Frame(right_controls, bg=C.BG)
        speed_frame.pack(pady=3)

        tk.Label(speed_frame, text="速度:", bg=C.BG, fg=C.TEXT_SUB, font=FONT_SMALL).pack(side=tk.LEFT)
        self.speed_var = tk.StringVar(value="1.0x")
        speed_combo = ttk.Combobox(speed_frame, textvariable=self.speed_var,
            values=["0.25x", "0.5x", "1.0x", "1.5x", "2.0x", "4.0x"],
            width=5, state="readonly")
        speed_combo.pack(side=tk.LEFT, padx=3)
        speed_combo.bind("<<ComboboxSelected>>", self._on_speed_changed)
        # StringVar trace 作为备用保障，确保倍速变化必触发
        self.speed_var.trace_add("write", lambda *_: self._on_speed_changed())

        # ─── 检测结果区域 ───
        result_frame = tk.Frame(parent, bg=C.PANEL, height=150)
        result_frame.pack(fill=tk.X, pady=(5, 0))
        result_frame.pack_propagate(False)

        tk.Label(result_frame, text="📋 检测日志",
            bg=C.PANEL, fg=C.TEXT, font=FONT_SECTION).pack(anchor=tk.W, padx=10, pady=(5, 0))

        result_text_frame = tk.Frame(result_frame, bg=C.CARD)
        result_text_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        scroll_y = tk.Scrollbar(result_text_frame)
        scroll_y.pack(side=tk.RIGHT, fill=tk.Y)

        self.text_result = tk.Text(result_text_frame,
            bg=C.BG, fg=C.TEXT, insertbackground=C.ACCENT,
            font=("Consolas", 9), yscrollcommand=scroll_y.set,
            relief=tk.FLAT, state="disabled", bd=0, padx=5, pady=3)
        self.text_result.pack(fill=tk.BOTH, expand=True)
        scroll_y.config(command=self.text_result.yview)

        # ─── 进度条 ───
        progress_frame = tk.Frame(parent, bg=C.BG, height=45)
        progress_frame.pack(fill=tk.X, pady=(5, 0))
        progress_frame.pack_propagate(False)

        self.stage_label = tk.Label(progress_frame, text="等待开始",
            bg=C.BG, fg=C.ACCENT, font=("Microsoft YaHei", 9, "bold"))
        self.stage_label.pack(side=tk.TOP, padx=10, pady=(3, 0))

        progress_bottom = tk.Frame(progress_frame, bg=C.BG)
        progress_bottom.pack(fill=tk.X, pady=(2, 5))

        tk.Label(progress_bottom, text="进度:", bg=C.BG, fg=C.TEXT_MUTED,
            font=FONT_BODY).pack(side=tk.LEFT, padx=10)

        self.progress_bar = tk.Canvas(progress_bottom, bg=C.CARD,
            height=12, highlightthickness=0)
        self.progress_bar.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=10, pady=0)
        self.progress_bar.create_rectangle(2, 2, 400, 10, fill=C.BORDER, outline="")
        self.progress_rect = self.progress_bar.create_rectangle(2, 2, 2, 10, fill=C.ACCENT, outline="")

        self.progress_percent = tk.Label(progress_bottom, text="0%",
            bg=C.BG, fg=C.ACCENT, font=("Consolas", 9), width=5)
        self.progress_percent.pack(side=tk.LEFT, padx=5)

    # ═════════════════════════════════════════
    # 参数面板
    # ═════════════════════════════════════════
    def _create_param_panel(self, parent: tk.Frame):
        """创建右侧参数配置面板"""

        # ═══════════ 功能配置 ═══════════
        tk.Label(parent, text="⚙ 功能配置",
                bg=C.PANEL, fg=C.TEXT, font=FONT_HEADER).pack(anchor=tk.W, padx=15, pady=(10, 5))

        # Row 1: 目标检测 + 轨迹跟踪
        quick_frame = tk.Frame(parent, bg=C.CARD)
        quick_frame.pack(fill=tk.X, padx=10, pady=(3, 0))

        cb_detect = tk.Checkbutton(
            quick_frame, text="🎯 目标检测",
            variable=self.var_enable_detection,
            bg=C.CARD, fg=C.GREEN,
            selectcolor=C.SURFACE, activebackground=C.CARD,
            activeforeground=C.GREEN, font=FONT_BODY,
            anchor=tk.W, padx=10, pady=2
        )
        cb_detect.pack(side=tk.LEFT)

        cb_track = tk.Checkbutton(
            quick_frame, text="🔗 轨迹跟踪",
            variable=self.var_enable_tracking,
            bg=C.CARD, fg=C.GREEN,
            selectcolor=C.SURFACE, activebackground=C.CARD,
            activeforeground=C.GREEN, font=FONT_BODY,
            anchor=tk.W, padx=10, pady=2
        )
        cb_track.pack(side=tk.LEFT)

        # Row 2: 道路边界检测
        quick_frame2 = tk.Frame(parent, bg=C.CARD)
        quick_frame2.pack(fill=tk.X, padx=10, pady=(0, 3))

        cb_road_detect = tk.Checkbutton(
            quick_frame2, text="🛤 道路边界检测",
            variable=self.var_enable_road_detect,
            bg=C.CARD, fg=C.GREEN,
            selectcolor=C.SURFACE, activebackground=C.CARD,
            activeforeground=C.GREEN, font=FONT_BODY,
            anchor=tk.W, padx=10, pady=2
        )
        cb_road_detect.pack(side=tk.LEFT)

        # ═══════════ 滚动区域 ═══════════
        canvas = tk.Canvas(parent, bg=C.PANEL, highlightthickness=0)
        scrollbar = tk.Scrollbar(parent, orient="vertical", command=canvas.yview)
        scroll_frame = tk.Frame(canvas, bg=C.PANEL)

        scroll_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        _canvas_window = canvas.create_window((0, 0), window=scroll_frame, anchor="nw")
        def _on_canvas_configure(event):
            canvas.itemconfig(_canvas_window, width=event.width)
        canvas.bind("<Configure>", _on_canvas_configure)
        canvas.configure(yscrollcommand=scrollbar.set)

        # ── 检测模型 ──
        model_section = tk.Frame(scroll_frame, bg=C.CARD)
        model_section.pack(fill=tk.X, padx=0, pady=(3, 5))

        tk.Label(model_section, text="🤖 检测模型",
                bg=C.CARD, fg=C.TEXT, font=FONT_SECTION).pack(anchor=tk.W, padx=10, pady=(8, 5))

        model_columns = tk.Frame(model_section, bg=C.CARD)
        model_columns.pack(fill=tk.X, padx=5)

        # 左栏：目标检测模型
        left_col = tk.Frame(model_columns, bg=C.CARD)
        left_col.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(5, 3))

        tk.Label(left_col, text="🎯 目标检测",
                bg=C.CARD, fg=C.ACCENT, font=("Microsoft YaHei", 9, "bold")).pack(anchor=tk.W, padx=5, pady=(2, 3))

        for i, (key, name) in enumerate([
            (1, "WALDO YOLOv8m (平衡)"),
            (2, "RT-DETR (高精度)"),
            (3, "YOLOv8l (大模型)")
        ]):
            rb = tk.Radiobutton(left_col, text=f"  {name}",
                variable=self.model_select_var, value=key,
                bg=C.CARD, fg=C.TEXT,
                selectcolor=C.SURFACE, activebackground=C.CARD,
                activeforeground=C.TEXT, font=FONT_BODY,
                anchor=tk.W, padx=10, pady=2)
            rb.pack(fill=tk.X)

        # 右栏
        right_col = tk.Frame(model_columns, bg=C.CARD)
        right_col.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(3, 5))

        tk.Label(right_col, text="🛤 道路边界检测",
                bg=C.CARD, fg=C.GREEN, font=("Microsoft YaHei", 9, "bold")).pack(anchor=tk.W, padx=5, pady=(2, 3))

        for key, name in [
            ("hsv", "HSV 色彩 (快速)"),
            ("segformer", "SegFormer (深度学习)"),
            ("hrnet_ocr", "HRNet-OCR (高精度)")
        ]:
            rb = tk.Radiobutton(right_col, text=f"  {name}",
                variable=self.road_model_var, value=key,
                bg=C.CARD, fg=C.TEXT,
                selectcolor=C.SURFACE, activebackground=C.CARD,
                activeforeground=C.TEXT, font=FONT_BODY,
                anchor=tk.W, padx=10, pady=2)
            rb.pack(fill=tk.X)

        # --- 检测参数 ---
        tk.Frame(scroll_frame, bg=C.BORDER, height=1).pack(fill=tk.X, padx=0)

        param_section = tk.Frame(scroll_frame, bg=C.CARD)
        param_section.pack(fill=tk.X, padx=0, pady=5)

        tk.Label(param_section, text="📊 检测参数",
                bg=C.CARD, fg=C.TEXT, font=FONT_SECTION).pack(anchor=tk.W, padx=10, pady=(8, 5))

        param_row1 = tk.Frame(param_section, bg=C.CARD)
        param_row1.pack(fill=tk.X, padx=5, pady=3)

        left_half = tk.Frame(param_row1, bg=C.CARD, width=1)
        left_half.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 2))
        self.conf_scale = FloatScale(left_half, "置信度阈值:", 0.01, 0.95, 0.2,
            command=self._on_confidence_changed)
        self.conf_scale.pack(fill=tk.X)

        right_half = tk.Frame(param_row1, bg=C.CARD, width=1)
        right_half.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(2, 0))
        self.skip_scale = IntScale(right_half, "跳帧间隔:", 0, 10, 2,
            command=self._on_skip_interval_changed)
        self.skip_scale.pack(fill=tk.X)

        cb_sahi = tk.Checkbutton(param_section, text="🧩 启用切片检测 (SAHI)",
            variable=self.var_use_sahi,
            bg=C.CARD, fg=C.GREEN, selectcolor=C.SURFACE,
            activebackground=C.CARD, activeforeground=C.GREEN,
            font=FONT_BODY, anchor=tk.W, padx=10, pady=4)
        cb_sahi.pack(fill=tk.X)

        param_row2 = tk.Frame(param_section, bg=C.CARD)
        param_row2.pack(fill=tk.X, padx=10, pady=3)

        slice_col = tk.Frame(param_row2, bg=C.CARD)
        slice_col.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        tk.Label(slice_col, text="切片尺寸", bg=C.CARD, fg=C.TEXT_MUTED, font=FONT_SMALL).pack(anchor=tk.W)
        slice_inner = tk.Frame(slice_col, bg=C.CARD); slice_inner.pack(fill=tk.X)
        tk.Label(slice_inner, text="H:", bg=C.CARD, fg=C.TEXT_SUB, font=FONT_MONO_SM).pack(side=tk.LEFT)
        self.slice_h_var = tk.IntVar(value=640)
        ttk.Combobox(slice_inner, textvariable=self.slice_h_var, values=[320,480,640,800,1024],
            width=5, state="readonly").pack(side=tk.LEFT, padx=(0, 8))
        tk.Label(slice_inner, text="W:", bg=C.CARD, fg=C.TEXT_SUB, font=FONT_MONO_SM).pack(side=tk.LEFT)
        self.slice_w_var = tk.IntVar(value=640)
        ttk.Combobox(slice_inner, textvariable=self.slice_w_var, values=[320,480,640,800,1024],
            width=5, state="readonly").pack(side=tk.LEFT)

        overlap_col = tk.Frame(param_row2, bg=C.CARD)
        overlap_col.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(5, 0))
        tk.Label(overlap_col, text="切片重叠率", bg=C.CARD, fg=C.TEXT_MUTED, font=FONT_SMALL).pack(anchor=tk.W)
        overlap_inner = tk.Frame(overlap_col, bg=C.CARD); overlap_inner.pack(fill=tk.X)
        tk.Label(overlap_inner, text="H:", bg=C.CARD, fg=C.TEXT_SUB, font=FONT_MONO_SM).pack(side=tk.LEFT)
        self.overlap_h_var = tk.DoubleVar(value=0.1)
        ttk.Combobox(overlap_inner, textvariable=self.overlap_h_var, values=[0.0,0.1,0.2,0.3,0.5],
            width=5, state="readonly").pack(side=tk.LEFT, padx=(0, 8))
        tk.Label(overlap_inner, text="W:", bg=C.CARD, fg=C.TEXT_SUB, font=FONT_MONO_SM).pack(side=tk.LEFT)
        self.overlap_w_var = tk.DoubleVar(value=0.1)
        ttk.Combobox(overlap_inner, textvariable=self.overlap_w_var, values=[0.0,0.1,0.2,0.3,0.5],
            width=5, state="readonly").pack(side=tk.LEFT)

        max_frame_frame = tk.Frame(param_section, bg=C.CARD)
        max_frame_frame.pack(fill=tk.X, padx=10, pady=5)
        tk.Label(max_frame_frame, text="最大处理帧数:", bg=C.CARD, fg=C.TEXT_SUB, font=FONT_BODY).pack(side=tk.LEFT)
        self.max_frames_var = tk.StringVar(value="全部")
        tk.Entry(max_frame_frame, textvariable=self.max_frames_var,
            bg=C.SURFACE, fg=C.TEXT, font=FONT_MONO_SM, width=10,
            insertbackground=C.ACCENT, relief=tk.FLAT).pack(side=tk.LEFT, padx=5)
        tk.Label(max_frame_frame, text="(留空=全部)", bg=C.CARD, fg=C.TEXT_HINT, font=FONT_TINY).pack(side=tk.LEFT)

        # --- 检测显示设置 ---
        tk.Frame(scroll_frame, bg=C.BORDER, height=1).pack(fill=tk.X, padx=0)

        display_section = tk.Frame(scroll_frame, bg=C.CARD)
        display_section.pack(fill=tk.X, padx=0, pady=5)

        tk.Label(display_section, text="🖥 检测显示设置",
                bg=C.CARD, fg=C.TEXT, font=FONT_SECTION).pack(anchor=tk.W, padx=10, pady=(8, 5))

        display_row = tk.Frame(display_section, bg=C.CARD)
        display_row.pack(fill=tk.X, padx=10, pady=4)

        cb_show_lines = tk.Checkbutton(display_row, text="📦 显示检测框和车型",
            variable=self.var_show_count_lines,
            bg=C.CARD, fg=C.GREEN, selectcolor=C.SURFACE,
            activebackground=C.CARD, activeforeground=C.GREEN,
            font=FONT_BODY, anchor=tk.W, padx=8, pady=2)
        cb_show_lines.pack(side=tk.LEFT, padx=(0, 10))

        cb_road_mask = tk.Checkbutton(display_row, text="🛤 显示道路边界",
            variable=self.var_show_road_mask,
            bg=C.CARD, fg=C.GREEN, selectcolor=C.SURFACE,
            activebackground=C.CARD, activeforeground=C.GREEN,
            font=FONT_BODY, anchor=tk.W, padx=8, pady=2)
        cb_road_mask.pack(side=tk.LEFT)

        # 显示轨迹（第二行）
        display_row2 = tk.Frame(display_section, bg=C.CARD)
        display_row2.pack(fill=tk.X, padx=10, pady=(0, 4))

        cb_show_traj = tk.Checkbutton(display_row2, text="🛤 显示轨迹",
            variable=self.var_show_trajectory,
            bg=C.CARD, fg=C.GREEN, selectcolor=C.SURFACE,
            activebackground=C.CARD, activeforeground=C.GREEN,
            font=FONT_BODY, anchor=tk.W, padx=8, pady=2)
        cb_show_traj.pack(side=tk.LEFT)

        # --- 流量统计方法 ---
        tk.Frame(scroll_frame, bg=C.BORDER, height=1).pack(fill=tk.X, padx=0)

        flow_section = tk.Frame(scroll_frame, bg=C.CARD)
        flow_section.pack(fill=tk.X, padx=0, pady=5)

        tk.Label(flow_section, text="📊 流量统计方法",
                bg=C.CARD, fg=C.TEXT, font=FONT_SECTION).pack(anchor=tk.W, padx=10, pady=(8, 5))

        flow_row = tk.Frame(flow_section, bg=C.CARD)
        flow_row.pack(fill=tk.X, padx=10, pady=4)

        cb_traffic_flow = tk.Checkbutton(flow_row, text="📏 画线流量统计",
            variable=self.var_traffic_count,
            bg=C.CARD, fg=C.GREEN, selectcolor=C.SURFACE,
            activebackground=C.CARD, activeforeground=C.GREEN,
            font=FONT_BODY, anchor=tk.W, padx=8, pady=2)
        cb_traffic_flow.pack(side=tk.LEFT, padx=(0, 10))

        cb_auto_flow = tk.Checkbutton(flow_row, text="📈 自动流量统计",
            variable=self.var_auto_traffic,
            bg=C.CARD, fg=C.GREEN, selectcolor=C.SURFACE,
            activebackground=C.CARD, activeforeground=C.GREEN,
            font=FONT_BODY, anchor=tk.W, padx=8, pady=2)
        cb_auto_flow.pack(side=tk.LEFT)

        # --- 输出设置 ---
        tk.Frame(scroll_frame, bg=C.BORDER, height=1).pack(fill=tk.X, padx=0)

        output_section = tk.Frame(scroll_frame, bg=C.CARD)
        output_section.pack(fill=tk.X, padx=0, pady=5)

        tk.Label(output_section, text="📤 输出设置",
                bg=C.CARD, fg=C.TEXT, font=FONT_SECTION).pack(anchor=tk.W, padx=10, pady=(8, 5))

        output_row = tk.Frame(output_section, bg=C.CARD)
        output_row.pack(fill=tk.X, padx=10, pady=4)

        cb_output = tk.Checkbutton(output_row, text="📹 输出检测视频",
            variable=self.var_write_output,
            bg=C.CARD, fg=C.GREEN, selectcolor=C.SURFACE,
            activebackground=C.CARD, activeforeground=C.GREEN,
            font=FONT_BODY, anchor=tk.W, padx=8, pady=2)
        cb_output.pack(side=tk.LEFT, padx=(0, 10))

        self.cb_traffic_xls = tk.Checkbutton(output_row, text="📊 输出流量统计表",
            variable=self.var_export_traffic,
            bg=C.CARD, fg=C.GREEN, selectcolor=C.SURFACE,
            activebackground=C.CARD, activeforeground=C.GREEN,
            font=FONT_BODY, anchor=tk.W, padx=8, pady=2)
        self.cb_traffic_xls.pack(side=tk.LEFT)

        # 联动：画线流量统计 或 自动流量统计 至少勾选一个时，输出流量统计表才可用
        def _on_flow_stats_toggle(*_args):
            has_line = self.var_traffic_count.get()
            has_auto = self.var_auto_traffic.get()
            state = "normal" if (has_line or has_auto) else "disabled"
            self.cb_traffic_xls.config(state=state)
            if not (has_line or has_auto):
                self.var_export_traffic.set(False)
        self.var_traffic_count.trace_add("write", _on_flow_stats_toggle)
        self.var_auto_traffic.trace_add("write", _on_flow_stats_toggle)
        _on_flow_stats_toggle()  # 初始同步

        # 轨迹输出设置（第二行，仅在自动流量统计勾选时可用）
        output_row2 = tk.Frame(output_section, bg=C.CARD)
        output_row2.pack(fill=tk.X, padx=10, pady=(0, 4))

        self.cb_export_traj_img = tk.Checkbutton(output_row2, text="🖼 输出分类轨迹图",
            variable=self.var_export_trajectory_img,
            bg=C.CARD, fg=C.GREEN, selectcolor=C.SURFACE,
            activebackground=C.CARD, activeforeground=C.GREEN,
            font=FONT_BODY, anchor=tk.W, padx=8, pady=2)
        self.cb_export_traj_img.pack(side=tk.LEFT, padx=(0, 10))

        self.cb_export_traj_data = tk.Checkbutton(output_row2, text="📋 输出轨迹坐标数据",
            variable=self.var_export_trajectory_data,
            bg=C.CARD, fg=C.GREEN, selectcolor=C.SURFACE,
            activebackground=C.CARD, activeforeground=C.GREEN,
            font=FONT_BODY, anchor=tk.W, padx=8, pady=2)
        self.cb_export_traj_data.pack(side=tk.LEFT)

        # 绑定：自动流量统计勾选状态联动轨迹输出复选框
        def _on_auto_traffic_toggle(*_args):
            enabled = self.var_auto_traffic.get()
            state = "normal" if enabled else "disabled"
            self.cb_export_traj_img.config(state=state)
            self.cb_export_traj_data.config(state=state)
        self.var_auto_traffic.trace_add("write", _on_auto_traffic_toggle)
        # 初始状态同步
        _on_auto_traffic_toggle()

        # ── 底部固定：操作按钮组 (2×2) 必须先于 canvas pack，否则被 expand=True 挤出视图 ──
        bottom_frame = tk.Frame(parent, bg=C.PANEL)
        bottom_frame.pack(side=tk.BOTTOM, fill=tk.X, padx=0, pady=0)

        tk.Frame(bottom_frame, bg=C.BORDER, height=1).pack(fill=tk.X, padx=10, pady=(0, 3))
        btn_panel = tk.Frame(bottom_frame, bg=C.CARD)
        btn_panel.pack(fill=tk.X, padx=10, pady=(5, 8))

        # Row 1
        r1 = tk.Frame(btn_panel, bg=C.CARD)
        r1.pack(fill=tk.X, pady=(2, 1))

        self.btn_select_video = tk.Button(r1, text="📹 选择视频", command=self.select_video,
            bg=C.ACCENT, fg=C.BG, relief=tk.FLAT, padx=8, pady=6,
            font=("Microsoft YaHei", 10, "bold"), cursor="hand2",
            activebackground=C.ACCENT_HOVER, activeforeground=C.BG)
        self.btn_select_video.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=2)

        self.btn_preview = tk.Button(r1, text="🔍 视频预览", command=self.toggle_play,
            bg=C.BORDER_LIGHT, fg=C.TEXT, relief=tk.FLAT, padx=8, pady=6,
            font=("Microsoft YaHei", 10, "bold"), cursor="hand2",
            activebackground=C.BLUE, activeforeground=C.BG)
        self.btn_preview.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=2)

        # Row 2
        r2 = tk.Frame(btn_panel, bg=C.CARD)
        r2.pack(fill=tk.X, pady=(1, 2))

        self.btn_start = tk.Button(r2, text="▶ 开始检测", command=self.start_detection,
            bg=C.GREEN, fg=C.BG, relief=tk.FLAT, padx=8, pady=6,
            font=("Microsoft YaHei", 10, "bold"), cursor="hand2",
            activebackground="#0ef0a0", activeforeground=C.BG)
        self.btn_start.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=2)

        self.btn_stop = tk.Button(r2, text="⏹ 停止检测", command=self.stop_detection,
            bg=C.BTN_DISABLED, fg=C.BTN_DISABLED_TEXT, relief=tk.FLAT, padx=8, pady=6,
            font=("Microsoft YaHei", 10, "bold"), cursor="hand2", state="disabled",
            activebackground=C.RED, activeforeground=C.BG)
        self.btn_stop.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=2)

        # canvas 扩展（必须在 bottom_frame 之后 pack，否则 expand=True 会将空间占满）
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    # ═════════════════════════════════════════
    # 文件导航浏览器
    # ═════════════════════════════════════════

    def _create_file_browser(self, parent: tk.Frame):
        """创建左侧文件导航面板"""
        tk.Label(parent, text="📁 文件浏览器",
                bg=C.PANEL, fg=C.TEXT, font=FONT_HEADER).pack(anchor=tk.W, padx=12, pady=(8, 4))

        # 工具栏
        tb = tk.Frame(parent, bg=C.PANEL)
        tb.pack(fill=tk.X, padx=8, pady=(0, 4))
        tk.Button(tb, text="🔄", command=self._scan_file_browser,
                  bg=C.CARD, fg=C.TEXT, relief=tk.FLAT, font=("Arial", 8),
                  cursor="hand2", activebackground=C.ACCENT).pack(side=tk.LEFT, padx=(0, 4))
        tk.Label(tb, text="data / outputs", bg=C.PANEL, fg=C.TEXT_MUTED,
                font=FONT_SMALL).pack(side=tk.LEFT)

        # Treeview
        tree_frame = tk.Frame(parent, bg=C.PANEL)
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=6, pady=(0, 6))

        self.file_tree = ttk.Treeview(tree_frame, show="tree", height=999,
                                      selectmode="browse")
        style = ttk.Style()
        style.configure("Treeview", background=C.BG, foreground=C.TEXT,
                       fieldbackground=C.BG, font=("Microsoft YaHei", 9))
        style.configure("Treeview.Heading", background=C.CARD, foreground=C.ACCENT,
                       font=("Microsoft YaHei", 9, "bold"))
        style.map("Treeview", background=[("selected", C.ACCENT)],
                 foreground=[("selected", C.BG)])

        file_scroll = tk.Scrollbar(tree_frame, orient="vertical", command=self.file_tree.yview)
        self.file_tree.configure(yscrollcommand=file_scroll.set)
        self.file_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        file_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        self.file_tree.bind("<Double-1>", self._on_file_double_click)
        self.file_tree.bind("<Button-3>", self._on_file_right_click)

        # 初始占位
        self._file_root_data = self.file_tree.insert("", tk.END, text="📂 data", tags=("folder",))
        self._file_root_out = self.file_tree.insert("", tk.END, text="📂 outputs", tags=("folder",))

        # 高亮标记
        self._file_highlighted: str = ""

    def _scan_file_browser(self, *args):
        """扫描 data/ 和 outputs/ 目录，刷新文件树（保留展开状态）"""
        import os, glob as _glob
        tree = self.file_tree

        # 保存展开状态
        expanded_paths = self._save_expanded_state()

        # 清除
        for child in tree.get_children():
            tree.delete(child)

        self._file_root_data = tree.insert("", tk.END, text="📂 data", open=True, tags=("folder",))
        self._file_root_out = tree.insert("", tk.END, text="📂 outputs", open=True, tags=("folder",))

        data_dir = PROJECT_ROOT / "data"
        out_dir = PROJECT_ROOT / "outputs"

        # 扫描 data/
        if data_dir.exists():
            self._scan_dir(data_dir, self._file_root_data, is_data=True)

        # 扫描 outputs/
        if out_dir.exists():
            self._scan_dir(out_dir, self._file_root_out, is_data=False)

        # 恢复展开状态
        self._restore_expanded_state(expanded_paths)

    def _save_expanded_state(self) -> set:
        """保存所有展开节点的路径标识"""
        tree = self.file_tree
        expanded = set()
        for item in tree.get_children():
            self._collect_expanded(tree, item, "", expanded)
        return expanded

    def _collect_expanded(self, tree, item, prefix, expanded):
        text = tree.item(item, "text").strip()
        path = f"{prefix}/{text}" if prefix else text
        if tree.item(item, "open"):
            expanded.add(path)
            for child in tree.get_children(item):
                self._collect_expanded(tree, child, path, expanded)

    def _restore_expanded_state(self, expanded_paths: set):
        """恢复节点展开状态"""
        tree = self.file_tree

        def _restore(item, prefix):
            text = tree.item(item, "text").strip()
            path = f"{prefix}/{text}" if prefix else text
            if path in expanded_paths:
                tree.item(item, open=True)
            for child in tree.get_children(item):
                _restore(child, path)

        for root_item in tree.get_children():
            _restore(root_item, "")

    def _scan_dir(self, directory: Path, parent_node: str, is_data: bool):
        """扫描目录文件添加到树节点"""
        import os
        tree = self.file_tree
        exts = {".mp4", ".avi", ".mov", ".mkv", ".png", ".jpg", ".jpeg", ".csv", ".xlsx"}

        files = []
        dirs = []
        try:
            for entry in sorted(directory.iterdir(), key=lambda x: (not x.is_dir(), x.name.lower())):
                if entry.is_dir() and not entry.name.startswith("."):
                    dirs.append(entry)
                elif entry.is_file() and entry.suffix.lower() in exts:
                    files.append(entry)
        except PermissionError:
            return

        # 子目录
        for d in dirs:
            sub_node = tree.insert(parent_node, tk.END, text=f"📁 {d.name}", tags=("folder",))
            self._scan_dir(d, sub_node, is_data)

        # 文件
        for f in files:
            ext = f.suffix.lower()
            if ext in {".mp4", ".avi", ".mov", ".mkv"}:
                icon = "🎬"
            elif ext in {".png", ".jpg", ".jpeg"}:
                icon = "🖼"
            elif ext == ".csv":
                icon = "📊"
            elif ext == ".xlsx":
                icon = "📋"
            else:
                icon = "📄"

            tags = ("file",)
            if ext in {".mp4", ".avi", ".mov", ".mkv"}:
                tags = ("video",)

            item = tree.insert(parent_node, tk.END, text=f"{icon} {f.name}",
                              values=(str(f),), tags=tags)

            # 检查是否有关联的截面线配置（periods 结构）
            if is_data and ext in {".mp4", ".avi", ".mov", ".mkv"}:
                full_cfg = self._load_config_yaml(f)
                if full_cfg and "periods" in full_cfg:
                    for period_name, period_data in full_cfg["periods"].items():
                        st = period_data.get("start_time", "")
                        et = period_data.get("end_time", "")
                        time_str = f"({st} - {et})" if st and et else ""
                        period_node = tree.insert(item, tk.END,
                            text=f"  ⏱ {period_name} {time_str}", tags=("period",),
                            values=(str(f), period_name))
                        entries_d = period_data.get("entries", {})
                        for entry_name, entry_data in entries_d.items():
                            if isinstance(entry_data, dict):
                                entry_node = tree.insert(period_node, tk.END,
                                    text=f"    ├ {entry_name}", tags=("entry",),
                                    values=(str(f), period_name, entry_name))
                                lines = entry_data.get("lines", [])
                                dirs = []
                                for line in lines:
                                    d = line.get("direction", "")
                                    ld = line.get("line_id", "")
                                    if d and d not in [x[0] for x in dirs]:
                                        dirs.append((d, ld))
                                if not dirs:
                                    dirs.append(("(无线)", ""))
                                for dname, lid in dirs:
                                    tree.insert(entry_node, tk.END,
                                        text=f"      └ {dname}", tags=("direction",),
                                        values=(str(f), period_name, entry_name, dname, lid))

            # 高亮当前视频
            if self.video_path and f.resolve() == self.video_path.resolve():
                tree.selection_set(item)
                tree.see(item)

    def _find_config_for_video(self, video_path: Path) -> dict[str, list[str]]:
        """查找视频对应的截面线配置，返回 {进口道名: [转向方向列表]} (兼容旧版)"""
        full = self._load_config_yaml(video_path)
        if not full:
            return {}
        # 平面化: periods → entries → lines → directions
        result: dict[str, list[str]] = {}
        for period_name, period_data in full.get("periods", {}).items():
            prefix = f"{period_name}/" if len(full.get("periods", {})) > 1 else ""
            for entry_name, entry_data in period_data.get("entries", {}).items():
                if isinstance(entry_data, dict):
                    lines = entry_data.get('lines', [])
                    dirs = []
                    for line in lines:
                        d = line.get('direction', '')
                        if d and d not in dirs:
                            dirs.append(d)
                    key = f"{prefix}{entry_name}"
                    result[key] = dirs if dirs else ["(无线)"]
        return result

    def _load_config_yaml(self, video_path: Path) -> dict | None:
        """加载视频 YAML 配置（支持 periods 结构，兼容旧版 entries 结构）"""
        config_dirs = [
            PROJECT_ROOT / "configs" / "traffic_lines",
            PROJECT_ROOT / "configs" / "cross_section_lines",
        ]
        stem = video_path.stem
        for cfg_dir in config_dirs:
            if cfg_dir.exists():
                for yf in cfg_dir.iterdir():
                    if yf.suffix in (".yaml", ".yml") and yf.stem == stem:
                        try:
                            import yaml
                            with open(yf, 'r', encoding='utf-8') as fcfg:
                                data = yaml.safe_load(fcfg) or {}
                            # 旧版兼容：只有 entries 无 periods → 自动包装为"全时段"
                            if "periods" not in data and "entries" in data:
                                fps = 30.0
                                cap = cv2.VideoCapture(str(video_path))
                                if cap.isOpened():
                                    total = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
                                    fps = cap.get(cv2.CAP_PROP_FPS) or 30.0
                                    cap.release()
                                else:
                                    total = 9999
                                ts = int(total / max(1.0, fps)) if fps > 0 else 0
                                m, s = divmod(ts, 60); h, m = divmod(m, 60)
                                data = {
                                    "video_name": data.get("video_name", video_path.name),
                                    "periods": {
                                        "全时段": {
                                            "start": 0, "end": total,
                                            "start_time": "00:00:00",
                                            "end_time": f"{h:02d}:{m:02d}:{s:02d}",
                                            "entries": data.get("entries", {}),
                                        }
                                    }
                                }
                            return data
                        except Exception:
                            pass
        return None

    def _save_config_yaml(self, video_path: Path, data: dict) -> str | None:
        """保存 YAML 配置到 cross_section_lines/ 目录"""
        import yaml
        config_dir = PROJECT_ROOT / "configs" / "cross_section_lines"
        config_dir.mkdir(parents=True, exist_ok=True)
        safe_name = video_path.stem.replace(" ", "_").replace(".", "_")
        config_path = config_dir / f"{safe_name}.yaml"
        data.setdefault("video_name", video_path.name)
        with open(config_path, "w", encoding="utf-8") as f:
            yaml.dump(data, f, allow_unicode=True, default_flow_style=False, sort_keys=False)
        return str(config_path)

    def _find_config_path(self, video_path: Path) -> Path | None:
        """查找视频对应的 YAML 配置文件路径"""
        config_dirs = [
            PROJECT_ROOT / "configs" / "traffic_lines",
            PROJECT_ROOT / "configs" / "cross_section_lines",
        ]
        stem = video_path.stem
        for cfg_dir in config_dirs:
            if cfg_dir.exists():
                for yf in cfg_dir.iterdir():
                    if yf.suffix in (".yaml", ".yml") and yf.stem == stem:
                        return yf
        return None

    def _seconds_to_time_str(self, seconds: float) -> str:
        m, s = divmod(int(seconds), 60); h, m = divmod(m, 60)
        return f"{h:02d}:{m:02d}:{s:02d}"

    def _frames_to_time_str(self, frame: int, fps: float) -> str:
        return self._seconds_to_time_str(frame / max(1.0, fps))

    def _on_file_double_click(self, event):
        """双击文件浏览器节点 —— 时段双击跳转到对应范围"""
        tree = self.file_tree
        item = tree.identify_row(event.y)
        if not item: return
        tags = str(tree.item(item, "tags"))
        values = tree.item(item, "values") or []

        # 时段节点：双击 → 设置起终点范围并跳转
        if "period" in tags and len(values) >= 2:
            video_path = Path(values[0])
            period_name = values[1]
            self._apply_period_range(video_path, period_name)
            return

        # 其他节点：打开对应视频
        video_path = self._find_video_for_node(item)
        if video_path and video_path.exists():
            self._open_video_from_browser(video_path)

    def _apply_period_range(self, video_path: Path, period_name: str):
        """应用时段范围：打开视频（如未打开）→ 设置起终点 → 跳转到起点"""
        data = self._load_config_yaml(video_path)
        if not data: return
        pd = data.get("periods", {}).get(period_name)
        if not pd:
            self._log(f"时段配置不存在: {period_name}")
            return

        start = int(pd.get("start", 0))
        end = int(pd.get("end", self.total_frames or 9999))

        # 如果视频未打开或不是当前视频，先打开
        current_open = self.video_path and self.video_path.resolve() == video_path.resolve()
        if not current_open:
            self._open_video_from_browser(video_path)
            # _load_video 会重置 start/end，需要 after 等加载完成后覆盖
            self.root.after(400, lambda: self._do_set_period_range_ui(start, end, period_name))
        else:
            self._do_set_period_range_ui(start, end, period_name)

    def _do_set_period_range_ui(self, start: int, end: int, period_name: str = ""):
        """设置UI起终点范围并跳转到起点（主线程调用）"""
        self.start_frame = start
        self.end_frame = min(end, self.total_frames - 1) if self.total_frames > 0 else end
        # 同步播放线程
        if self.player_thread and self.player_thread.is_alive():
            self.player_thread._start_frame = self.start_frame
            self.player_thread._end_frame = self.end_frame
        # 更新时间线
        self.timeline.configure(
            min_val=0,
            max_val=max(1, self.total_frames - 1),
            start_value=self.start_frame,
            end_value=self.end_frame,
        )
        self.timeline.set_playhead(self.start_frame)
        # 加载该时段的截面线
        self._load_period_lines(period_name)
        # 暂停播放并跳转到起点帧
        self._pause()
        self._seek_absolute(self.start_frame)
        self._update_frame_info()
        label = f"时段「{period_name}」" if period_name else ""
        self._log(f"{label}范围已加载: 帧 {self.start_frame} - {self.end_frame}")

    def _load_period_lines(self, period_name: str):
        """从 YAML 加载指定时段的截面线到 self.count_lines"""
        if not self.video_path or not period_name:
            return
        data = self._load_config_yaml(self.video_path)
        if not data:
            return
        periods = data.get("periods", {})
        pd = periods.get(period_name)
        if not pd:
            return
        entries = pd.get("entries", {})
        lines = []
        for entry_name, entry_data in entries.items():
            if not isinstance(entry_data, dict):
                continue
            for ld in entry_data.get("lines", []):
                try:
                    from src.cross_section_counter import CrossSectionLine
                    cl = CrossSectionLine(
                        line_id=ld.get("line_id", ""),
                        entry_name=entry_name,
                        direction=ld.get("direction", ""),
                        x1=int(ld.get("x1", 0)), y1=int(ld.get("y1", 0)),
                        x2=int(ld.get("x2", 0)), y2=int(ld.get("y2", 0)),
                        color=tuple(ld.get("color", (0, 255, 255))),
                        is_drawn=ld.get("is_drawn", True),
                    )
                    lines.append(cl)
                except Exception:
                    pass
        if lines:
            self.count_lines = lines
            self._log(f"已加载时段「{period_name}」的 {len(lines)} 条截面线")

    def _find_video_for_node(self, item: str) -> Path | None:
        """向上查找节点树，获取所属视频文件路径"""
        tree = self.file_tree
        while item:
            values = tree.item(item, "values")
            tags = tree.item(item, "tags")
            if "video" in str(tags) and values:
                return Path(values[0])
            item = tree.parent(item)
        return None

    def _on_file_right_click(self, event):
        """右键文件浏览器节点 —— 按节点层级分发菜单"""
        tree = self.file_tree
        item = tree.identify_row(event.y)
        if not item: return
        tree.selection_set(item)

        # 统一读取 tags（兼容 tkinter 版本差异：可能返回 string / tuple / list）
        raw_tags = tree.item(item, "tags")
        if isinstance(raw_tags, str):
            tag_list = [raw_tags]
        elif isinstance(raw_tags, (list, tuple)):
            tag_list = list(raw_tags)
        else:
            tag_list = []
        tag_set = set(tag_list)

        # 统一读取 values（同上兼容处理）
        raw_vals = tree.item(item, "values")
        if isinstance(raw_vals, str):
            vals = [raw_vals]
        elif isinstance(raw_vals, (list, tuple)):
            vals = list(raw_vals)
        else:
            vals = []
        vals_len = len(vals)

        under_outputs = self._is_under_outputs(item)
        print(f"[右键] item={item} tags={tag_list} vals_len={vals_len} under_outputs={under_outputs}")

        try:
            if "direction" in tag_set and vals_len >= 5:
                self._show_direction_menu(item, Path(vals[0]), vals[1], vals[2], vals[3], vals[4],
                                          event.x_root, event.y_root)
            elif "entry" in tag_set and vals_len >= 3:
                print(f"[菜单→进口道] vals={vals}")
                self._show_entry_menu(item, Path(vals[0]), vals[1], vals[2],
                                      event.x_root, event.y_root)
            elif "period" in tag_set and vals_len >= 2:
                self._show_period_menu(item, Path(vals[0]), vals[1],
                                       event.x_root, event.y_root)
            elif "video" in tag_set and not under_outputs:
                vp = Path(vals[0]) if vals else None
                if vp and vp.exists():
                    self._show_data_video_menu(item, vp, event.x_root, event.y_root)
            elif under_outputs and vals:
                fp = Path(vals[0])
                if fp.exists():
                    self._show_output_file_menu(item, fp, event.x_root, event.y_root)
        except Exception as e:
            import traceback
            print(f"[右键菜单异常] tags={tag_list} vals={vals} err={e}")
            traceback.print_exc()

    def _is_under_outputs(self, item: str) -> bool:
        """判断节点是否在 outputs 树下"""
        tree = self.file_tree
        while item:
            if tree.item(item, "text") == "📂 outputs":
                return True
            item = tree.parent(item)
        return False

    # ═══════════ L1: data/视频右键菜单 ═══════════

    def _show_data_video_menu(self, item: str, video_path: Path, x: int, y: int):
        menu = tk.Menu(self.root, tearoff=0, bg=C.CARD, fg=C.TEXT, font=("Microsoft YaHei", 9))
        menu.add_command(label="🎬 打开视频", font=("Microsoft YaHei", 9, "bold"),
                         command=lambda: self._open_video_from_browser(video_path))
        menu.add_separator()
        menu.add_command(label="✏ 重命名",
                         command=lambda: self._rename_file_node(item, video_path))
        menu.add_command(label="⏱ 添加统计时段",
                         command=lambda: self._add_period_dialog(video_path))
        menu.add_separator()
        menu.add_command(label="🗑 删除视频",
                         command=lambda: self._delete_video_node(item, video_path))
        menu.post(x, y)

    def _add_period_dialog(self, video_path: Path):
        """弹出对话框添加统计时段"""
        data = self._load_config_yaml(video_path) or {}
        periods = data.get("periods", {})
        # 生成默认名称
        idx = 1
        while f"时段{idx}" in periods:
            idx += 1

        dialog = tk.Toplevel(self.root)
        dialog.title("添加统计时段")
        dialog.configure(bg=C.CARD)
        dialog.resizable(False, False)
        x = self.root.winfo_x() + 200; y = self.root.winfo_y() + 200
        dialog.geometry(f"350x200+{x}+{y}")

        tk.Label(dialog, text="时段名称:", bg=C.CARD, fg=C.TEXT, font=FONT_BODY).pack(pady=(15, 2))
        name_var = tk.StringVar(value=f"时段{idx}")
        tk.Entry(dialog, textvariable=name_var, bg=C.SURFACE, fg=C.TEXT, font=FONT_MONO_SM,
                 width=20, relief=tk.FLAT, insertbackground=C.ACCENT).pack(pady=2)

        tk.Label(dialog, text="(起止时刻默认为视频全时刻，后续可修改)", bg=C.CARD,
                 fg=C.TEXT_MUTED, font=FONT_TINY).pack(pady=(5, 15))

        btn_f = tk.Frame(dialog, bg=C.CARD); btn_f.pack()
        tk.Button(btn_f, text="确认", bg=C.ACCENT, fg=C.BG, relief=tk.FLAT,
                  font=FONT_BODY, padx=15, cursor="hand2",
                  command=lambda: self._do_add_period(dialog, video_path, name_var.get(), data)
                  ).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_f, text="取消", bg=C.BORDER_LIGHT, fg=C.TEXT, relief=tk.FLAT,
                  font=FONT_BODY, padx=15, cursor="hand2",
                  command=dialog.destroy).pack(side=tk.LEFT, padx=5)

    def _do_add_period(self, dialog, video_path: Path, name: str, data: dict):
        name = name.strip()
        if not name:
            messagebox.showwarning("警告", "名称不能为空"); return
        periods = data.setdefault("periods", {})
        if name in periods:
            messagebox.showwarning("警告", "时段名称已存在"); return
        fps = self.fps if hasattr(self, 'fps') and self.fps > 0 else 30.0
        total = self.total_frames if hasattr(self, 'total_frames') and self.total_frames > 0 else 9999
        # 使用系统当前起终点作为默认时段范围
        sf = getattr(self, 'start_frame', 0) or 0
        ef = getattr(self, 'end_frame', total) or total
        periods[name] = {
            "start": sf, "end": ef,
            "start_time": self._frames_to_time_str(sf, fps),
            "end_time": self._frames_to_time_str(ef, fps),
            "entries": {}
        }
        self._save_config_yaml(video_path, data)
        dialog.destroy()
        self._scan_file_browser()
        self._log(f"已添加统计时段: {name} (帧 {sf}-{ef})")

    def _rename_file_node(self, item: str, path: Path):
        new_name = simpledialog.askstring("重命名", "新文件名:", initialvalue=path.stem)
        if not new_name: return
        new_path = path.with_name(new_name + path.suffix)
        if new_path.exists():
            messagebox.showwarning("警告", "目标文件已存在"); return
        try:
            path.rename(new_path)
            # 同步重命名 YAML 配置
            cfg_path = self._find_config_path(path)
            if cfg_path:
                new_cfg = cfg_path.with_name(new_path.stem.replace(" ", "_").replace(".", "_") + ".yaml")
                if not new_cfg.exists():
                    cfg_path.rename(new_cfg)
            self._scan_file_browser()
            self._log(f"已重命名: {path.name} → {new_path.name}")
        except Exception as e:
            messagebox.showerror("错误", f"重命名失败: {e}")

    def _delete_video_node(self, item: str, video_path: Path):
        if not messagebox.askyesno("确认删除", f"确定要删除视频文件？\n{video_path.name}\n\n此操作不可恢复！"):
            return
        try:
            # 删除 YAML 配置
            for cfg in [self._find_config_path(video_path)]:
                if cfg and cfg.exists():
                    cfg.unlink()
            video_path.unlink()
            if self.video_path and self.video_path == video_path:
                self.video_path = None
            self._scan_file_browser()
            self._log(f"已删除视频: {video_path.name}")
        except Exception as e:
            messagebox.showerror("错误", f"删除失败: {e}")

    # ═══════════ L2: 统计时段右键菜单 ═══════════

    def _show_period_menu(self, item: str, video_path: Path, period_name: str, x: int, y: int):
        menu = tk.Menu(self.root, tearoff=0, bg=C.CARD, fg=C.TEXT, font=("Microsoft YaHei", 9))
        menu.add_command(label="✏ 重命名",
                         command=lambda: self._rename_period(item, video_path, period_name))
        menu.add_command(label="🕐 设置起止时刻",
                         command=lambda: self._set_period_range_dialog(video_path, period_name))
        menu.add_separator()
        menu.add_command(label="➕ 添加进口道",
                         command=lambda: self._add_entry_dialog(video_path, period_name))
        menu.add_command(label="🧹 清空进口道",
                         command=lambda: self._clear_entries(video_path, period_name))
        menu.add_separator()
        menu.add_command(label="🗑 删除时段",
                         command=lambda: self._delete_period(item, video_path, period_name))
        menu.post(x, y)

    def _rename_period(self, item: str, video_path: Path, old_name: str):
        new_name = simpledialog.askstring("重命名时段", "新名称:", initialvalue=old_name)
        if not new_name or new_name == old_name: return
        data = self._load_config_yaml(video_path)
        if not data: return
        periods = data.get("periods", {})
        if new_name in periods:
            messagebox.showwarning("警告", "时段名称已存在"); return
        periods[new_name] = periods.pop(old_name)
        self._save_config_yaml(video_path, data)
        self._scan_file_browser()
        self._log(f"时段已重命名: {old_name} → {new_name}")

    def _set_period_range_dialog(self, video_path: Path, period_name: str):
        data = self._load_config_yaml(video_path)
        if not data: return
        pd = data["periods"].get(period_name)
        if not pd: return
        fps = self.fps if self.fps > 0 else 30.0

        dialog = tk.Toplevel(self.root)
        dialog.title(f"设置时段范围 - {period_name}")
        dialog.configure(bg=C.CARD); dialog.resizable(False, False)
        x = self.root.winfo_x() + 200; y = self.root.winfo_y() + 150
        dialog.geometry(f"400x200+{x}+{y}")

        row1 = tk.Frame(dialog, bg=C.CARD); row1.pack(pady=(15, 5))
        tk.Label(row1, text="起始帧:", bg=C.CARD, fg=C.TEXT, font=FONT_BODY, width=8).pack(side=tk.LEFT)
        start_var = tk.IntVar(value=pd.get("start", 0))
        tk.Entry(row1, textvariable=start_var, bg=C.SURFACE, fg=C.TEXT, font=FONT_MONO_SM,
                 width=10, relief=tk.FLAT).pack(side=tk.LEFT, padx=5)
        tk.Label(row1, text="起始时间:", bg=C.CARD, fg=C.TEXT_MUTED, font=FONT_SMALL).pack(side=tk.LEFT)
        st_label = tk.Label(row1, text=self._frames_to_time_str(pd.get("start", 0), fps),
                            bg=C.CARD, fg=C.ACCENT, font=FONT_SMALL)
        st_label.pack(side=tk.LEFT, padx=5)

        row2 = tk.Frame(dialog, bg=C.CARD); row2.pack(pady=5)
        tk.Label(row2, text="结束帧:", bg=C.CARD, fg=C.TEXT, font=FONT_BODY, width=8).pack(side=tk.LEFT)
        end_var = tk.IntVar(value=pd.get("end", self.total_frames))
        tk.Entry(row2, textvariable=end_var, bg=C.SURFACE, fg=C.TEXT, font=FONT_MONO_SM,
                 width=10, relief=tk.FLAT).pack(side=tk.LEFT, padx=5)
        tk.Label(row2, text="结束时间:", bg=C.CARD, fg=C.TEXT_MUTED, font=FONT_SMALL).pack(side=tk.LEFT)
        et_label = tk.Label(row2, text=self._frames_to_time_str(pd.get("end", self.total_frames), fps),
                            bg=C.CARD, fg=C.ACCENT, font=FONT_SMALL)
        et_label.pack(side=tk.LEFT, padx=5)

        def _update_labels(*_):
            st_label.config(text=self._frames_to_time_str(start_var.get(), fps))
            et_label.config(text=self._frames_to_time_str(end_var.get(), fps))
        start_var.trace_add("write", _update_labels)
        end_var.trace_add("write", _update_labels)

        btn_f = tk.Frame(dialog, bg=C.CARD); btn_f.pack(pady=(10, 15))
        tk.Button(btn_f, text="确认", bg=C.ACCENT, fg=C.BG, relief=tk.FLAT,
                  font=FONT_BODY, padx=15, cursor="hand2",
                  command=lambda: self._do_set_period_range(dialog, video_path, period_name,
                      start_var.get(), end_var.get(), fps)
                  ).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_f, text="取消", bg=C.BORDER_LIGHT, fg=C.TEXT, relief=tk.FLAT,
                  font=FONT_BODY, padx=15, cursor="hand2",
                  command=dialog.destroy).pack(side=tk.LEFT, padx=5)

    def _do_set_period_range(self, dialog, video_path, period_name, start, end, fps):
        data = self._load_config_yaml(video_path)
        if not data: return
        pd = data["periods"][period_name]
        pd["start"] = start
        pd["end"] = end
        pd["start_time"] = self._frames_to_time_str(start, fps)
        pd["end_time"] = self._frames_to_time_str(end, fps)
        self._save_config_yaml(video_path, data)
        dialog.destroy()
        self._scan_file_browser()
        self._log(f"时段 {period_name} 范围已更新: 帧 {start}-{end}")

    def _delete_period(self, item: str, video_path: Path, period_name: str):
        if not messagebox.askyesno("确认删除", f"确定要删除时段 '{period_name}'？"):
            return
        data = self._load_config_yaml(video_path)
        if not data: return
        del data["periods"][period_name]
        if not data["periods"]:
            data["periods"] = {"全时段": {"start": 0, "end": self.total_frames,
                "start_time": "00:00:00", "end_time": self._frames_to_time_str(self.total_frames, 30),
                "entries": {}}}
        self._save_config_yaml(video_path, data)
        self._scan_file_browser()
        self._log(f"已删除时段: {period_name}")

    # ═══════════ L3: 进口道右键菜单 ═══════════

    def _show_entry_menu(self, item: str, video_path: Path, period_name: str,
                          entry_name: str, x: int, y: int):
        menu = tk.Menu(self.root, tearoff=0, bg=C.CARD, fg=C.TEXT, font=("Microsoft YaHei", 9))
        menu.add_command(label="✏ 重命名",
                         command=lambda: self._rename_entry(item, video_path, period_name, entry_name))
        menu.add_separator()
        for label, dir_name in [("➕ 添加直行", "直行"), ("➕ 添加左转", "左转"),
                                 ("➕ 添加右转", "右转"), ("➕ 添加掉头", "掉头")]:
            menu.add_command(label=label,
                command=lambda vp=video_path, pn=period_name, en=entry_name, dn=dir_name:
                    self._add_direction_and_draw(vp, pn, en, dn))
        menu.add_separator()
        menu.add_command(label="🗑 删除进口道",
                         command=lambda: self._delete_entry(item, video_path, period_name, entry_name))
        menu.post(x, y)

    def _add_entry_dialog(self, video_path: Path, period_name: str):
        name = simpledialog.askstring("添加进口道", "进口道名称:")
        if not name: return
        data = self._load_config_yaml(video_path) or {"periods": {}}
        periods = data.setdefault("periods", {})
        fps = self.fps if self.fps > 0 else 30.0
        total = self.total_frames if self.total_frames > 0 else 9999
        sf = getattr(self, 'start_frame', 0) or 0
        ef = getattr(self, 'end_frame', total) or total
        if period_name not in periods:
            periods[period_name] = {
                "start": sf, "end": ef,
                "start_time": self._frames_to_time_str(sf, fps),
                "end_time": self._frames_to_time_str(ef, fps),
                "entries": {}
            }
        entries = periods[period_name].setdefault("entries", {})
        if name in entries:
            messagebox.showwarning("警告", "进口道名称已存在"); return
        entries[name] = {"lines": []}
        self._save_config_yaml(video_path, data)
        self._scan_file_browser()
        self._log(f"已添加进口道: {name} (时段: {period_name})")

    def _clear_entries(self, video_path: Path, period_name: str):
        if not messagebox.askyesno("确认", f"确定要清空时段 '{period_name}' 下所有进口道？"):
            return
        data = self._load_config_yaml(video_path)
        if not data: return
        data["periods"][period_name]["entries"] = {}
        self._save_config_yaml(video_path, data)
        self._scan_file_browser()
        self._log(f"已清空时段 {period_name} 的所有进口道")

    def _rename_entry(self, item, video_path, period_name, old_name):
        new_name = simpledialog.askstring("重命名进口道", "新名称:", initialvalue=old_name)
        if not new_name or new_name == old_name: return
        data = self._load_config_yaml(video_path)
        if not data: return
        entries = data["periods"][period_name]["entries"]
        if new_name in entries:
            messagebox.showwarning("警告", "进口道名称已存在"); return
        entries[new_name] = entries.pop(old_name)
        self._save_config_yaml(video_path, data)
        self._scan_file_browser()
        self._log(f"进口道已重命名: {old_name} → {new_name}")

    def _delete_entry(self, item, video_path, period_name, entry_name):
        if not messagebox.askyesno("确认删除", f"确定要删除进口道 '{entry_name}'？"):
            return
        data = self._load_config_yaml(video_path)
        if not data: return
        del data["periods"][period_name]["entries"][entry_name]
        self._save_config_yaml(video_path, data)
        self._scan_file_browser()
        self._log(f"已删除进口道: {entry_name}")

    # ═══════════ L4: 方向截面右键菜单 ═══════════

    def _show_direction_menu(self, item: str, video_path: Path, period_name: str,
                              entry_name: str, dir_name: str, line_id: str, x: int, y: int):
        menu = tk.Menu(self.root, tearoff=0, bg=C.CARD, fg=C.TEXT, font=("Microsoft YaHei", 9))
        menu.add_command(label="✏ 修改截面",
                         command=lambda: self._modify_direction(
                             video_path, period_name, entry_name, dir_name, line_id))
        menu.add_separator()
        menu.add_command(label="🗑 删除截面",
                         command=lambda: self._delete_direction(
                             video_path, period_name, entry_name, line_id))
        menu.post(x, y)

    # ═══════════ L5: outputs 文件右键菜单 ═══════════

    def _show_output_file_menu(self, item: str, file_path: Path, x: int, y: int):
        menu = tk.Menu(self.root, tearoff=0, bg=C.CARD, fg=C.TEXT, font=("Microsoft YaHei", 9))
        menu.add_command(label="✏ 重命名",
                         command=lambda: self._rename_file_node(item, file_path))
        menu.add_separator()
        menu.add_command(label="🗑 删除",
                         command=lambda: self._delete_output_file(item, file_path))
        menu.post(x, y)

    def _delete_output_file(self, item: str, file_path: Path):
        if not messagebox.askyesno("确认删除", f"确定要删除文件？\n{file_path.name}"):
            return
        try:
            file_path.unlink()
            self._scan_file_browser()
            self._log(f"已删除: {file_path.name}")
        except Exception as e:
            messagebox.showerror("错误", f"删除失败: {e}")

    # ═══════════ 视频加载 ═══════════

    def _open_video_from_browser(self, path: Path):
        """从文件浏览器打开视频"""
        if not path.exists():
            self._log(f"文件不存在: {path}")
            return
        self._stop_playback()
        self.video_path = path
        self._log(f"打开视频: {path.name}")
        self._load_video()

    def _load_browser_config(self, video_path: Path):
        """加载浏览器选中视频的截面线配置"""
        lines = load_cs_lines_config(video_path, PROJECT_ROOT)
        if lines:
            self.count_lines = lines
            self._refresh_tree()
            self._log(f"已加载 {len(lines)} 条截面线配置")
        else:
            self._log("未找到该视频的截面线配置")

    # ═══════════ 画线与编辑模式 ═══════════

    def _add_direction_and_draw(self, video_path: Path, period_name: str,
                                 entry_name: str, direction: str):
        """添加转向方向 → 进入画线模式"""
        if not self.video_path or self.video_path.resolve() != video_path.resolve():
            self._open_video_from_browser(video_path)
            self.root.after(500, lambda: self._try_enter_draw_mode(
                video_path, period_name, entry_name, direction))
        else:
            self._enter_draw_mode(video_path, period_name, entry_name, direction)

    def _try_enter_draw_mode(self, video_path, period_name, entry_name, direction, retry=0):
        """确认视频加载完成后进入画线模式（带重试）"""
        if not self.video_label:
            if retry < 10:
                self.root.after(200, lambda: self._try_enter_draw_mode(
                    video_path, period_name, entry_name, direction, retry + 1))
            return
        self._enter_draw_mode(video_path, period_name, entry_name, direction)

    def _enter_draw_mode(self, video_path: Path, period_name: str,
                          entry_name: str, direction: str):
        """进入画线模式：单击起点→移动预览→单击终点→Enter/双击确认"""
        if not self.video_label:
            print("[DrawMode] video_label is None, aborting")
            return
        # 停止播放
        self._pause()
        self._send_command(PlayerCommand.STOP)
        # 清理旧绑定（不刷新显示，避免触发 resize 导致画面缩放）
        self._unbind_draw_events()

        self._draw_state = {
            "video_path": str(video_path), "period_name": period_name,
            "entry_name": entry_name, "direction": direction,
            "x1": None, "y1": None, "x2": None, "y2": None,
            "phase": "start",  # start → draw → confirm
        }

        # 显示提示
        if not hasattr(self, 'lbl_draw_hint') or not self.lbl_draw_hint:
            self.lbl_draw_hint = tk.Label(
                self.root, text="", bg=C.ACCENT, fg=C.BG,
                font=("Microsoft YaHei", 10), padx=10, pady=4
            )
        self.lbl_draw_hint.config(
            text=f"画线: {entry_name}→{direction} | 单击起点 → 移动 → 单击终点 | Enter确认 / Esc取消")
        self.lbl_draw_hint.place(relx=0.5, y=10, anchor="n")

        self.video_label.bind("<Button-1>", self._on_draw_canvas_click)
        self.video_label.bind("<Motion>", self._on_draw_canvas_move)
        self.root.bind("<Return>", self._on_draw_confirm)
        self.root.bind("<Double-1>", self._on_draw_confirm)
        self.root.bind("<Escape>", self._on_draw_cancel)

    def _unbind_draw_events(self):
        """仅解绑画线事件，不刷新显示（避免触发帧 resize）"""
        try:
            self.video_label.unbind("<Button-1>")
            self.video_label.unbind("<Motion>")
        except Exception:
            pass
        self.root.unbind("<Return>")
        self.root.unbind("<Double-1>")
        self.root.unbind("<Escape>")

    def _cancel_draw_mode(self):
        """取消画线/编辑模式，恢复绑定 + 刷新帧显示"""
        self._unbind_draw_events()
        if hasattr(self, 'lbl_draw_hint') and self.lbl_draw_hint:
            self.lbl_draw_hint.place_forget()
        self._draw_state = None
        # 恢复帧显示
        self._seek_absolute(self.current_frame)

    def _on_draw_canvas_click(self, event):
        ds = getattr(self, '_draw_state', None)
        if not ds: return
        # 转换鼠标坐标到视频坐标
        vx, vy = self._mouse_to_video_coords(event.x, event.y)
        if vx < 0 or vy < 0: return
        if ds["phase"] == "start":
            ds["x1"], ds["y1"] = vx, vy
            ds["phase"] = "draw"
            self.lbl_draw_hint.config(text=f"画线: {ds['entry_name']}→{ds['direction']} | 起点已定，单击设置终点 | Enter确认 / Esc取消")
        elif ds["phase"] == "draw":
            ds["x2"], ds["y2"] = vx, vy
            ds["phase"] = "confirm"
            self.lbl_draw_hint.config(text=f"画线: {ds['entry_name']}→{ds['direction']} | 终点已定 | Enter确认 / Esc取消 / 单击继续画")
            self._seek_absolute(self.current_frame)  # 刷新显示

    def _on_draw_canvas_move(self, event):
        ds = getattr(self, '_draw_state', None)
        if not ds or ds["phase"] != "draw": return
        vx, vy = self._mouse_to_video_coords(event.x, event.y)
        if vx < 0 or vy < 0: return
        # 实时预览：在保持显示尺寸的前提下画临时线
        frame = self._render_current_frame_copy()
        if frame is not None:
            import cv2
            h, w = frame.shape[:2]
            # 计算当前显示尺寸（同 _render_ui_frame）
            display_w, display_h = 920, 540
            ratio = min(display_w / w, display_h / h)
            if ratio < 1:
                frame = cv2.resize(frame, (int(w * ratio), int(h * ratio)))
            # 在缩放后的帧上画线
            x1, y1 = int(ds["x1"] * ratio), int(ds["y1"] * ratio)
            x2, y2 = int(vx * ratio), int(vy * ratio)
            cv2.line(frame, (x1, y1), (x2, y2), (0, 255, 255), 2)
            rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            img = ImageTk.PhotoImage(Image.fromarray(rgb))
            self.video_label.configure(image=img)
            self.video_label._preview_image = img

    def _mouse_to_video_coords(self, label_x: int, label_y: int) -> tuple:
        """Label 坐标 → 原始视频坐标"""
        if not hasattr(self, '_video_width') or not self._video_width:
            return (-1, -1)
        label_w = self.video_label.winfo_width() or 1
        label_h = self.video_label.winfo_height() or 1
        ratio = min(label_w / self._video_width, label_h / self._video_height)
        disp_w = int(self._video_width * ratio)
        disp_h = int(self._video_height * ratio)
        offset_x = (label_w - disp_w) // 2
        offset_y = (label_h - disp_h) // 2
        vx = int((label_x - offset_x) / ratio)
        vy = int((label_y - offset_y) / ratio)
        return (vx, vy)

    def _video_to_label_coords(self, vx: int, vy: int) -> tuple:
        """原始视频坐标 → Label 坐标"""
        label_w = self.video_label.winfo_width() or 1
        label_h = self.video_label.winfo_height() or 1
        ratio = min(label_w / self._video_width, label_h / self._video_height)
        disp_w = int(self._video_width * ratio)
        disp_h = int(self._video_height * ratio)
        offset_x = (label_w - disp_w) // 2
        offset_y = (label_h - disp_h) // 2
        return (int(offset_x + vx * ratio), int(offset_y + vy * ratio))

    def _render_current_frame_copy(self) -> np.ndarray | None:
        """获取当前帧的副本用于预览"""
        import cv2
        if self.current_frame < 0 or not self.video_path:
            cap = cv2.VideoCapture(str(self.video_path))
            if cap.isOpened():
                cap.set(cv2.CAP_PROP_POS_FRAMES, max(0, self.current_frame))
                ret, frame = cap.read()
                cap.release()
                if ret: return frame
            return None
        frame_data = self.frame_buffer.get_cached(self.current_frame)
        if frame_data is not None:
            return frame_data.copy()
        return None

    def _on_draw_confirm(self, event=None):
        ds = getattr(self, '_draw_state', None)
        if not ds or "x2" not in ds or ds["x2"] is None: return
        self._save_drawn_line(ds)
        self._cancel_draw_mode()

    def _on_draw_cancel(self, event=None):
        self._cancel_draw_mode()
        self._log("画线已取消")

    def _save_drawn_line(self, ds: dict):
        """将画的线保存到 YAML"""
        import uuid
        video_path = Path(ds["video_path"])
        data = self._load_config_yaml(video_path) or {"periods": {}}
        periods = data.setdefault("periods", {})
        pd = periods.setdefault(ds["period_name"], {
            "start": 0, "end": self.total_frames,
            "start_time": "00:00:00",
            "end_time": self._frames_to_time_str(self.total_frames, 30),
            "entries": {}
        })
        entries = pd.setdefault("entries", {})
        entry = entries.setdefault(ds["entry_name"], {"lines": []})
        new_line = {
            "line_id": f"line_{uuid.uuid4().hex[:8]}",
            "entry_name": ds["entry_name"],
            "direction": ds["direction"],
            "x1": ds["x1"], "y1": ds["y1"],
            "x2": ds["x2"], "y2": ds["y2"],
            "is_drawn": True,
        }
        entry["lines"].append(new_line)
        self._save_config_yaml(video_path, data)
        self._scan_file_browser()
        self._log(f"已保存截面线: {ds['entry_name']}→{ds['direction']} ({ds['x1']},{ds['y1']})→({ds['x2']},{ds['y2']})")

    def _delete_direction(self, video_path: Path, period_name: str,
                           entry_name: str, line_id: str):
        """从 YAML 删除指定截面线"""
        if not line_id: return
        if not messagebox.askyesno("确认删除", f"确定要删除截面线？\n{entry_name} → {line_id}"):
            return
        data = self._load_config_yaml(video_path)
        if not data: return
        lines = data["periods"][period_name]["entries"][entry_name].get("lines", [])
        lines[:] = [l for l in lines if l.get("line_id") != line_id]
        self._save_config_yaml(video_path, data)
        self._scan_file_browser()
        self._log(f"已删除截面线: {line_id}")

    def _modify_direction(self, video_path: Path, period_name: str,
                            entry_name: str, dir_name: str, line_id: str):
        """进入截面修改模式：高亮截面线，可拖拽端点"""
        if not self.video_path or self.video_path.resolve() != video_path.resolve():
            self._open_video_from_browser(video_path)
            self.root.after(300, lambda: self._enter_edit_mode(
                video_path, period_name, entry_name, dir_name, line_id))
        else:
            self._enter_edit_mode(video_path, period_name, entry_name, dir_name, line_id)

    def _enter_edit_mode(self, video_path: Path, period_name: str,
                           entry_name: str, dir_name: str, line_id: str):
        """进入编辑模式"""
        data = self._load_config_yaml(video_path)
        if not data: return
        # 查找 lines
        lines = data["periods"][period_name]["entries"][entry_name].get("lines", [])
        target = None
        for l in lines:
            if l.get("line_id") == line_id:
                target = l
                break
        if not target:
            self._log("未找到截面线"); return

        self._cancel_draw_mode()
        self._edit_state = {
            "video_path": str(video_path), "period_name": period_name,
            "entry_name": entry_name, "line_id": line_id,
            "x1": target["x1"], "y1": target["y1"],
            "x2": target["x2"], "y2": target["y2"],
            "dragging": None,  # "start" or "end"
        }

        self.lbl_draw_hint.config(
            text=f"编辑截面线: {entry_name}→{dir_name} | 拖拽端点修改 | 点击空白/Enter保存 | Esc取消")
        self.lbl_draw_hint.place(relx=0.5, y=10, anchor="n")

        self.video_label.bind("<Button-1>", self._on_edit_click)
        self.video_label.bind("<B1-Motion>", self._on_edit_drag)
        self.video_label.bind("<ButtonRelease-1>", self._on_edit_release)
        self.root.bind("<Return>", self._on_edit_confirm)
        self.root.bind("<Escape>", self._on_edit_cancel)

        # 刷新显示高亮的线
        self._seek_absolute(self.current_frame)

    def _on_edit_click(self, event):
        es = getattr(self, '_edit_state', None)
        if not es: return
        vx, vy = self._mouse_to_video_coords(event.x, event.y)
        if vx < 0 or vy < 0:
            # 点击了视频外部 → 确认保存
            self._on_edit_confirm()
            return
        # 判断是否点中了端点（阈值10像素）
        d1 = ((vx - es["x1"])**2 + (vy - es["y1"])**2)**0.5
        d2 = ((vx - es["x2"])**2 + (vy - es["y2"])**2)**0.5
        if d1 < 15:
            es["dragging"] = "start"
        elif d2 < 15:
            es["dragging"] = "end"
        else:
            # 点击其他位置 → 确认保存
            self._on_edit_confirm()

    def _on_edit_drag(self, event):
        es = getattr(self, '_edit_state', None)
        if not es or not es["dragging"]: return
        vx, vy = self._mouse_to_video_coords(event.x, event.y)
        if vx < 0 or vy < 0: return
        if es["dragging"] == "start":
            es["x1"], es["y1"] = vx, vy
        else:
            es["x2"], es["y2"] = vx, vy
        # 实时刷新
        self._seek_absolute(self.current_frame)

    def _on_edit_release(self, event):
        es = getattr(self, '_edit_state', None)
        if es:
            es["dragging"] = None

    def _on_edit_confirm(self, event=None):
        es = getattr(self, '_edit_state', None)
        if not es: return
        # 保存到 YAML
        data = self._load_config_yaml(Path(es["video_path"]))
        if data:
            lines = data["periods"][es["period_name"]]["entries"][es["entry_name"]].get("lines", [])
            for l in lines:
                if l.get("line_id") == es["line_id"]:
                    l["x1"] = es["x1"]; l["y1"] = es["y1"]
                    l["x2"] = es["x2"]; l["y2"] = es["y2"]
                    break
            self._save_config_yaml(Path(es["video_path"]), data)
            self._log(f"截面线已更新: {es['entry_name']}")
        self._cancel_edit_mode()

    def _on_edit_cancel(self, event=None):
        self._cancel_edit_mode()
        self._log("截面线修改已取消")

    def _cancel_edit_mode(self):
        self.video_label.unbind("<Button-1>")
        self.video_label.unbind("<B1-Motion>")
        self.video_label.unbind("<ButtonRelease-1>")
        self.root.unbind("<Return>")
        self.root.unbind("<Escape>")
        if hasattr(self, 'lbl_draw_hint') and self.lbl_draw_hint:
            self.lbl_draw_hint.place_forget()
        self._edit_state = None
        self._seek_absolute(self.current_frame)
        """从文件浏览器打开视频"""
        if not path.exists():
            self._log(f"文件不存在: {path}")
            return
        self._stop_playback()
        self.video_path = path
        self._log(f"打开视频: {path.name}")
        self._load_video()

    def _load_browser_config(self, video_path: Path):
        """加载浏览器选中视频的截面线配置"""
        lines = load_cs_lines_config(video_path, PROJECT_ROOT)
        if lines:
            self.count_lines = lines
            self._refresh_tree()
            self._log(f"已加载 {len(lines)} 条截面线配置")
        else:
            self._log("未找到该视频的截面线配置")

    def _stop_playback(self):
        """停止播放"""
        self._pause()
        self._send_command(PlayerCommand.STOP)
        self.frame_buffer.clear()

    # ═════════════════════════════════════════
    # 核心播放控制逻辑（线程安全）
    # ═════════════════════════════════════════

    def _send_command(self, cmd_type: PlayerCommand, value=None):
        """向播放线程发送命令（线程安全）"""
        try:
            self.command_queue.put_nowait(Command(cmd_type, value))
        except Full:
            print(f"[UI] 命令队列已满，无法发送: {cmd_type}")

    def toggle_play(self):
        """切换播放/暂停 - 通过命令队列与后台线程通信"""
        if not self.video_path:
            messagebox.showwarning("警告", "请先选择视频文件")
            return

        if not self.player_thread or not self.player_thread.is_alive():
            self._start_player()
            return

        self.is_playing = not self.is_playing
        if self.is_playing:
            # 播放前确保从有效位置开始
            if self.current_frame < self.start_frame or self.current_frame >= self.end_frame:
                self._seek_absolute(self.start_frame)
            self.play_btn.config(text="❚❚", bg=C.GREEN, fg=C.BG)
            self._send_command(PlayerCommand.PLAY)
            self._log("播放视频...")
            self._start_display_loop()
        else:
            self._pause()

    def _start_player(self):
        """启动视频播放线程"""
        if self.player_thread and self.player_thread.is_alive():
            return

        self.frame_buffer.clear()
        # 清空命令队列
        while not self.command_queue.empty():
            try:
                self.command_queue.get_nowait()
            except Empty:
                break

        self.player_thread = VideoPlayerThread(
            self.video_path,
            self.frame_buffer,
            self.command_queue,
            self.fps,
            start_frame=self.start_frame,
            end_frame=self.end_frame,
        )
        self.player_thread.start()
        
        # 等待线程初始化完成
        time.sleep(0.2)
        
        self.total_frames = self.player_thread.get_total_frames()
        self.timeline.configure(
            min_val=0,
            max_val=max(1, self.total_frames - 1),
            start_value=self.start_frame,
            end_value=self.end_frame,
        )
        
        self.is_playing = True
        self.play_btn.config(text="❚❚", bg=C.GREEN, fg=C.BG)
        self._send_command(PlayerCommand.PLAY)
        # 同步当前倍速到新线程
        try:
            cur_speed = float(self.speed_var.get().replace("x", ""))
            self._send_command(PlayerCommand.SPEED, cur_speed)
        except ValueError:
            pass
        self._start_display_loop()

    def _start_display_loop(self):
        """启动显示循环 - 主线程定时器"""
        self._display_frame()

    def _render_ui_frame(self, frame: np.ndarray) -> ImageTk.PhotoImage:
        """将原始帧绘制截面线后，缩放并转为 PhotoImage"""
        annotated = self._draw_lines_on_frame(frame)

        # 编辑模式：高亮正在修改的截面线 + 拖拽手柄
        es = getattr(self, '_edit_state', None)
        if es and es.get("x1"):
            cv2.line(annotated, (es["x1"], es["y1"]), (es["x2"], es["y2"]), (0, 255, 255), 3)
            cv2.circle(annotated, (es["x1"], es["y1"]), 10, (0, 255, 255), -1)
            cv2.circle(annotated, (es["x2"], es["y2"]), 10, (0, 0, 255), -1)
            cv2.circle(annotated, (es["x1"], es["y1"]), 12, (0, 200, 255), 2)
            cv2.circle(annotated, (es["x2"], es["y2"]), 12, (0, 0, 200), 2)

        h, w = annotated.shape[:2]
        display_size = (920, 540)
        ratio = min(display_size[0] / w, display_size[1] / h)
        if ratio < 1 and ratio > 0:
            new_w, new_h = int(w * ratio), int(h * ratio)
            annotated = cv2.resize(annotated, (new_w, new_h), interpolation=cv2.INTER_LINEAR)
        rgb = cv2.cvtColor(annotated, cv2.COLOR_BGR2RGB)
        image = Image.fromarray(rgb)
        return ImageTk.PhotoImage(image)

    def _display_frame(self):
        """显示循环 - 从缓冲区获取帧并更新UI"""
        if not self.is_playing:
            return

        frame_data = self.frame_buffer.get()

        if frame_data:
            frame_num, frame = frame_data
            self.current_frame = frame_num
            self.current_photo = self._render_ui_frame(frame)

            # 检查是否达到终点范围
            if self.current_frame >= self.end_frame:
                self._pause()
                if self.play_timer:
                    self.root.after_cancel(self.play_timer)
                    self.play_timer = None
                self.frame_buffer.clear()
                self._seek_absolute(self.start_frame)
                return

            # 更新UI（主线程安全）
            self.video_label.configure(image=self.current_photo, text="")
            self._update_frame_info()
            self.timeline.set_playhead(frame_num)
        else:
            # 无新帧，检查线程状态
            if self.player_thread and not self.player_thread.is_alive():
                self._pause()
                return

        # 根据当前播放倍数动态调整刷新间隔
        speed = getattr(self.player_thread, '_speed', 1.0) if self.player_thread else 1.0
        delay = max(8, int(1000 / max(1.0, self.fps * speed)))
        self.play_timer = self.root.after(delay, self._display_frame)

    def _display_frame_image(self, frame_num: int, frame: np.ndarray):
        """显示指定的帧图像"""
        self.current_frame = frame_num
        self.current_photo = self._render_ui_frame(frame)
        self.video_label.configure(image=self.current_photo, text="")
        self._update_frame_info()
        self.timeline.set_playhead(frame_num)

    def _pause(self):
        """暂停播放"""
        self.is_playing = False
        self._send_command(PlayerCommand.PAUSE)
        if self.play_timer:
            self.root.after_cancel(self.play_timer)
            self.play_timer = None
        self.play_btn.config(text="▶", bg=C.ACCENT, fg=C.BG)

    def stop_playback(self):
        """停止播放并回到起帧位置"""
        self._pause()
        self.frame_buffer.clear()  # 清除残留帧，避免重播时从旧位置开始
        self._send_command(PlayerCommand.STOP)  # STOP 包含 seek 到 start_frame
        self.current_frame = self.start_frame
        self.timeline.set_playhead(self.start_frame)
        self._update_frame_info()

    def _seek_absolute(self, frame_num: int):
        """绝对跳转（限制在起终点范围内）"""
        if not self.player_thread:
            return
        # 先限制在视频总帧数内，再限制在起终点范围内
        frame_num = max(0, min(frame_num, self.total_frames - 1))
        frame_num = max(self.start_frame, min(frame_num, self.end_frame))
        self._send_command(PlayerCommand.SEEK, frame_num)
        # 立即更新UI显示
        self.current_frame = frame_num
        self.timeline.set_playhead(frame_num)
        self._update_frame_info()
        # 启动显示循环，让跳转后的帧能显示到界面上
        self._poll_seek_frame(frame_num)

    def _poll_seek_frame(self, target_frame: int, retry: int = 0):
        """轮询帧缓冲区并显示seek后的帧"""
        if retry > 10:  # 最多轮询10次（约500ms）
            return
        frame_data = self.frame_buffer.get()
        if frame_data:
            frame_num, frame = frame_data
            self.current_frame = frame_num
            self.current_photo = self._render_ui_frame(frame)
            self.video_label.configure(image=self.current_photo, text="")
            self._update_frame_info()
            self.timeline.set_playhead(frame_num)
        else:
            # 还没加载完成，继续等待
            self.root.after(50, self._poll_seek_frame, target_frame, retry + 1)

    def _seek_relative(self, offset: int):
        """相对跳转"""
        target = self.current_frame + offset
        self._seek_absolute(target)

    def _on_speed_changed(self, event=None):
        """速度变化"""
        speed_str = self.speed_var.get().replace("x", "")
        try:
            speed = float(speed_str)
            self._send_command(PlayerCommand.SPEED, speed)
        except ValueError:
            pass

    # ═════════════════════════════════════════
    # 时间轴回调
    # ═════════════════════════════════════════

    def _on_timeline_changed(self, action: str, *args, seek_frame: bool = False):
        """时间轴变化回调"""
        if action == "range":
            self.start_frame, self.end_frame = args
            # 同步到播放线程（线程内部有独立副本，需实时更新）
            if self.player_thread and self.player_thread.is_alive():
                self.player_thread._start_frame = self.start_frame
                self.player_thread._end_frame = self.end_frame
            if seek_frame:
                # 根据最近拖动的手柄跳转
                if hasattr(self, '_last_dragged_handle') and self._last_dragged_handle == "end":
                    self._seek_absolute(self.end_frame)
                else:
                    self._seek_absolute(self.start_frame)
        elif action == "seek":
            self._seek_absolute(args[0])

    # ═════════════════════════════════════════
    # 视频加载与选择
    # ═════════════════════════════════════════

    def select_video(self):
        """选择视频"""
        if self.is_detecting:
            messagebox.showwarning("警告", "检测正在进行中，请先停止")
            return

        path = filedialog.askopenfilename(
            title="选择视频文件",
            filetypes=[
                ("视频文件", "*.mp4 *.avi *.mov *.mkv"),
                ("所有文件", "*.*")
            ]
        )
        if path:
            self.video_path = Path(path)
            self._log(f"打开视频: {self.video_path.name}")
            self._load_video()

    def _load_video(self):
        """加载视频 - 启动播放线程获取信息"""
        if not self.video_path:
            return

        # 停止现有播放
        if self.player_thread:
            self._send_command(PlayerCommand.EXIT)
            self.player_thread.stop()
            self.player_thread.join(timeout=1.0)
            self.player_thread = None

        # 彻底重置所有视频相关状态
        self.start_frame = 0
        self.end_frame = 0
        self.current_frame = 0
        self.total_frames = 0
        self._detect_stats = {}
        self._detect_class_labels = {}
        self._detect_model_name = ""
        self._traffic_stats = None
        # 重置检测状态
        self.is_detecting = False
        self.is_playing = False
        if self.play_timer:
            self.root.after_cancel(self.play_timer)
            self.play_timer = None

        # 先获取视频总帧数，确定起终点范围（按新视频长度重置）
        cap = cv2.VideoCapture(str(self.video_path))
        if cap.isOpened():
            self.total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
            self.fps = cap.get(cv2.CAP_PROP_FPS) or 30.0
            self._video_width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
            self._video_height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
            cap.release()
        else:
            self._log("错误: 无法打开视频文件")
            return
        self.start_frame = 0
        self.end_frame = max(0, self.total_frames - 1)

        # 显示第一帧
        self._show_first_frame()
        
        self._pause()
        self.frame_buffer.clear()

        # 启动新线程（此时 start/end 已经是新视频的正确范围）
        self._start_player()
        self._pause()  # 加载后暂停，等待用户操作

        self._log(f"视频已加载: {self.video_path.name} ({self._video_width}×{self._video_height})")
        self._log(f"总帧数: {self.total_frames}")
        self._log(f"检测范围: 帧 0 - {self.total_frames - 1}")

        # 自动加载该视频的截面线配置
        lines = load_cs_lines_config(self.video_path, PROJECT_ROOT)
        if lines:
            self.count_lines = lines
            self._refresh_tree()
            self._log(f"已自动加载截面线配置，共 {len(lines)} 条线")
            # 预加载有画线后，默认勾选启用流量统计
            drawn_lines = [cl for cl in self.count_lines if cl.is_drawn]
            if drawn_lines:
                self.var_traffic_count.set(True)
                self._log("已自动勾选「检测时启用流量统计」")
        else:
            self.count_lines.clear()
            self._refresh_tree()
            self._entry_counter = 0
            self.var_traffic_count.set(False)

        self._scan_file_browser()

    def _show_first_frame(self):
        """显示视频第一帧"""
        if not self.video_path:
            return
        
        cap = cv2.VideoCapture(str(self.video_path))
        if cap.isOpened():
            ret, frame = cap.read()
            if ret:
                self._display_frame_image(0, frame)
        cap.release()

    # ═════════════════════════════════════════
    # 检测功能
    # ═════════════════════════════════════════

    def start_detection(self):
        """开始检测"""
        if not self.video_path:
            messagebox.showwarning("警告", "请先选择视频文件")
            return

        # 流量统计校验
        if self.var_traffic_count.get():
            drawn_lines = [cl for cl in self.count_lines if cl.is_drawn]
            if not drawn_lines:
                messagebox.showwarning("警告", "已启用流量统计，但未配置任何截面线\n请在「流量统计设置」中添加进口道和转向方向并画线")
                return

        self._save_config()

        self.is_detecting = True
        self._pause()
        # 视频指针同步到起点帧（支持时段/自定义范围）
        self._seek_absolute(self.start_frame)

        self.btn_start.config(state="disabled", bg=C.BTN_DISABLED, fg=C.BTN_DISABLED_TEXT)
        self.btn_stop.config(state="normal", bg=C.RED, fg=C.BG)
        self.btn_preview.config(state="disabled")
        self.btn_select_video.config(state="disabled", bg=C.BTN_DISABLED, fg=C.BTN_DISABLED_TEXT)
        pass  # 状态指示器已移除

        self._update_stage_label("准备中...")
        self._log("=" * 60)
        self._log("【无人机航拍视频检测系统】")
        self._log(f"视频: {self.video_path.name}")
        self._log(f"检测范围: 帧 {self.start_frame} - {self.end_frame}")

        threading.Thread(target=self._run_detection_streaming, daemon=True).start()

    def _run_detection_streaming(self):
        """流式检测 - 委托 predict_video.run_inference 处理"""
        try:
            from src import predict_video as _pv

            model_select = self.model_select_var.get()
            device = resolve_device(self.device_var.get())
            model_config = _pv.MODEL_OPTIONS[model_select]
            model_path = PROJECT_ROOT / model_config["path"]
            model_type = model_config["type"]
            model_name = model_config["name"]

            if not model_path.exists():
                raise FileNotFoundError(f"模型文件不存在: {model_path}")

            model_class_config = MODEL_CLASS_CONFIGS[model_select]
            target_classes = model_class_config["target_classes"]
            class_labels_zh = model_class_config["labels"]

            output_dir = PROJECT_ROOT / "outputs"
            output_dir.mkdir(parents=True, exist_ok=True)
            output_video = output_dir / f"detect_{model_name}_{int(time.time())}.mp4"

            self.root.after(0, self._log, f"使用模型: {model_name}")
            # 准备流量统计截面线
            _count_lines = None
            if self.var_traffic_count.get():
                _count_lines = [cl for cl in self.count_lines if cl.is_drawn]
            # 跟踪启用条件：用户勾选轨迹跟踪 AND (画线统计或自动流量统计需要跟踪)
            need_tracking_for_flow = bool(_count_lines) or self.var_auto_traffic.get()
            _enable_tracking = self.var_enable_tracking.get() and need_tracking_for_flow
            _enable_detection = self.var_enable_detection.get()

            # 判断是否有任何输出选项被选中
            _has_output = (self.var_write_output.get() or self.var_export_traffic.get() or
                           self.var_export_trajectory_img.get() or self.var_export_trajectory_data.get())

            self.root.after(0, self._log, f"设备: {device}")
            self.root.after(0, self._log,
                f"检测范围: 帧 {self.start_frame} - {self.end_frame}")
            self.root.after(0, self._log,
                f"目标检测: {'启用' if _enable_detection else '关闭（仅过帧）'}")
            self.root.after(0, self._log,
                f"SAHI切片: {'启用' if self.var_use_sahi.get() else '关闭（整帧推理）'}")
            self.root.after(0, self._log,
                f"轨迹跟踪: {'启用' if _enable_tracking else '关闭'}")
            self.root.after(0, self._log,
                f"视频输出: {'启用' if self.var_write_output.get() else '关闭'}")
            self.root.after(0, self._log,
                f"车道边界检测: {'启用' if self.var_enable_road_detect.get() else '关闭'}")
            if not _has_output:
                self.root.after(0, self._log, "⚠ 无输出选项选中，仅执行检测不生成文件")
            if _count_lines:
                self.root.after(0, self._log, f"流量统计: 启用，共 {len(_count_lines)} 条截面线")

            self.root.after(0, self._update_stage_label, "阶段1: 初始化模型")
            self.root.after(0, self._log, "[1/4] 初始化检测模型...")
            self.root.after(0, self._update_progress_bar, 2)
            # 保存检测参数供 UI 预览使用
            self._detect_class_labels = class_labels_zh
            self._detect_model_name = model_name

            # 调用 predict_video 的核心推理
            _detect_start_time = time.time()  # 记录检测开始时间
            _stop_event = threading.Event()
            self._stop_event = _stop_event   # 保存引用供 stop_detection 使用
            result = _pv.run_inference(
                model_select=model_select,
                model_path=model_path,
                model_type=model_type,
                model_name=model_name,
                input_video=self.video_path,
                output_video=output_video,
                device=device,
                confidence_threshold=self.confidence_threshold,
                target_classes=target_classes,
                class_labels_zh=class_labels_zh,
                slice_height=self.slice_h_var.get(),
                slice_width=self.slice_w_var.get(),
                overlap_height_ratio=self.overlap_h_var.get(),
                overlap_width_ratio=self.overlap_w_var.get(),
                detect_skip_interval=self.skip_interval,
                start_frame=self.start_frame,
                end_frame=self.end_frame,
                label_font_size=18,
                enable_tracking=_enable_tracking,
                count_lines=_count_lines,
                show_labels=self.var_show_count_lines.get(),
                use_sahi=self.var_use_sahi.get(),
                write_output=self.var_write_output.get(),
                enable_road_detect=self.var_enable_road_detect.get(),
                road_model_type=self.road_model_var.get(),
                enable_detection=_enable_detection,
                enable_auto_traffic=self.var_auto_traffic.get(),
                show_trajectory=self.var_show_trajectory.get(),
                stop_event=_stop_event,
                log_callback=lambda msg: self.root.after(0, self._log, msg),
                progress_callback=lambda cur, total, frame, dcount, stats, tstats:
                    self._on_detect_progress(cur, total, frame, dcount, stats, tstats, model_name),
            )
            frame_count = result["frame_count"]
            traffic_stats = result.get("traffic_stats")
            auto_flows = result.get("auto_flows")
            auto_traffic_counter = result.get("auto_traffic_counter")

            # 检测完成
            self.root.after(0, self._update_stage_label, "阶段4: 生成报告")
            self.root.after(0, self._log, "[4/4] 生成检测报告...")
            self.root.after(0, self._update_progress_bar, 98)

            self.root.after(0, self._log, "=" * 60)
            self.root.after(0, self._log, "【检测结果】")
            self.root.after(0, self._log, f"处理帧数: {frame_count}")
            if self.var_write_output.get():
                self.root.after(0, self._log, f"输出视频: {output_video.name}")
            else:
                self.root.after(0, self._log, "输出视频: 未启用")

            _elapsed = time.time() - _detect_start_time
            _elapsed_str = self._format_time(_elapsed)
            duration_sec = (self.end_frame - self.start_frame) / max(1, self.fps)
            duration_str = self._format_time(duration_sec)

            # ── 流量统计结果输出（日志）──
            has_cs = traffic_stats is not None and self.var_traffic_count.get()
            has_auto_flow = auto_flows is not None and self.var_auto_traffic.get()

            if has_cs:
                self.root.after(0, self._log, "-" * 30)
                self.root.after(0, self._log, "【截面流量统计】")
                self.root.after(0, self._log, f"  实际检测用时: {_elapsed_str}")
                by_entry = traffic_stats.get("by_entry", {})
                for entry_name in sorted(by_entry.keys()):
                    for direction in sorted(by_entry[entry_name].keys()):
                        counts = by_entry[entry_name][direction]
                        total = sum(counts.values())
                        self.root.after(0, self._log, f"  {entry_name} → {direction}: {total} 辆")

            if has_auto_flow:
                self.root.after(0, self._log, "-" * 30)
                self.root.after(0, self._log, "【自动流量统计（轨迹聚类）】")
                self.root.after(0, self._log, f"  实际检测用时: {_elapsed_str}")
                self.root.after(0, self._log, f"  聚类结果: {len(auto_flows)} 个方向")
                from src.auto_traffic_counter import get_flow_color
                for flow in auto_flows:
                    dir_label = f"方向{flow.flow_id}" if flow.flow_id >= 0 else "其他/噪声"
                    self.root.after(0, self._log,
                        f"  {dir_label}: {flow.vehicle_count} 辆 "
                        f"(角度 {flow.mean_angle:.1f}°)")
                    for cid, cnt in sorted(flow.by_class.items()):
                        cname = class_labels_zh.get(cid, f"类别{cid}")
                        self.root.after(0, self._log, f"    {cname}: {cnt}")

                # 自动保存轨迹图和轨迹数据（按用户勾选）
                if auto_traffic_counter is not None:
                    stem = self.video_path.stem
                    ts = int(time.time())
                    if self.var_export_trajectory_img.get():
                        img_path = output_dir / f"trajectory_flow_{stem}_{ts}.png"
                        self.root.after(200, self._auto_save_trajectory_image,
                                        auto_traffic_counter, auto_flows, str(img_path))
                    if self.var_export_trajectory_data.get():
                        data_path = output_dir / f"trajectory_data_{stem}_{ts}.csv"
                        self.root.after(250, self._auto_save_trajectory_data,
                                        auto_traffic_counter, auto_flows, str(data_path))

            # ── 统一 Excel 导出（根据勾选组合包含不同统计方法）──
            if self.var_export_traffic.get():
                excel_path = output_dir / f"traffic_{self.video_path.stem}_{int(time.time())}.xlsx"
                export_sources = {}
                if has_cs:
                    export_sources["cs"] = traffic_stats
                if has_auto_flow:
                    export_sources["auto"] = (auto_traffic_counter, auto_flows)
                if export_sources:
                    self.root.after(300, self._export_combined_excel,
                                    export_sources, class_labels_zh,
                                    self.video_path.name, duration_str, _elapsed_str, str(excel_path))
            else:
                self.root.after(0, self._log,
                    f"自动流量统计: auto_flows={auto_flows is not None and len(auto_flows) if auto_flows else 'None'}, "
                    f"var={self.var_auto_traffic.get()}, counter={auto_traffic_counter is not None}")

            self.root.after(0, self._log, "=" * 60)
            self.root.after(0, self._detection_complete, frame_count, str(output_video))

        except Exception as e:
            import traceback
            self.root.after(0, self._log, f"检测出错: {str(e)}")
            self.root.after(0, self._log, traceback.format_exc())
            self.root.after(0, self.stop_detection)

    def _on_detect_progress(self, frame_idx: int, total: int, frame: np.ndarray,
                            detect_count: int, stats: dict, traffic_stats: dict, model_name: str):
        """检测进度回调（后台线程调用，用 after 调度到主线程）
        frame_idx: 已处理帧计数 (0-based，从检测开始累计)
        total: 起终点范围内的总帧数 (end_frame - start_frame)
        """
        self._detect_stats = dict(stats)
        self._traffic_stats = traffic_stats
        # frame_idx 是 predict_video 内部累加的帧计数，已相对于 start_frame
        processed = min(frame_idx + 1, total)  # +1 转为 1-based 计数
        progress = 15 + (processed / max(1, total)) * 80
        self.root.after(0, self._update_progress_bar, progress)
        if detect_count % 10 == 0:
            self.root.after(0, self._log,
                f"  进度: {processed}/{total} ({processed/max(1,total)*100:.0f}%)")
        # 记录当前检测参数供预览显示使用
        self._detect_model_name = model_name
        ui_frame = frame.copy()
        self.root.after(0, self._show_detect_preview, frame_idx, total, ui_frame)

    def _draw_osd(self, frame, frame_id, total, stats, class_labels_zh, model_name):
        """绘制OSD信息（使用中文字体）"""
        y_offset = 30
        font_path = str(resolve_chinese_font_path())
        cv2.rectangle(frame, (5, 10), (320, 180), (255, 255, 255), -1)
        cv2.rectangle(frame, (5, 10), (320, 180), (100, 100, 100), 1)

        put_chinese_text(frame, f"帧: {frame_id}/{total}",
                         (10, y_offset), font_size=14, color=(0, 0, 0), font_path=font_path)
        y_offset += 25

        total_det = sum(stats.values())
        put_chinese_text(frame, f"检测数: {total_det}",
                         (10, y_offset), font_size=14, color=(0, 0, 0), font_path=font_path)
        y_offset += 25

        for cid, name in class_labels_zh.items():
            if stats.get(cid, 0) > 0:
                put_chinese_text(frame, f"{name}: {stats[cid]}",
                                 (10, y_offset), font_size=12, color=(50, 50, 50), font_path=font_path)
                y_offset += 20

        put_chinese_text(frame, f"[Model: {model_name}]",
                         (frame.shape[1] - 180, 10), font_size=12, color=(0, 0, 0), font_path=font_path)

    def _update_detection_frame(self, frame_id: int, annotated_frame: np.ndarray):
        """更新检测帧显示（主线程调用）"""
        self.current_frame = frame_id
        
        # 缩放到显示尺寸
        h, w = annotated_frame.shape[:2]
        ratio = min(self.display_width / w, self.display_height / h)
        if ratio < 1 and ratio > 0:
            display_frame = cv2.resize(annotated_frame, (int(w * ratio), int(h * ratio)))
        else:
            display_frame = annotated_frame

        frame_rgb = cv2.cvtColor(display_frame, cv2.COLOR_BGR2RGB)
        image = Image.fromarray(frame_rgb)
        self.current_photo = ImageTk.PhotoImage(image)
        self.video_label.configure(image=self.current_photo, text="")
        self._update_frame_info()
        self.timeline.set_playhead(frame_id)

    def _show_detect_preview(self, frame_idx: int, total_frames: int, frame: np.ndarray):
        """显示检测预览帧（主线程调用）- 添加 OSD 信息和截面线"""
        annotated = frame.copy()

        # 绘制流量统计截面线（始终绘制，与车型标注开关独立）
        annotated = self._draw_lines_on_frame(annotated)

        # 标注道路边界（由"启用车道边界检测"开关控制）
        if self.var_enable_road_detect.get():
            annotated = self._draw_road_mask(annotated)

        # OSD 信息
        labels = getattr(self, '_detect_class_labels', {})
        model_name = getattr(self, '_detect_model_name', '')
        stats = getattr(self, '_detect_stats', {})
        self._draw_osd(annotated, frame_idx, total_frames, stats, labels, model_name)

        # 缩放到显示尺寸
        h, w = annotated.shape[:2]
        ratio = min(self.display_width / w, self.display_height / h)
        if ratio < 1 and ratio > 0:
            display_frame = cv2.resize(annotated, (int(w * ratio), int(h * ratio)))
        else:
            display_frame = annotated

        frame_rgb = cv2.cvtColor(display_frame, cv2.COLOR_BGR2RGB)
        image = Image.fromarray(frame_rgb)
        self.current_photo = ImageTk.PhotoImage(image)
        self.video_label.configure(image=self.current_photo, text="")
        # 时间轴指针跟随检测位置
        self.timeline.set_playhead(frame_idx)

    def _detection_complete(self, frame_count: int, output_path: str):
        """检测完成"""
        self._log(f"✓ 检测完成! 共处理 {frame_count} 帧")
        if self.var_write_output.get():
            self._log(f"✓ 输出视频已保存: {output_path}")
        else:
            self._log("✓ 仅统计模式，未输出视频")

        # 补充显示流量统计结果（确保在日志栏中可见）
        traffic_stats = getattr(self, '_traffic_stats', None)
        if traffic_stats and self.var_traffic_count.get():
            self._log("-" * 30)
            self._log("【流量统计汇总】")
            by_entry = traffic_stats.get("by_entry", {})
            if by_entry:
                for entry_name in sorted(by_entry.keys()):
                    for direction in sorted(by_entry[entry_name].keys()):
                        counts = by_entry[entry_name][direction]
                        total = sum(counts.values())
                        self._log(f"  {entry_name} → {direction}: {total} 辆")
            else:
                self._log("  暂无流量统计数据")
            self._log("=" * 60)

        self.stop_detection()

    def stop_detection(self):
        """停止检测"""
        self.is_detecting = False

        # 通知推理线程停止
        stop_event = getattr(self, '_stop_event', None)
        if stop_event:
            stop_event.set()
            self._stop_event = None

        self.btn_start.config(state="normal", bg=C.GREEN, fg=C.BG)
        self.btn_stop.config(state="disabled", bg=C.BTN_DISABLED, fg=C.BTN_DISABLED_TEXT)
        self.btn_preview.config(state="normal", bg=C.BORDER_LIGHT, fg=C.TEXT)
        self.btn_select_video.config(state="normal", bg=C.ACCENT, fg=C.BG)
        self._update_stage_label("已停止")

        self._log("检测已停止")

    # ═════════════════════════════════════════
    # 交叉口标注
    # ═════════════════════════════════════════

    # ═════════════════════════════════════════
    # 流量统计线标注与画线
    # ═════════════════════════════════════════

    def _get_video_real_coords(self, event_x: int, event_y: int) -> Optional[Tuple[int, int]]:
        """将视频 Label 上的鼠标坐标转换为原始视频坐标"""
        if not self.video_path:
            return None
        label_w = self.video_label.winfo_width()
        label_h = self.video_label.winfo_height()

        # 获取当前帧原始尺寸
        cap = cv2.VideoCapture(str(self.video_path))
        if not cap.isOpened():
            return None
        w = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
        h = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
        cap.release()

        ratio = min(label_w / w, label_h / h)
        offset_x = (label_w - w * ratio) / 2
        offset_y = (label_h - h * ratio) / 2
        real_x = int((event_x - offset_x) / ratio)
        real_y = int((event_y - offset_y) / ratio)
        if 0 <= real_x < w and 0 <= real_y < h:
            return (real_x, real_y)
        return None

    def _draw_lines_on_frame(self, frame: np.ndarray) -> np.ndarray:
        """在帧上绘制所有已配置的截面线 + 实时通过车辆数（中文）"""
        annotated = frame.copy()
        font_path = str(resolve_chinese_font_path())
        # 获取当前流量统计（检测过程中实时更新）
        traffic_stats = getattr(self, '_traffic_stats', None)
        by_entry = traffic_stats.get("by_entry", {}) if traffic_stats else {}

        for cl in self.count_lines:
            if not cl.is_drawn:
                continue
            # 截面线
            cv2.line(annotated, (cl.x1, cl.y1), (cl.x2, cl.y2), cl.color, 3)
            cv2.circle(annotated, (cl.x1, cl.y1), 5, (0, 255, 0), -1)
            cv2.circle(annotated, (cl.x2, cl.y2), 5, (0, 0, 255), -1)

            # 实时计数
            count = 0
            if by_entry and cl.entry_name in by_entry:
                counts = by_entry[cl.entry_name].get(cl.direction, {})
                count = sum(counts.values()) if isinstance(counts, dict) else 0

            mx, my = (cl.x1 + cl.x2) // 2, (cl.y1 + cl.y2) // 2
            label = f"{cl.entry_name}→{cl.direction}: {count}辆"
            tw, th = get_chinese_text_size(label, font_size=14, font_path=font_path)
            tx, ty = mx - tw // 2, my - 8
            # 半透明背景
            overlay = annotated.copy()
            cv2.rectangle(overlay, (tx - 3, ty - th - 3), (tx + tw + 3, ty + 3), (0, 0, 0), -1)
            cv2.addWeighted(overlay, 0.5, annotated, 0.5, 0, annotated)
            put_chinese_text(annotated, label, (tx, ty), font_size=14, color=(255, 255, 255), font_path=font_path)
        return annotated

    def _show_trajectory_popup(self, counter, flows):
        """在悬浮窗口中绘制聚类后的轨迹（直接在主线程调用）"""
        import traceback

        try:
            self._log(f"_show_trajectory_popup 被调用, flows={len(flows)}, counter={counter is not None}")

            # 生成轨迹图
            self._log("正在生成轨迹图...")
            cv_image = counter.draw_flow_canvas(
                flows,
                canvas_width=500,
                canvas_height=400,
            )
            self._log(f"轨迹图生成完成, shape={cv_image.shape}")

            rgb_image = cv2.cvtColor(cv_image, cv2.COLOR_BGR2RGB)
            img = ImageTk.PhotoImage(Image.fromarray(rgb_image))
            self._log("PhotoImage 创建成功")

            # 创建悬浮窗口
            popup = tk.Toplevel(self.root)
            popup.title("自动流量统计 — 轨迹聚类图")
            popup.configure(bg=C.BG)
            popup.resizable(True, True)

            # 居中定位
            popup.update_idletasks()
            pw, ph = 820, 530
            try:
                root_x = self.root.winfo_x()
                root_y = self.root.winfo_y()
                root_w = self.root.winfo_width()
                root_h = self.root.winfo_height()
                px = root_x + root_w - pw - 30
                py = root_y + max(30, (root_h - ph) // 2)
            except tk.TclError:
                px, py = 100, 100
            popup.geometry(f"{pw}x{ph}+{px}+{py}")

            # 标题
            title_frame = tk.Frame(popup, bg=C.BG)
            title_frame.pack(fill=tk.X, padx=10, pady=(10, 5))
            tk.Label(
                title_frame, text=f"轨迹聚类结果 - {len(flows)} 个方向",
                bg=C.BG, fg=C.TEXT,
                font=("Microsoft YaHei", 12, "bold"),
            ).pack(side=tk.LEFT)

            # 按钮组
            btn_frame = tk.Frame(title_frame, bg=C.BG)
            btn_frame.pack(side=tk.RIGHT)
            tk.Button(
                btn_frame, text="X 关闭", command=popup.destroy,
                bg=C.BORDER_LIGHT, fg=C.TEXT_SUB,
                relief=tk.FLAT, font=("Microsoft YaHei", 9),
                cursor="hand2", padx=10, pady=2,
            ).pack(side=tk.RIGHT, padx=(6, 0))
            tk.Button(
                btn_frame, text="保存聚类结果表",
                command=lambda: self._save_trajectory_table(counter, flows),
                bg=C.ACCENT, fg=C.BG,
                relief=tk.FLAT, font=("Microsoft YaHei", 9),
                cursor="hand2", padx=10, pady=2,
            ).pack(side=tk.RIGHT, padx=(6, 0))
            tk.Button(
                btn_frame, text="保存轨迹聚类图",
                command=lambda: self._save_trajectory_image(cv_image),
                bg=C.GREEN, fg=C.BG,
                relief=tk.FLAT, font=("Microsoft YaHei", 9),
                cursor="hand2", padx=10, pady=2,
            ).pack(side=tk.RIGHT, padx=(6, 0))

            # 图像 + 统计表 左右并排
            main_body = tk.Frame(popup, bg=C.BG)
            main_body.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 5))

            # 左：轨迹图像
            canvas = tk.Canvas(main_body, bg=C.PANEL, highlightthickness=1,
                              highlightbackground=C.BORDER, width=500)
            canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 5))
            canvas.create_image(3, 3, anchor=tk.NW, image=img)
            canvas.image = img

            # 右：分方向车型统计表（Treeview）
            table_frame = tk.Frame(main_body, bg=C.BG)
            table_frame.pack(side=tk.LEFT, fill=tk.BOTH, padx=(5, 0))

            tk.Label(table_frame, text="分方向车型统计",
                    bg=C.BG, fg=C.TEXT,
                    font=("Microsoft YaHei", 9, "bold")).pack(anchor=tk.W, pady=(0, 3))

            tree_scroll = tk.Scrollbar(table_frame, orient=tk.VERTICAL)
            tree_scroll.pack(side=tk.RIGHT, fill=tk.Y)
            tree = ttk.Treeview(table_frame, columns=("count",), show="tree headings",
                               height=14, selectmode="none",
                               yscrollcommand=tree_scroll.set)
            tree_scroll.configure(command=tree.yview)
            tree.heading("#0", text="方向 / 车型")
            tree.heading("count", text="数量")
            tree.column("#0", width=140, anchor=tk.W)
            tree.column("count", width=50, anchor=tk.CENTER)
            tree.pack(fill=tk.BOTH, expand=True)

            from src.auto_traffic_counter import get_flow_color
            class_labels = getattr(self, '_detect_class_labels', {})
            for i, flow in enumerate(flows):
                c = get_flow_color(i)
                color_hex = f"#{c[2]:02x}{c[1]:02x}{c[0]:02x}"
                if flow.flow_id >= 0:
                    flow_label = f"方向{flow.flow_id} ({flow.vehicle_count})"
                else:
                    flow_label = f"其他/噪声 ({flow.vehicle_count})"
                node = tree.insert("", tk.END, text=flow_label,
                                  values=(str(flow.vehicle_count),),
                                  tags=(f"flow_{i}",))
                tree.tag_configure(f"flow_{i}", background=color_hex,
                                  foreground="#ffffff" if sum(c) < 400 else "#000000")
                for cid, cnt in sorted(flow.by_class.items()):
                    cname = class_labels.get(cid, f"类别{cid}")
                    tree.insert(node, tk.END, text=f"  {cname}",
                               values=(str(cnt),))
                tree.item(node, open=True)

            tree.tag_configure("flow_", background=C.PANEL)

            # 图例
            legend = tk.Frame(popup, bg=C.BG)
            legend.pack(fill=tk.X, padx=10, pady=(5, 8))
            for i, flow in enumerate(flows):
                c = get_flow_color(i)
                color_hex = f"#{c[2]:02x}{c[1]:02x}{c[0]:02x}"
                label = counter.get_flow_label(flow)
                ef = tk.Frame(legend, bg=C.BG)
                ef.pack(side=tk.LEFT, padx=(0, 15))
                tk.Label(ef, bg=color_hex, width=3, height=1).pack(side=tk.LEFT, padx=(0, 4))
                tk.Label(ef, text=label, bg=C.BG, fg=C.TEXT_SUB,
                         font=("Microsoft YaHei", 8)).pack(side=tk.LEFT)

            # 确保窗口显示
            popup.lift()
            popup.focus_force()
            self._log(f"轨迹图窗口已弹出: {len(flows)} 个方向")
        except Exception as e:
            self._log(f"轨迹图绘制异常: {e}")
            self._log(traceback.format_exc())

    def _auto_save_trajectory_image(self, counter, flows, output_path: str):
        """自动保存分类轨迹图（不使用对话框）"""
        try:
            cv_image = counter.draw_flow_canvas(flows, canvas_width=800, canvas_height=600)
            output_path = Path(output_path)
            output_path.parent.mkdir(parents=True, exist_ok=True)
            cv2.imwrite(str(output_path), cv_image)
            self._log(f"✓ 分类轨迹图已保存: {output_path}")
        except Exception as e:
            self._log(f"保存轨迹图失败: {e}")

    def _auto_save_trajectory_data(self, counter, flows, output_path: str):
        """自动保存轨迹坐标数据（不使用对话框）"""
        try:
            import csv, math
            fps = max(self.fps, 1.0)
            tracker_to_flow = {}
            for flow in flows:
                for traj in flow.trajectories:
                    tracker_to_flow[traj.tracker_id] = flow.flow_id

            output_path = Path(output_path)
            output_path.parent.mkdir(parents=True, exist_ok=True)
            with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(["车辆ID", "时间(s)", "帧号", "X坐标", "Y坐标",
                                 "瞬时方向角(°)", "车型ID", "车型名称", "聚类方向ID", "聚类方向"])
                for flow in flows:
                    direction_label = counter.get_flow_label(flow) if flow.flow_id >= 0 else "其他/噪声"
                    for traj in flow.trajectories:
                        pts = traj.points
                        for i, pt in enumerate(pts):
                            ts_val = pt.frame_idx / fps
                            if i > 0:
                                prev_pt = pts[i - 1]
                                dx = pt.x - prev_pt.x
                                dy = pt.y - prev_pt.y
                                angle = math.degrees(math.atan2(-dy, dx))
                                if angle < 0:
                                    angle += 360
                                angle_str = f"{angle:.1f}"
                            else:
                                angle_str = ""
                            writer.writerow([
                                traj.tracker_id, f"{ts_val:.2f}", pt.frame_idx,
                                f"{pt.x:.1f}", f"{pt.y:.1f}", angle_str,
                                traj.class_id,
                                self._detect_class_labels.get(traj.class_id, f"类别{traj.class_id}"),
                                flow.flow_id if flow.flow_id >= 0 else -1, direction_label,
                            ])
            total_records = sum(len(f.trajectories) for f in flows)
            self._log(f"✓ 轨迹坐标数据已保存: {output_path} (共 {total_records} 条轨迹)")
        except Exception as e:
            self._log(f"保存轨迹数据失败: {e}")

    def _save_trajectory_image(self, cv_image):
        """保存轨迹聚类图到 PNG 文件"""
        from tkinter import filedialog
        import time as _time
        path = filedialog.asksaveasfilename(
            title="保存轨迹聚类图",
            defaultextension=".png",
            filetypes=[("PNG 图片", "*.png"), ("JPG 图片", "*.jpg")],
            initialfile=f"trajectory_flow_{int(_time.time())}.png",
        )
        if path:
            cv2.imwrite(path, cv_image)
            self._log(f"轨迹聚类图已保存: {path}")

    def _save_trajectory_table(self, counter, flows):
        """导出轨迹聚类结果表（含车辆ID、帧号、坐标、车型、方向）"""
        from tkinter import filedialog
        from src.auto_traffic_counter import get_flow_color
        import time as _time
        import csv

        path = filedialog.asksaveasfilename(
            title="保存轨迹聚类结果表",
            defaultextension=".csv",
            filetypes=[("CSV 表格", "*.csv"), ("所有文件", "*.*")],
            initialfile=f"trajectory_data_{int(_time.time())}.csv",
        )
        if not path:
            return

        try:
            import math
            fps = max(self.fps, 1.0)

            # 构建 tracker_id → flow_id 映射
            tracker_to_flow: dict = {}
            for flow in flows:
                for traj in flow.trajectories:
                    tracker_to_flow[traj.tracker_id] = flow.flow_id

            with open(path, "w", encoding="utf-8-sig", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(["车辆ID", "时间(s)", "帧号", "X坐标", "Y坐标", "瞬时方向角(°)", "车型ID", "车型名称", "聚类方向ID", "聚类方向"])

                for flow in flows:
                    direction_label = counter.get_flow_label(flow) if flow.flow_id >= 0 else "其他/噪声"
                    for traj in flow.trajectories:
                        pts = traj.points
                        for i, pt in enumerate(pts):
                            ts = pt.frame_idx / fps
                            # 计算瞬时方向角：从上一个点到当前点的方向
                            if i > 0:
                                prev_pt = pts[i - 1]
                                dx = pt.x - prev_pt.x
                                dy = pt.y - prev_pt.y
                                angle = math.degrees(math.atan2(-dy, dx))
                                if angle < 0:
                                    angle += 360
                                angle_str = f"{angle:.1f}"
                            else:
                                angle_str = ""
                            writer.writerow([
                                traj.tracker_id,
                                f"{ts:.2f}",
                                pt.frame_idx,
                                f"{pt.x:.1f}",
                                f"{pt.y:.1f}",
                                angle_str,
                                traj.class_id,
                                self._detect_class_labels.get(traj.class_id, f"类别{traj.class_id}"),
                                flow.flow_id if flow.flow_id >= 0 else -1,
                                direction_label,
                            ])

            self._log(f"轨迹聚类结果表已保存: {path} (共 {sum(1 for f in flows for t in f.trajectories for _ in t.points)} 条记录)")
        except Exception as e:
            import traceback
            self._log(f"保存轨迹结果表失败: {e}")
            self._log(traceback.format_exc())

    def _draw_road_mask(self, frame: np.ndarray) -> np.ndarray:
        """绘制道路边界掩码（在帧上叠加绿色道路区域）"""
        if self._road_segmenter is None:
            try:
                from src.road_segmentation import RoadSegmenter
                model_type = self.road_model_var.get()
                self._road_segmenter = RoadSegmenter(model_type=model_type)
            except Exception as e:
                self._log(f"道路分割初始化失败: {e}")
                return frame

        try:
            mask = self._road_segmenter.predict(frame)
            annotated = self._road_segmenter.get_mask_overlay(frame, mask, alpha=0.2)
            return annotated
        except Exception as e:
            # 静默失败，不影响检测
            return frame

    def _on_video_click(self, event):
        """视频画面点击事件 - 用于画线"""
        if not self._line_draw_mode:
            return
        coords = self._get_video_real_coords(event.x, event.y)
        if coords is None:
            return
        rx, ry = coords

        if self._line_draw_mode == "start":
            self._line_start_point = (rx, ry)
            self._line_draw_mode = "end"
            self._log(f"线段起点: ({rx}, {ry})，请点击终点")
            if self.lbl_draw_hint:
                self.lbl_draw_hint.config(text=f"画线: {self._line_temp_entry}→{self._line_temp_direction} | 请点击终点 (ESC取消)")
            # 绑定鼠标移动事件用于橡皮筋
            self.video_label.bind("<Motion>", self._on_video_motion)
        elif self._line_draw_mode == "end":
            sx, sy = self._line_start_point
            # 查找或创建 CountLine
            existing = None
            for cl in self.count_lines:
                if cl.entry_name == self._line_temp_entry and cl.direction == self._line_temp_direction:
                    existing = cl
                    break
            color = get_entry_color(self._entry_index_of(self._line_temp_entry))
            if existing:
                existing.x1, existing.y1 = sx, sy
                existing.x2, existing.y2 = rx, ry
                existing.color = color
                existing.is_drawn = True
                self._log(f"更新截面线: {existing.entry_name}→{existing.direction}")
            else:
                new_line = CrossSectionLine(
                    line_id=f"line_{uuid.uuid4().hex[:8]}",
                    entry_name=self._line_temp_entry,
                    direction=self._line_temp_direction,
                    x1=sx, y1=sy, x2=rx, y2=ry,
                    color=color,
                    is_drawn=True,
                )
                self.count_lines.append(new_line)
                self._log(f"新增截面线: {new_line.entry_name}→{new_line.direction}")
            self._finish_line_drawing()
            self._refresh_tree()
            self._save_traffic_config_silent()

    def _on_video_motion(self, event):
        """鼠标移动事件 - 橡皮筋画线预览"""
        if self._line_draw_mode != "end" or self._line_start_point is None:
            return
        coords = self._get_video_real_coords(event.x, event.y)
        if coords is None:
            return
        # 获取当前帧并绘制预览
        cap = cv2.VideoCapture(str(self.video_path))
        if not cap.isOpened():
            return
        cap.set(cv2.CAP_PROP_POS_FRAMES, self.current_frame)
        ret, frame = cap.read()
        cap.release()
        if not ret:
            return
        # 绘制已有线 + 橡皮筋
        frame = self._draw_lines_on_frame(frame)
        sx, sy = self._line_start_point
        rx, ry = coords
        cv2.line(frame, (sx, sy), (rx, ry), (0, 255, 255), 2)
        cv2.circle(frame, (sx, sy), 6, (0, 255, 0), -1)
        cv2.circle(frame, (rx, ry), 6, (255, 0, 0), -1)
        # 缩放到显示尺寸
        h, w = frame.shape[:2]
        ratio = min(self.display_width / w, self.display_height / h)
        if ratio < 1 and ratio > 0:
            frame = cv2.resize(frame, (int(w * ratio), int(h * ratio)))
        rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        image = Image.fromarray(rgb)
        self.current_photo = ImageTk.PhotoImage(image)
        self.video_label.configure(image=self.current_photo, text="")

    def _finish_line_drawing(self):
        """结束画线模式"""
        self._line_draw_mode = None
        self._line_start_point = None
        self._line_temp_entry = ""
        self._line_temp_direction = ""
        self.video_label.unbind("<Motion>")
        if self.lbl_draw_hint:
            self.lbl_draw_hint.config(text="")
        self._log("画线完成")
        # 有画线后默认勾选启用流量统计
        drawn_lines = [cl for cl in self.count_lines if cl.is_drawn]
        if drawn_lines:
            self.var_traffic_count.set(True)
            self._log("已自动勾选「检测时启用流量统计」")
        # 刷新当前帧显示，确保新画的线立即呈现
        if self.video_path:
            self._seek_absolute(self.current_frame)

    def _cancel_line_drawing(self, _event=None):
        """取消画线"""
        if self._line_draw_mode:
            self._line_draw_mode = None
            self._line_start_point = None
            self.video_label.unbind("<Motion>")
            if self.lbl_draw_hint:
                self.lbl_draw_hint.config(text="")
            self._log("已取消画线")
            # 刷新显示
            self._seek_absolute(self.current_frame)
        return "break"

    # ─── 树形操作 ───

    def _entry_index_of(self, entry_name: str) -> int:
        """获取进口道在树中的索引（用于颜色分配）"""
        entries = []
        for child in self.tree_traffic.get_children():
            text = self.tree_traffic.item(child, "text")
            entries.append(text)
        try:
            return entries.index(entry_name)
        except ValueError:
            return len(entries)

    def _refresh_tree(self):
        """根据 count_lines 刷新树形控件"""
        # 清空树
        for item in self.tree_traffic.get_children():
            self.tree_traffic.delete(item)
        # 按 entry_name 分组
        entries: Dict[str, List[CrossSectionLine]] = {}
        for cl in self.count_lines:
            entries.setdefault(cl.entry_name, []).append(cl)
        for entry_name in sorted(entries.keys()):
            entry_node = self.tree_traffic.insert("", tk.END, text=entry_name, tags=("entry",))
            self.tree_traffic.item(entry_node, tags=("entry",))
            for cl in entries[entry_name]:
                status = "已画线" if cl.is_drawn else "未画线"
                self.tree_traffic.insert(entry_node, tk.END, text=cl.direction,
                                         values=(status,), tags=("direction",))

    def _on_add_entry(self):
        """添加进口道"""
        self._entry_counter += 1
        entry_name = f"进口道{self._entry_counter}"
        # 确保名称唯一
        existing_names = set()
        for child in self.tree_traffic.get_children():
            existing_names.add(self.tree_traffic.item(child, "text"))
        while entry_name in existing_names:
            self._entry_counter += 1
            entry_name = f"进口道{self._entry_counter}"
        node = self.tree_traffic.insert("", tk.END, text=entry_name, tags=("entry",))
        self.tree_traffic.selection_set(node)
        self._rename_tree_node(node)

    def _on_tree_right_click(self, event):
        """树节点右键菜单（已迁移至文件浏览器右键菜单系统）"""
        # 此 handler 不再使用，进口道/方向管理请通过左侧文件浏览器右键操作
        pass

    def _on_tree_double_click(self, event):
        """树节点双击 - 重命名"""
        item = self.tree_traffic.identify_row(event.y)
        if item:
            self._rename_tree_node(item)

    # 注意: _show_entry_menu 和 _show_direction_menu 已迁移至文件浏览器右键菜单系统
    # 旧版定义位于上方 "L3: 进口道右键菜单" 和 "L4: 方向截面右键菜单"

    def _show_empty_area_menu(self, x, y):
        """树形控件空白处/标题行右键菜单（添加进口道或转向）"""
        menu = tk.Menu(self.root, tearoff=0, bg=C.CARD, fg=C.TEXT,
                       font=("Microsoft YaHei", 9))

        # 添加进口道
        menu.add_command(label="添加进口道", command=self._on_add_entry)

        # 如果已有进口道，提供快捷添加转向选项
        entries = self.tree_traffic.get_children()
        if entries:
            menu.add_separator()
            add_dir_menu = tk.Menu(menu, tearoff=0, bg=C.CARD, fg=C.TEXT,
                                   font=("Microsoft YaHei", 9))
            for entry_node in entries:
                entry_name = self.tree_traffic.item(entry_node, "text")
                entry_sub = tk.Menu(add_dir_menu, tearoff=0, bg=C.CARD, fg=C.TEXT,
                                    font=("Microsoft YaHei", 9))
                for direction in ["直行", "左转", "右转", "掉头"]:
                    entry_sub.add_command(
                        label=direction,
                        command=lambda en=entry_node, d=direction: self._add_direction(en, d))
                entry_sub.add_command(
                    label="自定义...",
                    command=lambda en=entry_node: self._add_custom_direction(en))
                add_dir_menu.add_cascade(label=entry_name, menu=entry_sub)
            menu.add_cascade(label="➡️ 添加转向至", menu=add_dir_menu)

        menu.post(x, y)

    def _add_direction(self, entry_node: str, direction: str):
        """添加预设转向方向"""
        entry_name = self.tree_traffic.item(entry_node, "text")
        # 检查是否已存在
        for child in self.tree_traffic.get_children(entry_node):
            if self.tree_traffic.item(child, "text") == direction:
                messagebox.showwarning("警告", f"该进口道已存在「{direction}」方向")
                return
        self.tree_traffic.insert(entry_node, tk.END, text=direction, values=("未画线",), tags=("direction",))
        self._start_line_drawing(entry_name, direction)

    def _add_custom_direction(self, entry_node: str):
        """添加自定义转向方向"""
        entry_name = self.tree_traffic.item(entry_node, "text")
        direction = filedialog.asksaveasfilename(
            title=f"为「{entry_name}」添加自定义方向",
            defaultextension="",
            initialfile="自定义方向",
            filetypes=[("所有文件", "*.*")]
        )
        if not direction:
            return
        # 只取文件名作为方向名
        direction = Path(direction).name
        for child in self.tree_traffic.get_children(entry_node):
            if self.tree_traffic.item(child, "text") == direction:
                messagebox.showwarning("警告", f"该进口道已存在「{direction}」方向")
                return
        self.tree_traffic.insert(entry_node, tk.END, text=direction, values=("未画线",), tags=("direction",))
        self._start_line_drawing(entry_name, direction)

    def _rename_tree_node(self, item: str):
        """重命名树节点（显示 Entry 框）"""
        bbox = self.tree_traffic.bbox(item, column="#0")
        if not bbox:
            return
        x, y, w, h = bbox
        entry = tk.Entry(self.tree_traffic, font=("Microsoft YaHei", 9))
        entry.place(x=x, y=y, width=w, height=h)
        old_text = self.tree_traffic.item(item, "text")
        entry.insert(0, old_text)
        entry.select_range(0, tk.END)
        entry.focus_set()

        def _commit(_event=None):
            new_text = entry.get().strip()
            entry.destroy()
            if new_text and new_text != old_text:
                parent = self.tree_traffic.parent(item)
                if parent == "":
                    # 进口道重命名
                    for cl in self.count_lines:
                        if cl.entry_name == old_text:
                            cl.entry_name = new_text
                    self.tree_traffic.item(item, text=new_text)
                else:
                    # 方向重命名
                    entry_name = self.tree_traffic.item(parent, "text")
                    for cl in self.count_lines:
                        if cl.entry_name == entry_name and cl.direction == old_text:
                            cl.direction = new_text
                    self.tree_traffic.item(item, text=new_text)
                self._save_traffic_config_silent()

        def _cancel(_event=None):
            entry.destroy()

        entry.bind("<Return>", _commit)
        entry.bind("<Escape>", _cancel)
        entry.bind("<FocusOut>", _commit)

    def _delete_tree_item(self, item: str):
        """删除树节点"""
        parent = self.tree_traffic.parent(item)
        if parent == "":
            # 删除进口道及其所有线
            entry_name = self.tree_traffic.item(item, "text")
            self.count_lines = [cl for cl in self.count_lines if cl.entry_name != entry_name]
            self.tree_traffic.delete(item)
            self._log(f"已删除进口道: {entry_name}")
        else:
            # 删除方向
            entry_name = self.tree_traffic.item(parent, "text")
            direction = self.tree_traffic.item(item, "text")
            self.count_lines = [cl for cl in self.count_lines
                                if not (cl.entry_name == entry_name and cl.direction == direction)]
            self.tree_traffic.delete(item)
            self._log(f"已删除转向: {entry_name}→{direction}")
        self._save_traffic_config_silent()

    def _modify_line(self, item: str, parent_item: str):
        """修改画线"""
        entry_name = self.tree_traffic.item(parent_item, "text")
        direction = self.tree_traffic.item(item, "text")
        self._start_line_drawing(entry_name, direction)

    def _delete_line(self, item: str, parent_item: str):
        """删除画线（保留方向节点）"""
        entry_name = self.tree_traffic.item(parent_item, "text")
        direction = self.tree_traffic.item(item, "text")
        for cl in self.count_lines:
            if cl.entry_name == entry_name and cl.direction == direction:
                cl.is_drawn = False
                cl.x1 = cl.y1 = cl.x2 = cl.y2 = 0
        self.tree_traffic.item(item, values=("未画线",))
        self._log(f"已删除画线: {entry_name}→{direction}")
        self._seek_absolute(self.current_frame)
        self._save_traffic_config_silent()

    def _move_direction(self, item: str, old_parent: str, new_parent: str, new_entry_name: str):
        """移动方向到其他进口道"""
        old_entry_name = self.tree_traffic.item(old_parent, "text")
        direction = self.tree_traffic.item(item, "text")
        for cl in self.count_lines:
            if cl.entry_name == old_entry_name and cl.direction == direction:
                cl.entry_name = new_entry_name
        self.tree_traffic.move(item, new_parent, tk.END)
        self._log(f"已将 {direction} 移动到 {new_entry_name}")
        self._save_traffic_config_silent()

    def _start_line_drawing(self, entry_name: str, direction: str):
        """开始画线模式"""
        self._line_draw_mode = "start"
        self._line_start_point = None
        self._line_temp_entry = entry_name
        self._line_temp_direction = direction
        self._log(f"进入画线模式: {entry_name}→{direction}，请点击起点")
        if not self.lbl_draw_hint:
            self.lbl_draw_hint = tk.Label(
                self.video_label.master,
                text="", bg=C.ACCENT, fg=C.BG,
                font=("Microsoft YaHei", 10), padx=10, pady=4
            )
            self.lbl_draw_hint.place(relx=0.5, y=10, anchor="n")
        self.lbl_draw_hint.config(text=f"画线: {entry_name}→{direction} | 请点击起点 (ESC取消)")
        # 绑定 ESC 取消
        self.root.bind("<Escape>", self._cancel_line_drawing)

    # ─── 菜单命令 ───

    def start_traffic_line_mark(self):
        """开始流量统计线标注（菜单命令）"""
        if not self.video_path:
            messagebox.showwarning("警告", "请先选择视频文件")
            return
        if not self.tree_traffic.get_children():
            messagebox.showinfo("提示", "请先点击「设置进口道」添加进口道，然后右键添加转向方向")
            return
        self._log("进入流量统计线标注模式。右键进口道可添加转向方向，点击视频画面进行画线")

    def clear_all_traffic_lines(self):
        """清除所有流量统计线"""
        if messagebox.askyesno("确认", "确定要清除所有进口道和截面线配置吗？"):
            self.count_lines.clear()
            for item in self.tree_traffic.get_children():
                self.tree_traffic.delete(item)
            self._entry_counter = 0
            self._log("已清除所有流量统计线")
            self._seek_absolute(self.current_frame)
            if self.video_path:
                self._save_traffic_config_silent()

    def save_traffic_config(self):
        """手动保存截面线配置"""
        if not self.video_path:
            messagebox.showwarning("警告", "请先选择视频文件")
            return
        path = save_cs_lines_config(self.count_lines, self.video_path, PROJECT_ROOT)
        self._log(f"截面线配置已保存: {path.name}")
        messagebox.showinfo("成功", f"配置已保存:\n{path.name}")

    def _save_traffic_config_silent(self):
        """静默保存截面线配置"""
        if self.video_path:
            try:
                save_cs_lines_config(self.count_lines, self.video_path, PROJECT_ROOT)
            except Exception:
                pass

    def load_traffic_config(self):
        """加载截面线配置"""
        if not self.video_path:
            messagebox.showwarning("警告", "请先选择视频文件")
            return
        lines = load_cs_lines_config(self.video_path, PROJECT_ROOT)
        if lines:
            self.count_lines = lines
            self._refresh_tree()
            self._log(f"已加载截面线配置，共 {len(lines)} 条线")
            self._seek_absolute(self.current_frame)
        else:
            messagebox.showinfo("提示", "未找到该视频对应的截面线配置")

    def _export_combined_excel(self, export_sources: dict, class_labels_zh: dict,
                               video_name: str, duration_str: str,
                               detect_time_str: str, excel_path: str):
        """导出统一流量统计 Excel（主线程调用）
        export_sources 可含:
          'cs'   → traffic_stats (截面线统计)
          'auto' → (auto_traffic_counter, auto_flows) (自动聚类统计)
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

            path = Path(excel_path)
            path.parent.mkdir(parents=True, exist_ok=True)
            wb = Workbook()

            # 通用样式
            header_font = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")
            header_fill = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")
            header_align = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"),
            )

            # ── 截面线统计 Sheet ──
            if "cs" in export_sources:
                traffic_stats = export_sources["cs"]
                ws = wb.active
                ws.title = "截面流量统计"

                ws.merge_cells("A1:D1")
                ws["A1"] = f"Video: {video_name}"; ws["A1"].font = Font(name="Microsoft YaHei", bold=True, size=12)
                ws.merge_cells("A2:D2")
                ws["A2"] = f"Duration: {duration_str}"; ws["A2"].font = Font(name="Microsoft YaHei", size=11)
                if detect_time_str:
                    ws.merge_cells("A3:D3")
                    ws["A3"] = f"Detection Time: {detect_time_str}"
                    ws["A3"].font = Font(name="Microsoft YaHei", size=11, color="4A90D9")

                # 收集所有 class_id
                by_entry = traffic_stats.get("by_entry", {})
                all_cids = set()
                for entry_name in by_entry:
                    for direction in by_entry[entry_name]:
                        for cid in by_entry[entry_name][direction]:
                            all_cids.add(cid)
                sorted_cids = sorted(all_cids)

                offset = 4 if detect_time_str else 3
                hr = offset + 1
                headers = ["进口道", "方向"] + [class_labels_zh.get(c, f"类别{c}") for c in sorted_cids] + ["合计"]
                for ci, h in enumerate(headers, 1):
                    c = ws.cell(row=hr, column=ci, value=h)
                    c.font = header_font; c.fill = header_fill; c.alignment = header_align; c.border = thin_border

                dr = hr + 1
                for entry_name in sorted(by_entry.keys()):
                    for direction in sorted(by_entry[entry_name].keys()):
                        counts = by_entry[entry_name][direction]
                        row_vals = [entry_name, direction]
                        row_total = 0
                        for cid in sorted_cids:
                            v = counts.get(cid, 0)
                            row_vals.append(v)
                            row_total += v
                        row_vals.append(row_total)
                        for ci, val in enumerate(row_vals, 1):
                            c = ws.cell(row=dr, column=ci, value=val)
                            c.alignment = Alignment(horizontal="center", vertical="center")
                            c.border = thin_border
                        dr += 1

                from openpyxl.utils import get_column_letter
                ws.column_dimensions["A"].width = 18
                ws.column_dimensions["B"].width = 12
                for i, cid in enumerate(sorted_cids, 3):
                    name = class_labels_zh.get(cid, f"类别{cid}")
                    ws.column_dimensions[get_column_letter(i)].width = max(10, len(name) * 2 + 2)
                ws.column_dimensions[get_column_letter(len(headers))].width = 10

            # ── 自动聚类统计 Sheet ──
            if "auto" in export_sources:
                auto_traffic_counter, auto_flows = export_sources["auto"]
                from src.auto_traffic_counter import _angle_to_label
                ws = wb.create_sheet("自动聚类统计")

                ws.merge_cells("A1:C1")
                ws["A1"] = f"Video: {video_name}"; ws["A1"].font = Font(name="Microsoft YaHei", bold=True, size=12)
                ws.merge_cells("A2:C2")
                ws["A2"] = f"Duration: {duration_str}"; ws["A2"].font = Font(name="Microsoft YaHei", size=11)
                if detect_time_str:
                    ws.merge_cells("A3:C3")
                    ws["A3"] = f"Detection Time: {detect_time_str}"
                    ws["A3"].font = Font(name="Microsoft YaHei", size=11, color="4A90D9")

                # 收集所有 class_id for auto
                auto_cids = set()
                for flow in auto_flows:
                    auto_cids.update(flow.by_class.keys())
                sorted_auto_cids = sorted(auto_cids)

                offset = 4 if detect_time_str else 3
                hr = offset + 1
                headers = ["方向"] + [class_labels_zh.get(c, f"类别{c}") for c in sorted_auto_cids] + ["合计"]
                for ci, h in enumerate(headers, 1):
                    c = ws.cell(row=hr, column=ci, value=h)
                    c.font = header_font; c.fill = header_fill; c.alignment = header_align; c.border = thin_border

                dr = hr + 1
                for flow in auto_flows:
                    dir_name = _angle_to_label(flow.mean_angle) if flow.flow_id >= 0 else "其他/噪声"
                    label = f"{dir_name} (Flow {flow.flow_id})"
                    row_vals = [label]
                    row_total = 0
                    for cid in sorted_auto_cids:
                        v = flow.by_class.get(cid, 0)
                        row_vals.append(v)
                        row_total += v
                    row_vals.append(row_total)
                    for ci, val in enumerate(row_vals, 1):
                        c = ws.cell(row=dr, column=ci, value=val)
                        c.alignment = Alignment(horizontal="center", vertical="center")
                        c.border = thin_border
                    dr += 1

                from openpyxl.utils import get_column_letter
                ws.column_dimensions["A"].width = 24
                for i, cid in enumerate(sorted_auto_cids, 2):
                    name = class_labels_zh.get(cid, f"类别{cid}")
                    ws.column_dimensions[get_column_letter(i)].width = max(10, len(name) * 2 + 2)
                ws.column_dimensions[get_column_letter(len(headers))].width = 10

            wb.save(str(path))
            sheet_count = len(export_sources)
            self._log(f"✓ 流量统计表已导出: {path.name} (共 {sheet_count} 个工作表)")
            if messagebox.askyesno("完成", f"流量统计结果已保存到:\n{path}\n\n是否打开所在文件夹？"):
                import subprocess
                subprocess.run(["explorer", "/select,", str(path)])
        except Exception as e:
            import traceback
            self._log(f"导出 Excel 失败: {e}")
            self._log(traceback.format_exc())

    # ═════════════════════════════════════════
    # 配置与文件操作
    # ═════════════════════════════════════════

    def load_config_file(self):
        """加载配置文件"""
        path = filedialog.askopenfilename(
            title="加载配置文件",
            filetypes=[("YAML文件", "*.yaml *.yml"), ("所有文件", "*.*")]
        )
        if path:
            try:
                with open(path, 'r', encoding='utf-8') as f:
                    self.config = yaml.safe_load(f) or {}

                self.model_select_var.set(self.config.get('model_select', 3))
                self.conf_scale.set(self.config.get('confidence_threshold', 0.2))
                self.skip_scale.set(self.config.get('detect_skip_interval', 2))
                self.slice_h_var.set(self.config.get('slice_height', 640))
                self.slice_w_var.set(self.config.get('slice_width', 640))
                self.overlap_h_var.set(self.config.get('overlap_height_ratio', 0.1))
                self.overlap_w_var.set(self.config.get('overlap_width_ratio', 0.1))

                self._log(f"已加载配置: {path}")
                messagebox.showinfo("成功", "配置文件加载成功")
            except Exception as e:
                messagebox.showerror("错误", f"加载配置失败: {e}")

    def export_results(self):
        """导出结果"""
        messagebox.showinfo("提示", "检测结果已自动保存到 outputs 目录")

    # ═════════════════════════════════════════
    # UI更新辅助方法
    # ═════════════════════════════════════════

    def _update_frame_info(self):
        """更新帧信息和时间显示"""
        self.frame_info.config(text=f"帧: {self.current_frame} / {self.total_frames}")

        current_time = self.current_frame / max(1, self.fps)
        total_time = self.total_frames / max(1, self.fps)
        self.time_info.config(text=self._format_time(current_time))

        remaining = total_time - current_time
        self.time_remaining.config(text=f"-{self._format_time(remaining)}")

    def _format_time(self, seconds: float) -> str:
        """格式化时间"""
        m, s = divmod(int(seconds), 60)
        h, m = divmod(m, 60)
        return f"{h:02d}:{m:02d}:{s:02d}"

    def _update_progress_bar(self, percent: float):
        """更新进度条"""
        canvas_width = self.progress_bar.winfo_width()
        if canvas_width > 2:
            x = 2 + (canvas_width - 4) * percent / 100
            x = max(2, min(x, canvas_width - 2))
            self.progress_bar.coords(self.progress_rect, 2, 2, x, 10)
            self.progress_percent.config(text=f"{percent:.0f}%")

    def _update_stage_label(self, stage: str):
        """更新阶段标签"""
        self.stage_label.config(text=stage)

    def _log(self, message: str):
        """添加日志"""
        self.text_result.config(state="normal")
        self.text_result.insert(tk.END, message + "\n")
        self.text_result.see(tk.END)
        self.text_result.config(state="disabled")

    def _on_confidence_changed(self, val):
        self.confidence_threshold = val

    def _on_skip_interval_changed(self, val):
        self.skip_interval = val

    # ═════════════════════════════════════════
    # 窗口关闭处理
    # ═════════════════════════════════════════

    def _on_close(self):
        """窗口关闭处理 - 优雅退出所有线程"""
        self.is_detecting = False
        self.is_playing = False
        
        # 停止播放线程
        if self.player_thread:
            self._send_command(PlayerCommand.EXIT)
            self.player_thread.stop()
            self.player_thread.join(timeout=2.0)
        
        # 取消定时器
        if self.play_timer:
            self.root.after_cancel(self.play_timer)
        
        self.root.destroy()

    # ═════════════════════════════════════════
    # 帮助对话框
    # ═════════════════════════════════════════

    def show_help(self):
        help_text = """【使用说明】

1. 打开视频
   - 点击"选择视频文件"或菜单"文件->打开视频"
   - 支持 MP4, AVI, MOV, MKV 格式

2. 视频预览
   - 点击 ▶ 播放按钮或按空格键开始/暂停
   - 使用 ◀◀ ▶▶ 按钮逐帧播放
   - 拖动时间轴可实时预览不同位置

3. 设置时间范围
   - 拖动时间轴上的绿色/红色滑块选择检测起止时间
   - 点击 |◀ ▶| 跳转到选择范围的开始/结束

4. 流量统计设置
   - 点击"设置进口道"添加进口道
   - 右键进口道 → 添加转向方向（直行/左转/右转/掉头/自定义）
   - 添加方向后自动进入画线模式：点击起点 → 点击终点
   - 右键方向节点可：修改画线、删除画线、移动到其他进口道
   - 勾选"检测时启用流量统计"将在检测时自动计数

5. 选择检测模型
   - WALDO YOLOv8m: 平衡模式
   - RT-DETR: 高精度模式
   - YOLOv8l: 大模型模式

6. 开始检测
   - 配置参数后点击"开始检测"
   - 检测完成后若启用流量统计，自动导出 Excel 报表"""
        messagebox.showinfo("使用说明", help_text)

    def show_param_help(self):
        help_text = """【参数说明】

置信度阈值 (0.01-0.95)
- 控制检测框的置信度下限
- 值越高，检测越严格
- 建议: 航拍视频使用0.2-0.3

跳帧间隔 (0-10)
- 0表示每帧都检测
- N表示每N帧跳过一次检测
- 值越大，速度越快

切片尺寸 (SAHI)
- 将大图分割成小块进行检测
- 320-640: 适合1080p以下视频
- 800-1024: 适合4K视频

重叠率
- 切片之间的重叠比例
- 增加重叠可减少边界漏检
- 建议: 0.1-0.2"""
        messagebox.showinfo("参数说明", help_text)

    def show_about(self):
        messagebox.showinfo(
            "关于",
            "无人机航拍视频检测系统 v3.0 (流量统计版)\n\n"
            "基于 YOLOv8/RT-DETR + SAHI 切片检测 + ByteTrack 流量检测\n\n"
            "基于 HRNet-OCR 车道边界检测\n\n"
            "核心特性:\n"
            "  ✓ 线程安全视频播放\n"
            "  ✓ 双缓冲帧缓存\n"
            "  ✓ 防抖时间轴拖动\n"
            "  ✓ 流式检测处理\n"
            "  ✓ 增加IoU + IoM 的非极大值抑制\n"
            "  ✓ 增加道路边界检测\n"
            "  ✓ 进口道/转向方向流量统计\n"
            "  ✓ 基于轨迹聚类的自动流量统计\n"
            "  ✓ 自动导出 Excel 流量报表\n\n"
            "2026 UAV Detection System"
        )


# ═══════════════════════════════════════════════════════════════
# 入口
# ═══════════════════════════════════════════════════════════════

def main():
    root = tk.Tk()
    app = UAVDetectionUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()
