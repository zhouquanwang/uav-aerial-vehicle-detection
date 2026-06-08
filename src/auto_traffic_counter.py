"""
Automatic Traffic Flow Counter
基于轨迹聚类（DBSCAN）的自动流量方向分析

核心思路:
    1. 检测阶段 - 逐帧记录所有被跟踪车辆的轨迹坐标
    2. 后处理阶段 - 提取每条轨迹的起始→终止方向向量
    3. DBSCAN 方向聚类 - 自动发现流量方向簇（无需人工画线）
    4. 每个簇 = 一个流量方向，统计车数 / 车型分布

Author: UAV Detection System
"""

from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Tuple, Optional

import numpy as np

# scipy.signal.find_peaks 用于直方图峰值检测（延迟导入）


# ═══════════════════════════════════════════════════════════════
# 数据类
# ═══════════════════════════════════════════════════════════════

@dataclass
class TrajectoryPoint:
    """单个轨迹点"""
    x: float
    y: float
    frame_idx: int


@dataclass
class VehicleTrajectory:
    """一辆车的完整轨迹"""
    tracker_id: int
    class_id: int
    points: List[TrajectoryPoint] = field(default_factory=list)

    @property
    def start_point(self) -> Tuple[float, float]:
        return (self.points[0].x, self.points[0].y)

    @property
    def end_point(self) -> Tuple[float, float]:
        return (self.points[-1].x, self.points[-1].y)

    @property
    def length(self) -> int:
        return len(self.points)

    @property
    def displacement(self) -> float:
        """起点→终点的欧氏距离"""
        dx = self.end_point[0] - self.start_point[0]
        dy = self.end_point[1] - self.start_point[1]
        return float(np.sqrt(dx * dx + dy * dy))

    @property
    def direction_vector(self) -> Tuple[float, float]:
        """归一化方向向量 (dx, dy), |v|=1"""
        dx = self.end_point[0] - self.start_point[0]
        dy = self.end_point[1] - self.start_point[1]
        length = float(np.sqrt(dx * dx + dy * dy))
        if length < 1e-6:
            return (0.0, 0.0)
        return (dx / length, dy / length)

    @property
    def angle_deg(self) -> float:
        """方向角（度）：0°=右, 90°=上, 180°=左, 270°=下"""
        dx, dy = self.direction_vector
        angle = float(np.degrees(np.arctan2(-dy, dx)))
        if angle < 0:
            angle += 360
        return angle

    def as_xy_array(self) -> np.ndarray:
        """返回 (N, 2) 的坐标数组"""
        return np.array([[p.x, p.y] for p in self.points], dtype=np.float32)


@dataclass
class TrafficFlow:
    """一个自动检测到的流量方向（簇）"""
    flow_id: int
    mean_angle: float                  # 平均方向角（度）
    unit_vector: Tuple[float, float]   # 归一化平均方向向量
    vehicle_count: int                 # 该方向总车辆数
    by_class: Dict[int, int]           # 车型细分: {class_id: count}
    trajectories: List[VehicleTrajectory] = field(default_factory=list)


# ═══════════════════════════════════════════════════════════════
# 辅助函数
# ═══════════════════════════════════════════════════════════════

def _angle_to_label(angle_deg: float) -> str:
    """将角度转为可读方向名"""
    if 337.5 <= angle_deg < 360 or 0 <= angle_deg < 22.5:
        return "→ 东"
    elif 22.5 <= angle_deg < 67.5:
        return "↗ 东北"
    elif 67.5 <= angle_deg < 112.5:
        return "↑ 北"
    elif 112.5 <= angle_deg < 157.5:
        return "↖ 西北"
    elif 157.5 <= angle_deg < 202.5:
        return "← 西"
    elif 202.5 <= angle_deg < 247.5:
        return "↙ 西南"
    elif 247.5 <= angle_deg < 292.5:
        return "↓ 南"
    else:
        return "↘ 东南"


# ═══════════════════════════════════════════════════════════════
# 颜色表（BGR，用于绘图）
# ═══════════════════════════════════════════════════════════════

_FLOW_COLORS = [
    (231, 76, 60),    # 红
    (52, 152, 219),   # 蓝
    (46, 204, 113),   # 绿
    (155, 89, 182),   # 紫
    (241, 196, 15),   # 黄
    (230, 126, 34),   # 橙
    (26, 188, 156),   # 青
    (236, 112, 99),   # 粉
    (255, 99, 71),    # 番茄红
    (64, 224, 208),   # 绿松石
]


def get_flow_color(flow_index: int) -> Tuple[int, int, int]:
    return _FLOW_COLORS[flow_index % len(_FLOW_COLORS)]


# ═══════════════════════════════════════════════════════════════
# 核心计数器
# ═══════════════════════════════════════════════════════════════

class AutoTrafficCounter:
    """
    自动流量统计器（基于轨迹方向聚类）

    使用方式:
        counter = AutoTrafficCounter(min_traj_len=5, min_displacement=30)
        for each frame:
            for each tracked vehicle:
                counter.record(tracker_id, class_id, bbox, frame_idx)

        flows = counter.analyze()
        counter.export_flows_excel(flows, class_labels, output_path)
    """

    def __init__(
        self,
        *,
        min_traj_len: int = 5,
        min_displacement: float = 30.0,
        angle_threshold_deg: float = 10.0,
        min_samples: int = 3,
        bin_width_deg: float = 8.0,
        peak_prominence: int = 3,
        free_angle_threshold_deg: float = 35.0,
    ):
        """
        Args:
            min_traj_len: 最小轨迹点数（少于该值视为噪声）
            min_displacement: 最小位移（像素），低于该值视为静止/短途
            angle_threshold_deg: 方向聚类角度阈值（度），偏差在此范围内的归为一簇。
                注意：若数据密集且覆盖较大角度范围，过大的阈值会导致DBSCAN
                余弦距离产生"桥接效应"（chaining），多个方向被逐步连接为一个簇。
                推荐值：简单路口15-20°，复杂场景10-15°。
            min_samples: DBSCAN 最小样本数（一个方向簇至少需要的轨迹数）
            bin_width_deg: 直方图峰值检测的每格宽度（度），越小越精细
            peak_prominence: 峰值最低显著度（车辆数），低于此值的峰不视为独立方向
            free_angle_threshold_deg: 距离最近峰超过此角度识别为自由人/零散
        """
        self.min_traj_len = min_traj_len
        self.min_displacement = min_displacement
        self.angle_threshold_deg = angle_threshold_deg
        self.min_samples = min_samples
        self.bin_width_deg = bin_width_deg
        self.peak_prominence = peak_prominence
        self.free_angle_threshold_deg = free_angle_threshold_deg

        # tracker_id → VehicleTrajectory
        self._trajectories: Dict[int, VehicleTrajectory] = {}

    # ── 检测阶段 ────────────────────────────────────────────

    def record(
        self,
        tracker_id: int,
        class_id: int,
        bbox_xyxy: np.ndarray,
        frame_idx: int,
    ):
        """
        记录一帧中一个被跟踪车辆的位置。

        Args:
            tracker_id: ByteTrack 跟踪 ID
            class_id: 类别 ID
            bbox_xyxy: [x1, y1, x2, y2]
            frame_idx: 帧序号
        """
        x1, y1, x2, y2 = bbox_xyxy
        # 使用底部中心作为车辆位置（航拍俯视）
        cx = float((x1 + x2) / 2.0)
        cy = float(y2)

        if tracker_id not in self._trajectories:
            self._trajectories[tracker_id] = VehicleTrajectory(
                tracker_id=tracker_id, class_id=class_id
            )
        self._trajectories[tracker_id].points.append(
            TrajectoryPoint(x=cx, y=cy, frame_idx=frame_idx)
        )

    # ── 分析阶段（检测完成后调用）────────────────────────────

    def analyze(
        self,
        image_width: int = 1280,
        image_height: int = 720,
    ) -> List[TrafficFlow]:
        """
        分析所有轨迹 → 方向聚类 → 生成流量报告。

        算法: 直方图峰值检测 + 角度分配（彻底避免DBSCAN桥接效应）
        步骤:
            1. 过滤短轨迹/静止车辆
            2. 方向角度提取（0°=右, 90°=下, 180°=左, 270°=上）
            3. 构建圆形角度直方图 → 高斯平滑 → 峰值检测 → 每个峰=一个主流方向
            4. 每辆车的轨迹分配到最近峰（角度距离 < free_angle_threshold）
            5. 距所有峰都超过阈值的 → "自由人/零散"（flow_id=-1）

        Returns:
            List[TrafficFlow]，按车辆数降序排列
        """
        # Step 1: 过滤
        valid_trajs: List[VehicleTrajectory] = []
        for traj in self._trajectories.values():
            if traj.length >= self.min_traj_len and traj.displacement >= self.min_displacement:
                valid_trajs.append(traj)

        if len(valid_trajs) < self.min_samples:
            return self._fallback_single_flow(valid_trajs)

        # Step 2: 方向特征
        vectors = np.array([traj.direction_vector for traj in valid_trajs], dtype=np.float32)
        angles = np.array([traj.angle_deg for traj in valid_trajs], dtype=np.float32)

        # Step 3: 直方图峰值检测（无桥接效应）
        peak_angles, peak_bin_heights = self._detect_angle_peaks(angles)

        # Step 4: 将每辆车分配到最近的峰（或标记为"自由人"）
        labels, cluster_trajs, cluster_angles, cluster_vectors = self._assign_to_peaks(
            valid_trajs, angles, vectors, peak_angles,
        )

        # Step 5: 聚合 Flow
        flows: List[TrafficFlow] = []
        for label, trajs_list in cluster_trajs.items():
            if not trajs_list:
                continue
            c_angles = np.array(cluster_angles[label])
            c_vectors = np.array(cluster_vectors[label])

            mean_angle = float(np.mean(c_angles))
            mean_vx = float(np.mean(c_vectors[:, 0]))
            mean_vy = float(np.mean(c_vectors[:, 1]))
            norm = float(np.sqrt(mean_vx * mean_vx + mean_vy * mean_vy))
            mean_vector = (mean_vx / norm, mean_vy / norm) if norm > 1e-6 else (0.0, 0.0)

            by_class: Dict[int, int] = defaultdict(int)
            for traj in trajs_list:
                by_class[traj.class_id] += 1

            flows.append(TrafficFlow(
                flow_id=max(label, -1),
                mean_angle=mean_angle,
                unit_vector=mean_vector,
                vehicle_count=len(trajs_list),
                by_class=dict(by_class),
                trajectories=trajs_list,
            ))

        flows.sort(key=lambda f: f.vehicle_count, reverse=True)
        for i, flow in enumerate(flows):
            flow.flow_id = i if flow.flow_id >= 0 else flow.flow_id

        return flows

    # ── 直方图峰值检测 ────────────────────────────────────────

    def _detect_angle_peaks(self, angles: np.ndarray) -> Tuple[np.ndarray, np.ndarray]:
        """
        构建圆形角度直方图 → 高斯平滑 → 峰值检测。

        返回: (peak_angles_deg, peak_heights) 列表
        """
        try:
            from scipy.ndimage import gaussian_filter1d
            from scipy.signal import find_peaks
        except ImportError:
            raise ImportError(
                "AutoTrafficCounter 直方图模式需要 scipy。请安装: pip install scipy"
            )

        n_bins = max(30, int(360.0 / self.bin_width_deg))
        bin_edges = np.linspace(0, 360, n_bins + 1)
        hist, _ = np.histogram(angles, bins=bin_edges)
        bin_centers = (bin_edges[:-1] + bin_edges[1:]) / 2.0

        # 圆形延拓：首尾各复制 1/3 直方图长度，处理后去重
        pad_len = max(1, n_bins // 3)
        hist_padded = np.concatenate([hist[-pad_len:], hist, hist[:pad_len]])
        centers_padded = np.concatenate([
            bin_centers[-pad_len:] - 360,
            bin_centers,
            bin_centers[:pad_len] + 360,
        ])

        # 高斯平滑（轻度平滑，避免磨平真实峰）
        sigma = max(0.8, self.bin_width_deg / 6.0)
        hist_smooth = gaussian_filter1d(hist_padded.astype(float), sigma=sigma)

        # 峰值检测
        prominence = max(1, self.peak_prominence)
        peaks, props = find_peaks(hist_smooth, prominence=prominence, distance=max(2, int(15 / self.bin_width_deg)))

        peak_angles = centers_padded[peaks]
        peak_heights = hist_smooth[peaks]

        # 去重（圆形等效）：保留 prominence 更高的
        visited = set()
        keep_idx = []
        for i in range(len(peak_angles)):
            mod_deg = round(peak_angles[i] % 360, 1)
            if mod_deg not in visited:
                visited.add(mod_deg)
                keep_idx.append(i)

        return peak_angles[keep_idx], peak_heights[keep_idx]

    # ── 分配到峰 ───────────────────────────────────────────────

    def _assign_to_peaks(
        self,
        valid_trajs: List[VehicleTrajectory],
        angles: np.ndarray,
        vectors: np.ndarray,
        peak_angles: np.ndarray,
    ):
        """将每条轨迹分配到最近峰，或标记为自由人(-1)"""
        clusters_trajs = defaultdict(list)
        clusters_angles = defaultdict(list)
        clusters_vectors = defaultdict(list)

        for i, angle in enumerate(angles):
            if len(peak_angles) == 0:
                # 没有检测到峰，全归为自由人
                clusters_trajs[-1].append(valid_trajs[i])
                clusters_angles[-1].append(angle)
                clusters_vectors[-1].append(vectors[i])
                continue

            # 计算到各峰的角度距离（圆形距离）
            dists = np.abs(angle - peak_angles)
            dists = np.minimum(dists, 360 - dists)
            nearest_idx = int(np.argmin(dists))
            nearest_dist = dists[nearest_idx]

            if nearest_dist <= self.free_angle_threshold_deg:
                label = nearest_idx
            else:
                label = -1  # 自由人/零散

            clusters_trajs[label].append(valid_trajs[i])
            clusters_angles[label].append(angle)
            clusters_vectors[label].append(vectors[i])

        return None, clusters_trajs, clusters_angles, clusters_vectors

    # ── 后备 ──────────────────────────────────────────────────

    def _fallback_single_flow(self, trajs: List[VehicleTrajectory]) -> List[TrafficFlow]:
        """轨迹太少时退化为单一方向"""
        if not trajs:
            return []
        by_class: Dict[int, int] = defaultdict(int)
        for traj in trajs:
            by_class[traj.class_id] += 1
        return [TrafficFlow(
            flow_id=0,
            mean_angle=0.0,
            unit_vector=(0.0, 0.0),
            vehicle_count=len(trajs),
            by_class=dict(by_class),
            trajectories=trajs,
        )]

    @property
    def trajectory_count(self) -> int:
        return len(self._trajectories)

    @property
    def trajectories(self) -> List[VehicleTrajectory]:
        return list(self._trajectories.values())

    def get_flow_label(self, flow: TrafficFlow) -> str:
        """生成流量方向的可读标签"""
        direction = _angle_to_label(flow.mean_angle)
        return f"方向{flow.flow_id} {direction} ({flow.vehicle_count}辆)"

    # ── 绘图方法 ────────────────────────────────────────────

    def draw_flow_canvas(
        self,
        flows: List[TrafficFlow],
        canvas_width: int = 400,
        canvas_height: int = 500,
        bg_color: Tuple[int, int, int] = (248, 248, 248),
        margin: int = 20,
    ) -> np.ndarray:
        """
        在独立画布上绘制所有轨迹（不同方向不同颜色）。

        Args:
            flows: analyze() 返回的方向簇列表
            canvas_width, canvas_height: 画布尺寸
            bg_color: 背景色 (BGR)
            margin: 边距

        Returns:
            BGR 图像 (numpy array)
        """
        import cv2

        canvas = np.full((canvas_height, canvas_width, 3), bg_color, dtype=np.uint8)

        # 收集所有轨迹点计算全局边界
        all_points_x = []
        all_points_y = []
        for flow in flows:
            for traj in flow.trajectories:
                pts = traj.as_xy_array()
                all_points_x.extend(pts[:, 0].tolist())
                all_points_y.extend(pts[:, 1].tolist())

        if not all_points_x:
            # 画提示文字
            cv2.putText(
                canvas, "No trajectories", (canvas_width // 2 - 80, canvas_height // 2),
                cv2.FONT_HERSHEY_SIMPLEX, 0.7, (150, 150, 150), 1,
            )
            return canvas

        min_x, max_x = float(min(all_points_x)), float(max(all_points_x))
        min_y, max_y = float(min(all_points_y)), float(max(all_points_y))

        # 防除零
        range_x = max_x - min_x if max_x > min_x else 1
        range_y = max_y - min_y if max_y > min_y else 1

        # 缩放比例（保持宽高比）
        plot_w = canvas_width - 2 * margin
        plot_h = canvas_height - 2 * margin
        scale = min(plot_w / range_x, plot_h / range_y)

        offset_x = margin + (plot_w - range_x * scale) / 2.0
        offset_y = margin + (plot_h - range_y * scale) / 2.0

        def _to_canvas(x: float, y: float) -> Tuple[int, int]:
            """将图像坐标映射到画布坐标"""
            return (
                int(offset_x + (x - min_x) * scale),
                int(offset_y + (y - min_y) * scale),
            )

        # 绘制每个方向的轨迹
        for i, flow in enumerate(flows):
            color = get_flow_color(i)

            for traj in flow.trajectories:
                pts = traj.as_xy_array()
                if len(pts) < 2:
                    # 单点：画小圆
                    cx_c, cy_c = _to_canvas(pts[0, 0], pts[0, 1])
                    cv2.circle(canvas, (cx_c, cy_c), 2, color, -1)
                    continue

                # 多段折线
                canvas_pts = np.array([_to_canvas(p[0], p[1]) for p in pts], dtype=np.int32)
                cv2.polylines(canvas, [canvas_pts], False, color, thickness=2, lineType=cv2.LINE_AA)

                # 在轨迹终点画箭头
                if len(canvas_pts) >= 2:
                    last = canvas_pts[-1]
                    prev = canvas_pts[-2]
                    dx, dy = last[0] - prev[0], last[1] - prev[1]
                    length = float(np.sqrt(dx * dx + dy * dy))
                    if length > 2:
                        dx, dy = dx / length * 8, dy / length * 8
                        cv2.arrowedLine(canvas, prev, last, color, thickness=2, tipLength=0.4)

        # 图例由 tkinter 悬浮窗底部 Label 显示，支持中文

        return canvas

    # ── 导出方法 ────────────────────────────────────────────

    def export_flows_excel(
        self,
        flows: List[TrafficFlow],
        class_labels_zh: Dict[int, str],
        video_name: str,
        duration_str: str,
        detect_time_str: str = "",
        output_path: Path = None,
    ) -> Path:
        """
        导出自动流量统计为 Excel 文件。
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError("Excel 导出需要 openpyxl: pip install openpyxl")

        wb = Workbook()
        ws = wb.active
        ws.title = "自动流量统计"

        header_font = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        # 标题
        ws.merge_cells("A1:D1")
        ws["A1"] = f"Video: {video_name}"
        ws["A1"].font = Font(name="Microsoft YaHei", bold=True, size=12)
        ws["A1"].alignment = Alignment(horizontal="left")

        ws.merge_cells("A2:D2")
        ws["A2"] = f"Duration: {duration_str}"
        ws["A2"].font = Font(name="Microsoft YaHei", size=11)

        row_offset = 2
        if detect_time_str:
            row_offset += 1
            ws.merge_cells("A3:D3")
            ws["A3"] = f"Detection Time: {detect_time_str}"
            ws["A3"].font = Font(name="Microsoft YaHei", size=11, color="4A90D9")

        # 收集所有 class_ids
        all_class_ids: set = set()
        for flow in flows:
            all_class_ids.update(flow.by_class.keys())
        sorted_class_ids = sorted(all_class_ids)

        header_row = row_offset + 1
        headers = ["Flow Direction", "Angle"] + [class_labels_zh.get(cid, f"Class {cid}") for cid in sorted_class_ids] + ["Total"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        data_row = header_row + 1
        for flow in flows:
            dir_name = _angle_to_label(flow.mean_angle) if flow.flow_id >= 0 else "Other/Noise"
            row_values = [f"{dir_name} (Flow {flow.flow_id})", f"{flow.mean_angle:.1f}"]
            row_total = 0
            for cid in sorted_class_ids:
                c = flow.by_class.get(cid, 0)
                row_values.append(c)
                row_total += c
            row_values.append(row_total)

            for col_idx, val in enumerate(row_values, 1):
                cell = ws.cell(row=data_row, column=col_idx, value=val)
                cell.alignment = Alignment(horizontal="center" if col_idx > 2 else "left", vertical="center")
                cell.border = thin_border
            data_row += 1

        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 10
        for i, cid in enumerate(sorted_class_ids, 3):
            name = class_labels_zh.get(cid, f"Class {cid}")
            ws.column_dimensions[get_column_letter(i)].width = max(10, len(name) * 2 + 2)

        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_path))
        return output_path

    def reset(self):
        """清空所有记录，准备新一轮检测"""
        self._trajectories.clear()
